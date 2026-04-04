# =============
# ✅ IMPORTS 
# ============
from flask import Flask, render_template, request, jsonify, session, url_for, redirect, flash, send_from_directory, send_file, make_response
from flask_cors import CORS
import smtplib
import random
import json
import os
import time
from datetime import timedelta, datetime
import uuid
import re
import ssl
import csv
import io
import math
from collections import defaultdict
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from werkzeug.utils import secure_filename
import requests  # type: ignore[import]
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib import colors  # type: ignore[import]
from reportlab.lib.pagesizes import A4  # type: ignore[import]
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer  # type: ignore[import]
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore[import]
from reportlab.lib.enums import TA_CENTER  # type: ignore[import]
from reportlab.lib.units import inch, mm  # type: ignore[import]
from reportlab.pdfbase import pdfmetrics  # type: ignore[import]
from reportlab.pdfbase.ttfonts import TTFont  # type: ignore[import]
from reportlab.pdfbase.pdfmetrics import registerFontFamily  # type: ignore[import]
from dotenv import load_dotenv  # env loader
from sqlalchemy import create_engine

# PDF
from flask import make_response, request

#email
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText

# ===================================================
# IMPORTS
# ===================================================
import string
from collections import defaultdict

import psycopg2
import psycopg2.extras 
from psycopg2 import pool as psycopg2_pool
import atexit
import traceback
import socket
from urllib.parse import urlparse, unquote

DB_POOL = None


def _env_truthy(name, default=True):
    raw = os.getenv(name)
    if raw is None:
        return default
    return str(raw).strip().lower() not in {"0", "false", "no", "off", ""}


def _resolve_ipv4_hostaddr(host, port):
    """Resolve host to IPv4 address so deployments without IPv6 can connect."""
    try:
        infos = socket.getaddrinfo(host, int(port), socket.AF_INET, socket.SOCK_STREAM)
        if infos:
            return infos[0][4][0]
    except Exception:
        return None
    return None


def _supabase_project_ref_from_host(host):
    """
    Extract project ref from direct Supabase host:
    db.<project-ref>.supabase.co -> <project-ref>
    """
    m = re.match(r"^db\.([a-z0-9]+)\.supabase\.co$", (host or "").strip(), flags=re.IGNORECASE)
    return m.group(1) if m else None


def _alternate_supabase_pooler_host(host):
    """
    Toggle pooler host prefix between aws-0 and aws-1 for same region.
    Some projects are provisioned on one prefix only.
    """
    h = (host or "").strip().lower()
    m = re.match(r"^aws-(0|1)-([a-z0-9-]+)\.pooler\.supabase\.com$", h)
    if not m:
        return None
    alt_prefix = "1" if m.group(1) == "0" else "0"
    return f"aws-{alt_prefix}-{m.group(2)}.pooler.supabase.com"


class _PooledConnection:
    """Proxy connection that returns underlying conn to pool on close()."""

    def __init__(self, conn, pool_obj):
        self._conn = conn
        self._pool = pool_obj

    def __getattr__(self, name):
        return getattr(self._conn, name)

    def close(self):
        if self._conn is None:
            return
        try:
            # Keep pooled connections clean if caller forgot to commit/rollback.
            if getattr(self._conn, "status", None) != psycopg2.extensions.STATUS_READY:
                self._conn.rollback()
        except Exception:
            pass
        try:
            self._pool.putconn(self._conn)
        finally:
            self._conn = None


def _db_conn_params():
    # Deployment-friendly DSN support (Vercel/PythonAnywhere/custom envs)
    # Prefer pooler/prisma-style URLs first; direct DB URLs can be IPv6-only on some hosts.
    dsn_env_key = None
    for k in (
        "DB_DSN",
        "POSTGRES_PRISMA_URL",
        "SUPABASE_TRANSACTION_POOLER_URL",
        "SUPABASE_SESSION_POOLER_URL",
        "SUPABASE_POOLER_URL",
        "DATABASE_URL",
        "NEON_DATABASE_URL",
        "POSTGRES_URL",
        "SUPABASE_DB_URL",
    ):
        if os.getenv(k):
            dsn_env_key = k
            break
    dsn = (os.getenv(dsn_env_key) if dsn_env_key else "")
    dsn = (dsn or "").strip()

    db_host = (os.getenv("DB_HOST") or os.getenv("host") or "localhost").strip()
    db_name = (os.getenv("DB_NAME") or os.getenv("dbname") or "POS_Billing").strip()
    db_user = (os.getenv("DB_USER") or os.getenv("user") or "postgres").strip()
    db_pass = (os.getenv("DB_PASSWORD") or os.getenv("password") or "Pos@123").strip()
    db_port = int((os.getenv("DB_PORT") or os.getenv("port") or 5432))
    db_sslmode = (os.getenv("DB_SSLMODE") or ("require" if "supabase.co" in (db_host or "") else "prefer")).strip()
    db_connect_timeout = int(os.getenv("DB_CONNECT_TIMEOUT") or 5)
    force_ipv4 = _env_truthy("DB_FORCE_IPV4", True)
    pooler_host = (os.getenv("SUPABASE_POOLER_HOST") or "").strip()
    pooler_port = int(os.getenv("SUPABASE_POOLER_PORT") or 6543)
    supabase_region = (os.getenv("SUPABASE_REGION") or "").strip()
    deployed_region = (os.getenv("SUPABASE_DEFAULT_POOLER_REGION") or "ap-south-1").strip()
    is_deployed_runtime = bool(
        os.getenv("PYTHONANYWHERE_SITE")
        or os.getenv("PA_SITE")
        or os.getenv("WEBSITE_HOSTNAME")
        or os.getenv("RENDER")
        or os.getenv("VERCEL")
    )
    # Runtime-aware DSN selection so localhost and deployed can coexist.
    if is_deployed_runtime and os.getenv("DEPLOY_DB_DSN"):
        dsn_env_key = "DEPLOY_DB_DSN"
        dsn = (os.getenv("DEPLOY_DB_DSN") or "").strip()
    elif (not is_deployed_runtime) and os.getenv("LOCAL_DB_DSN"):
        dsn_env_key = "LOCAL_DB_DSN"
        dsn = (os.getenv("LOCAL_DB_DSN") or "").strip()
    elif (
        not is_deployed_runtime
        and dsn_env_key == "DB_DSN"
        and _env_truthy("LOCAL_PREFER_HOST_CONFIG", True)
        and not os.getenv("LOCAL_DB_DSN")
    ):
        # Keep local dev on host/user/password config unless LOCAL_DB_DSN is set.
        dsn = ""
        dsn_env_key = None

    # psycopg2 expects postgres:// or postgresql:// (not sqlalchemy dialect suffixes)
    if dsn.startswith("postgresql+psycopg2://"):
        dsn = "postgresql://" + dsn[len("postgresql+psycopg2://") :]
    elif dsn.startswith("postgres+psycopg2://"):
        dsn = "postgres://" + dsn[len("postgres+psycopg2://") :]

    # Auto-convert direct Supabase host to pooler host when region is provided.
    # We avoid guessing regions because wrong poolers cause "Tenant or user not found".
    project_ref = _supabase_project_ref_from_host(db_host)
    effective_region = supabase_region or (deployed_region if is_deployed_runtime else "")
    if not pooler_host and project_ref and effective_region:
        pooler_host = f"aws-0-{effective_region}.pooler.supabase.com"
        pooler_port = int(os.getenv("SUPABASE_POOLER_PORT") or 6543)
        print(f"Using derived Supabase pooler host: {pooler_host}:{pooler_port}")

    if pooler_host:
        db_host = pooler_host
        db_port = pooler_port
        # Supabase pooler expects tenant-suffixed username (e.g. postgres.<project_ref>).
        if project_ref and db_user and "." not in db_user:
            db_user = f"{db_user}.{project_ref}"

    if dsn:
        if dsn_env_key:
            print(f"DB DSN source: {dsn_env_key}")
        # Some providers use postgres://; psycopg2 accepts both, but normalize anyway.
        if dsn.startswith("postgres://"):
            dsn = "postgresql://" + dsn[len("postgres://") :]

        parsed = urlparse(dsn)
        host = (parsed.hostname or db_host or "").strip()
        port = int(parsed.port or db_port)
        database = ((parsed.path or "").lstrip("/") or db_name).strip()
        user = (unquote(parsed.username) if parsed.username is not None else db_user).strip()
        password = (unquote(parsed.password) if parsed.password is not None else db_pass).strip()

        params = {
            "host": host,
            "database": database,
            "user": user,
            "password": password,
            "port": port,
            "sslmode": db_sslmode,
            "connect_timeout": db_connect_timeout,
        }
        explicit_hostaddr = (os.getenv("DB_HOSTADDR") or "").strip()
        if explicit_hostaddr:
            params["hostaddr"] = explicit_hostaddr
            print(f"Using DB_HOSTADDR override: {explicit_hostaddr}")
        elif force_ipv4 and host and host not in {"localhost", "127.0.0.1"}:
            hostaddr = _resolve_ipv4_hostaddr(host, port)
            if hostaddr:
                params["hostaddr"] = hostaddr
                print(f"Using IPv4 DB hostaddr for {host}: {hostaddr}")
        return params

    params = {
        "host": db_host,
        "database": db_name,
        "user": db_user,
        "password": db_pass,
        "port": db_port,
        "sslmode": db_sslmode,
        "connect_timeout": db_connect_timeout,
    }
    explicit_hostaddr = (os.getenv("DB_HOSTADDR") or "").strip()
    if explicit_hostaddr:
        params["hostaddr"] = explicit_hostaddr
        print(f"Using DB_HOSTADDR override: {explicit_hostaddr}")
    elif force_ipv4 and db_host and db_host not in {"localhost", "127.0.0.1"}:
        hostaddr = _resolve_ipv4_hostaddr(db_host, db_port)
        if hostaddr:
            params["hostaddr"] = hostaddr
            print(f"Using IPv4 DB hostaddr for {db_host}: {hostaddr}")
    if (
        "supabase.co" in (params.get("host") or "")
        and (params.get("host") or "").startswith("db.")
        and "hostaddr" not in params
        and not os.getenv("SUPABASE_POOLER_HOST")
        and not os.getenv("SUPABASE_REGION")
        and not os.getenv("DB_DSN")
    ):
        print(
            "Supabase direct host detected without IPv4 override. "
            "Set DB_DSN to Supabase pooler URL (recommended), "
            "or set SUPABASE_POOLER_HOST/SUPABASE_REGION."
        )
    return params


def _init_db_pool():
    """Initialize global postgres pool once (lazy)."""
    global DB_POOL
    if DB_POOL is not None:
        return DB_POOL
    min_conn = int(os.getenv("DB_POOL_MINCONN") or 1)
    max_conn = int(os.getenv("DB_POOL_MAXCONN") or 20)
    DB_POOL = psycopg2_pool.ThreadedConnectionPool(min_conn, max_conn, **_db_conn_params())
    return DB_POOL


def _close_db_pool():
    global DB_POOL
    if DB_POOL is not None:
        try:
            DB_POOL.closeall()
        finally:
            DB_POOL = None


atexit.register(_close_db_pool)


def get_db_connection():
    """Get DB connection from global pool; fallback to direct connect."""
    params = _db_conn_params()
    try:
        p = _init_db_pool()
        conn = p.getconn()
        return _PooledConnection(conn, p)
    except Exception as e:
        print(f"DB pool get failed, falling back to direct connect: {e}")
        # Fallback if pool init/get fails for any reason.
        try:
            return psycopg2.connect(**params)
        except Exception as e2:
            # Supabase pooler can return "Tenant or user not found" when aws-0/aws-1 host
            # prefix is mismatched for a project. Try the alternate host once.
            msg = str(e2)
            host = str(params.get("host") or "")
            alt_host = _alternate_supabase_pooler_host(host)
            if alt_host and "Tenant or user not found" in msg:
                alt_params = dict(params)
                alt_params["host"] = alt_host
                alt_params.pop("hostaddr", None)  # recalculate DNS/IP for alternate host
                try:
                    print(f"Retrying DB connect with alternate pooler host: {alt_host}")
                    return psycopg2.connect(**alt_params)
                except Exception:
                    pass
            print(f"Direct DB connect failed: {e2}")
            print(traceback.format_exc())
            raise

# Base directory for building absolute paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

load_dotenv()  # Load variables from .env if present

# Pre-warm DB pool on startup to reduce first-login latency.
try:
    _init_db_pool()
except Exception as e:
    print(f"DB pool warmup skipped: {e}")

# =========================================
# ✅ SQLALCHEMY ENGINE (optional)
# - Used for Supabase/Postgres connectivity
# - Driven by lowercase keys in .env: user/password/host/port/dbname
# - Exposes `engine` for scripts like `main.py`
# =========================================
SQLALCHEMY_ENGINE = None
engine = None

_USER = os.getenv("user")
_PASSWORD = os.getenv("password")
_HOST = os.getenv("host")
_PORT = os.getenv("port")
_DBNAME = os.getenv("dbname")

if _USER and _PASSWORD and _HOST and _PORT and _DBNAME and _PASSWORD != "[YOUR-PASSWORD]":
    _DATABASE_URL = (
        f"postgresql+psycopg2://{_USER}:{_PASSWORD}@{_HOST}:{_PORT}/{_DBNAME}?sslmode=require"
    )
    try:
        SQLALCHEMY_ENGINE = create_engine(_DATABASE_URL)
        engine = SQLALCHEMY_ENGINE
    except Exception as e:
        # Keep app importable even if DB env is not configured yet.
        print(f"Failed to create SQLAlchemy engine: {e}")

def get_departments_from_db():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, code, name, branch, description 
        FROM departments
        ORDER BY id DESC
    """)

    rows = cur.fetchall()

    result = []
    for r in rows:
        result.append({
            "id": r[0],
            "code": r[1],
            "name": r[2],
            "branch": r[3],
            "description": r[4]
        })

    cur.close()
    conn.close()
    return result

def get_roles_from_db():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT r.id, d.name, r.branch, r.role_name, r.description, r.permissions
        FROM roles r
        JOIN departments d ON r.department_name = d.name
        ORDER BY r.id DESC
    """)

    rows = cur.fetchall()

    roles = []
    for r in rows:
        roles.append({
            "id": r[0],
            "department": r[1],
            "branch": r[2],
            "role": r[3],
            "description": r[4],
            "permissions": r[5]
        })

    cur.close()
    conn.close()
    return roles

def get_customers_from_db():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
    SELECT 
        customer_id,      -- 0
        name,             -- 1
        company,          -- 2
        customer_type,    -- 3
        status,           -- 4
        email,            -- 5
        phone,            -- 6  
        credit_limit,     -- 7
        city,             -- 8
        sales_rep,        -- 9
        company_type,     -- 10
        billing_address,  -- 11
        shipping_address  -- 12
    FROM customers
""")

    rows = cur.fetchall()

    customers = []
    for r in rows:
        customers.append({
    "customer_id": r[0],
    "name": r[1],
    "company": r[2],
    "customer_type": r[3],
    "status": r[4],
    "email": r[5],
    "phone": r[6],             
    "credit_limit": float(r[7] or 0),
    "city": r[8],
    "sales_rep": r[9],
    "company_type": r[10],
    "billing_address": r[11],
    "shipping_address": r[12]
})

    cur.close()
    conn.close()
    return customers

# =========================================
# PDF Font Setup (Supports ₹ Indian Rupee Symbol)
# =========================================
FONT_DIR = os.path.join(BASE_DIR, "static", "fonts")

pdfmetrics.registerFont(
    TTFont("DejaVuSans", os.path.join(FONT_DIR, "DejaVuSans.ttf"))
)

pdfmetrics.registerFont(
    TTFont("DejaVuSans-Bold", os.path.join(FONT_DIR, "DejaVuSans-Bold.ttf"))
)

registerFontFamily(
    "DejaVuSans",
    normal="DejaVuSans",
    bold="DejaVuSans-Bold",
    italic="DejaVuSans",
    boldItalic="DejaVuSans-Bold"
)

# =========================================
# ✅ EMAIL SENDER (SMTP / UNIVERSAL)
# =========================================
def send_email_universal(to_email, subject, body, from_email, password, smtp_server=None, port=None):
    """Send email using configured SMTP server."""
    smtp_server = smtp_server or os.getenv("SMTP_SERVER", "smtp.gmail.com")
    port = int(port or os.getenv("SMTP_PORT", "587"))
    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = to_email

    context = ssl.create_default_context()

    try:
        with smtplib.SMTP(smtp_server, port) as server:
            server.starttls(context=context)
            server.login(from_email, password)
            server.sendmail(from_email, to_email, msg.as_string())
        return True

    except Exception as e:
        print("❌ Email send error:", e)
        return False


# =========================================
# ✅ FLASK APP SETUP
# =========================================
app = Flask(__name__)
CORS(app)


# =========================================
# ✅ SESSION SETTINGS & GLOBAL CONSTANTS
# =========================================
app.secret_key = os.getenv("SECRET_KEY", "dev-secret-key")
app.permanent_session_lifetime = timedelta(days=7)     # for Remember Me
# Inactivity timeout for normal sessions when "Remember Me" is not checked (in seconds)
# 15 minutes = 900 seconds — after this, user is logged out and redirected to login
INACTIVITY_TIMEOUT = 900


# =========================================
# ✅ FILE PATH CONSTANTS (JSON FILES)
# =========================================
USER_FILE = os.path.join(app.root_path, "users.json")
ROLE_FILE = os.path.join(app.root_path, "roles.json")
FAILED_ATTEMPTS_FILE = os.path.join(app.root_path, "failed_attempts.json")
OTP_FILE = os.path.join(app.root_path, "email_otps.json")
DEPARTMENT_FILE = os.path.join(app.root_path, "departments.json")
UPLOAD_FOLDER = os.path.join(app.root_path, "static", "uploads")
ATTACHMENTS_FOLDER = os.path.join(app.root_path, "attachments")
PRODUCT_FILE = os.path.join(app.root_path, "product.json")
CATEGORY_FILE = os.path.join(app.root_path, "product_categories.json")
TAX_CODE_FILE = os.path.join(app.root_path, "product_tax_codes.json")
UOM_FILE = os.path.join(app.root_path, "product_uoms.json")
WAREHOUSE_FILE = os.path.join(app.root_path, "product_warehouses.json")
SIZE_FILE = os.path.join(app.root_path, "product_sizes.json")
COLOR_FILE = os.path.join(app.root_path, "product_colors.json")
SUPPLIER_FILE = os.path.join(app.root_path, "product_suppliers.json")
# CUSTOMER_FILE = os.path.join(app.root_path, "customer.json")
ENQUIRY_FILE = os.path.join(app.root_path, "new-enquiry.json")
ENQUIRY_PRODUCT_FILE = "data/enquiry_products.json"
QUOTATION_FILE = os.path.join(app.root_path, "quotation.json")
COMMENTS_FILE = os.path.join(app.root_path, "comments.json")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(ATTACHMENTS_FOLDER, exist_ok=True)
BILLS_FILE = os.path.join(app.root_path, "bills.json")
ATTACHMENTS_FOLDER = os.path.join(app.root_path, "attachments")
os.makedirs(ATTACHMENTS_FOLDER, exist_ok=True)
HOLD_FILE = os.path.join(app.root_path, "Hold-Billing.json")

DELIVERY_NOTE_FILE = os.path.join(app.root_path, "deliverynotes.json")



# =========================================
# ✅ EMAIL CONFIG (from environment)
# =========================================
EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS", "")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD", "")

SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SENDER_EMAIL = os.getenv("SENDER_EMAIL", EMAIL_ADDRESS)
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD", EMAIL_PASSWORD)

OTP_EXPIRY_MINUTES = int(os.getenv("OTP_EXPIRY_MINUTES", "1"))

# =========================================
# ✅ FORGOT PASSWORD + LOCKOUT CONFIG
# =========================================
_raw_base_url = os.getenv("APP_BASE_URL", "http://127.0.0.1:5000")
# Support multiple base URLs in APP_BASE_URL (comma-separated); use the first as primary
BASE_URL = _raw_base_url.split(",")[0].strip() if _raw_base_url else "http://127.0.0.1:5000"
RESET_SEND_COUNT = {}
MAX_RESET_SENDS = 5
LOCKOUT_THRESHOLD = 5
LOCKOUT_DURATION = 120  # seconds
RESET_TOKENS = {}
RESET_LOCK = {}
RESET_TOKEN_EXPIRY = 600

# =========================================
# ✅ SIGNUP OTP RATE LIMITING
# =========================================
# BUG_008: Too many OTP resend attempts should return HTTP 429 instead of 200.
# We keep a simple in‑memory counter of recent OTP sends per email.
OTP_SEND_COUNT = {}  # { email: [timestamps...] }
MAX_OTP_SENDS = 5    # max OTPs within the window
OTP_WINDOW_SECONDS = 5 * 60  # 5 minutes


# =========================================
# ✅ REGEX VALIDATION RULES
# =========================================
EMAIL_REGEX = re.compile(
    r"^[A-Za-z0-9._%+-]{3,40}@(gmail\.com|yahoo\.com|yahoo\.co\.in|outlook\.com|hotmail\.com|thestackly\.com|stackly\.in)$",
    re.IGNORECASE
)

MAX_EMAIL_LENGTH = 40
PHONE_REGEX = re.compile(r"^[0-9]{10}$")
NAME_REGEX = re.compile(r"^[A-Za-z\s]{3,20}$")


# =========================================
# ✅ QUOTATION / EMAIL OTP + RATE LIMIT CONFIG
# =========================================
OTP_EXPIRY_MINUTES = 1
MAX_ATTACHMENTS_PER_QUOTATION = 5
MAX_FILE_SIZE_MB = 10
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024
ALLOWED_EXTENSIONS = {"pdf", "doc", "docx", "xls", "xlsx", "jpg", "jpeg", "png"}
RATE_LIMIT_CONFIG = {
    "max_otp_attempts": 5,                # Max OTP attempts
    "otp_cooldown_minutes": 30,          # Cooldown after max attempts
    "max_emails_per_quotation": 3,       # Max emails per quotation
    "max_emails_per_recipient": 2,       # Max emails to same person
    "max_daily_emails_per_customer": 5,  # Max emails per day
    "min_time_between_emails_minutes": 5,  # Wait time between sends
    "requires_approval_after": 2,        # Require approval after 2 emails
}

# In-memory storage for rate limiting (quotation emails)
email_attempts = defaultdict(list)
otp_attempts = defaultdict(list)
otp_resend_attempts = defaultdict(list)
otp_blocked = defaultdict(dict)

# Shared email message containers for quotation emails
msg = MIMEMultipart("mixed")  # For both HTML and attachment
msg_alternative = MIMEMultipart("alternative")  # For HTML + plain text
msg.attach(msg_alternative)

# Simple in-memory OTP storage for quotation actions
otp_storage = {}


# =========================================
# ✅ CONTENT NEGOTIATION (HTML vs JSON for Postman)
# =========================================
def wants_json():
    """Check if client wants JSON response (API/Postman).
    True when: Accept: application/json, ?format=json, request.is_json,
    or Content-Type: application/json (many clients send this on GET in Postman)."""
    accept = (request.headers.get("Accept") or "").lower()
    if "application/json" in accept:
        return True
    if request.args.get("format") == "json":
        return True
    if request.is_json:
        return True
    ct = (request.headers.get("Content-Type") or "").lower()
    if "application/json" in ct:
        return True
    return False


# =========================================
# ✅ GLOBAL BEFORE_REQUEST
#   - AUTO SESSION TIMEOUT CHECK
# =========================================
@app.before_request
def auto_session_timeout():
    allowed_paths = [
        "/",                     # root
        "/login",
        "/signup",
        "/forgot-password",
        "/check-email",
        "/check-your-mail",
        "/reset-password",
        "/send_otp",
        "/verify_otp",
        "/send-reset-link",
        "/static",
    ]

    # allow static files
    if request.path.startswith("/static/"):
        return

    # allow the routes above
    if request.path in allowed_paths:
        return

    # for all other pages → check session timeout
    # Skip session check for some public pages above.
    # Special handling for JSON/API clients: return JSON 401 instead of redirecting to /login.
    is_api = request.path.startswith("/api/")

    if wants_json():
        # JSON clients (e.g. Postman) should get a JSON error, not an HTML redirect.
        if not check_session_timeout():
            return jsonify({"success": False, "message": "session_expired"}), 401
    else:
        # Normal browser HTML navigation: redirect to login on timeout for non-API routes.
        if not is_api and not check_session_timeout():
            return redirect(url_for("login", message="session_expired"))

    # =========================================
    # ✅ INNER LINK RESTRICTION
    #   - Prevent direct URL / new-tab access to
    #     internal pages without coming from
    #     another page inside the app.
    #   - Skip APIs and AJAX JSON endpoints.
    # =========================================
    # Only apply for normal HTML GET requests
    if request.method == "GET" and not request.path.startswith("/api/"):
        ref = request.referrer or ""
        # host_url example: "http://127.0.0.1:5000/"
        base = (request.host_url or "").rstrip("/")

        # If there is no referrer or it is from outside this app,
        # block direct navigation and send user to login.
        if (not ref) or (base and not ref.startswith(base)):
            return redirect(url_for("login", message="invalid_navigation"))


# =========================================
# ✅ SESSION DEFAULT HELPERS
# =========================================
def ensure_role():
    if "user" in session and "role" not in session:
        session["role"] = "user"


# =========================================
# ✅ JSON HELPERS — Users storage shape
# =========================================
# Persisted users.json records: full branch-user fields + password, never "id".
# DEFAULT_BRANCH_USER_PASSWORD applies when admin-created users have no password yet.
_USER_PHONE_COUNTRY_PREFIXES = tuple(
    sorted(
        ["+91", "+971", "+974", "+966", "+94", "+880", "+977", "+1", "+44", "+61"],
        key=len,
        reverse=True,
    )
)
DEFAULT_BRANCH_USER_PASSWORD = os.getenv("DEFAULT_BRANCH_USER_PASSWORD", "Stackly@123")


def _infer_country_and_contact_from_phone(phone: str):
    """Split +… phone into (country_code, contact_number) when prefix matches a known code."""
    phone = (phone or "").strip()
    if not phone:
        return "", ""
    if not phone.startswith("+"):
        return "", re.sub(r"\D", "", phone)
    digits_only = "".join(c for c in phone[1:] if c.isdigit())
    for prefix in _USER_PHONE_COUNTRY_PREFIXES:
        p_digits = prefix[1:]
        if digits_only.startswith(p_digits):
            rest = digits_only[len(p_digits) :]
            return prefix, rest
    return "", digits_only


def normalize_user_record_for_storage(u: dict) -> dict:
    """Normalize one user dict for users.json: drop id, ensure password + full field set."""
    if not isinstance(u, dict):
        return {}
    out = {k: v for k, v in u.items() if k != "id"}
    pwd = (out.get("password") or "").strip() if out.get("password") is not None else ""
    if not pwd:
        out["password"] = DEFAULT_BRANCH_USER_PASSWORD
    else:
        out["password"] = pwd
    name = (out.get("name") or "").strip()
    fn = (out.get("first_name") or "").strip()
    ln = (out.get("last_name") or "").strip()
    if not fn and name:
        parts = name.split(None, 1)
        fn = parts[0]
        ln = parts[1] if len(parts) > 1 else (ln or "")
    out["first_name"] = fn
    out["last_name"] = ln or ""
    out["name"] = name or f"{fn} {ln}".strip()
    cc = (out.get("country_code") or "").strip()
    cn = (out.get("contact_number") or "").strip()
    phone = (out.get("phone") or "").strip()
    if cc and cn:
        out["country_code"] = cc
        out["contact_number"] = cn
        out["phone"] = phone or f"{cc}{cn}"
    elif phone:
        icc, icn = _infer_country_and_contact_from_phone(phone)
        out["phone"] = phone
        out["country_code"] = icc
        out["contact_number"] = icn
    else:
        out["phone"] = phone
        out["country_code"] = cc
        out["contact_number"] = cn
    out["email"] = (out.get("email") or "").strip()
    out["role"] = (out.get("role") or "").strip() or "User"
    for key in ("branch", "department", "reporting_to", "available_branches", "employee_id"):
        val = out.get(key)
        out[key] = (val or "").strip() if val is not None else ""
    return out


def user_public_dict(u: dict) -> dict:
    """User object safe for JSON responses (no password or id)."""
    if not isinstance(u, dict):
        return {}
    return {k: v for k, v in u.items() if k not in ("password", "id")}


def load_users():
    """Read users from users.json as a list of dicts."""
    if not os.path.exists(USER_FILE):
        return []
    with open(USER_FILE, "r", encoding="utf-8") as f:
        try:
            data = json.load(f)
            if isinstance(data, list):
                return data
            if isinstance(data, dict):
                return list(data.values())
            return []
        except json.JSONDecodeError:
            return []




def save_users(data):
    """Write users back to users.json as list (no id; always password + full keys)."""
    if isinstance(data, dict):
        data = list(data.values())
    normalized = []
    for item in data:
        if isinstance(item, dict):
            norm = normalize_user_record_for_storage(item)
            item.clear()
            item.update(norm)
            normalized.append(item)
    with open(USER_FILE, "w", encoding="utf-8") as f:
        json.dump(normalized, f, indent=2, ensure_ascii=False)


def load_failed_attempts():
    if os.path.exists(FAILED_ATTEMPTS_FILE):
        try:
            with open(FAILED_ATTEMPTS_FILE, "r") as f:
                return json.load(f)
        except json.JSONDecodeError:
            return {}
    return {}


def save_failed_attempts(data):
    with open(FAILED_ATTEMPTS_FILE, "w") as f:
        json.dump(data, f)


# =========================================
# ✅ EMAIL WRAPPER HELPER
# =========================================
def send_email(to_email, subject, body):
    """Wrapper around universal email sender (supports Gmail, Outlook, Yahoo, etc.)."""
    try:
        return send_email_universal(
            to_email=to_email,
            subject=subject,
            body=body,
            from_email=EMAIL_ADDRESS,
            password=EMAIL_PASSWORD,
        )
    except Exception as e:
        print("Email send error:", e)
        return False


# =========================================
# ✅ OTP HELPERS
# =========================================
def load_otps():
    """Return dict: { email: {otp, verified, timestamp} }"""
    try:
        with open(OTP_FILE, "r") as f:
            data = json.load(f)
            if isinstance(data, dict):
                return data
            return {}
    except FileNotFoundError:
        return {}


def save_otps(otps: dict):
    with open(OTP_FILE, "w") as f:
        json.dump(otps, f, indent=2)


def generate_otp():
    return str(random.randint(100000, 999999))



def save_otp_in_db(email, otp):
    """Store/overwrite OTP for this email, mark as not verified yet."""
    email = (email or "").strip().lower()
    otps = load_otps()
    otps[email] = {
        "otp": otp,
        "verified": False,
        "timestamp": time.time(),
    }
    save_otps(otps)


def verify_otp_in_db(email, otp, expiry_seconds=300):
    """
    Check OTP for email with expiry.
    If valid, mark as verified and return True.
    """
    email = (email or "").strip().lower()
    otps = load_otps()
    entry = otps.get(email)

    if not entry:
        return False

    now = time.time()
    if now - entry.get("timestamp", 0) > expiry_seconds:
        return False

    if entry.get("otp") != otp:
        return False

    entry["verified"] = True
    otps[email] = entry
    save_otps(otps)
    return True


def is_email_otp_verified(email: str) -> bool:
    """Used during signup to ensure email's OTP was verified."""
    email = (email or "").strip().lower()
    otps = load_otps()
    entry = otps.get(email)

    if not entry:
        return False

    max_age = 10 * 60
    if time.time() - entry.get("timestamp", 0) > max_age:
        return False

    return bool(entry.get("verified"))


def send_otp_email(to_email, otp):
    """Send the signup OTP using the universal email helper."""
    subject = "Your OTP Verification Code - Stackly POS"
    body = (
        f"Hi,\n\n"
        f"Your OTP for Stackly POS signup is: {otp}\n"
        f"It is valid for 5 minutes.\n\n"
        f"If you did not request this, you can ignore this email.\n\n"
        f"- Stackly Team"
    )
    send_email(to_email, subject, body)


# =========================================
# ✅ DEPARTMENT HELPERS
# =========================================
def normalize_department_for_storage(d):
    """Departments.json stores code, name, branch, description — never id."""
    if not isinstance(d, dict):
        return {}
    return {k: v for k, v in d.items() if k != "id"}


def department_for_api(d):
    """Department dict safe for JSON (no id)."""
    if not isinstance(d, dict):
        return {}
    return {k: v for k, v in d.items() if k != "id"}


def _dept_code_key(d):
    return (d.get("code") or "").strip().lower()


def find_department_by_code(departments, code_ref):
    """Find a department by code (case-insensitive)."""
    if not code_ref or not isinstance(departments, list):
        return None
    cref = str(code_ref).strip().lower()
    if not cref:
        return None
    for d in departments:
        if isinstance(d, dict) and _dept_code_key(d) == cref:
            return d
    return None


def load_departments():
    if not os.path.exists(DEPARTMENT_FILE):
        return []
    try:
        with open(DEPARTMENT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            if not isinstance(data, list):
                return []
            return [normalize_department_for_storage(d) if isinstance(d, dict) else d for d in data]
    except Exception:
        return []


def save_departments(departments):
    """Persist departments without id field."""
    if not isinstance(departments, list):
        departments = []
    normalized = []
    for d in departments:
        if isinstance(d, dict):
            norm = normalize_department_for_storage(d)
            d.clear()
            d.update(norm)
            normalized.append(d)
    with open(DEPARTMENT_FILE, "w", encoding="utf-8") as f:
        json.dump(normalized, f, indent=2, ensure_ascii=False)


def load_products():
    
    if not os.path.exists(PRODUCT_FILE):
        return []

    try:
        with open(PRODUCT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError:
        return []

    if isinstance(data, dict):
        data = [data]

    products = []
    for p in data:
        if not isinstance(p, dict):
            continue

        pid = p.get("product_id")
        if pid is None:
            continue
        # ✅ always keep IDs as string
        p["product_id"] = str(pid)

        products.append(p)

    return products


def save_products(products):
    with open(PRODUCT_FILE, "w", encoding="utf-8") as f:
        json.dump(products, f, indent=2)


def generate_product_id():
    """Generate a new product_id in format 'P101', 'P102', ..."""
    products = load_products()
    if not products:
        return "P101"

    max_num = 0
    for p in products:
        pid = str(p.get("product_id", "")).strip()
        # Extract numeric part from formats like "P101", "101", "P-101", etc.
        match = re.search(r"(\d+)$", pid)
        if match:
            max_num = max(max_num, int(match.group(1)))

    return f"P{max_num + 1}"


@app.route('/api/products/new-id', methods=['GET'])
def get_new_product_id():
    """Returns the next auto-generated product ID"""
    product_id = generate_product_id()
    return jsonify({"productId": product_id})


# =========================================
# ✅ PRODUCT CATEGORY HELPERS
# =========================================
def load_product_categories():
    """
    Load saved product categories from JSON.
    Structure: [ { "product_type": "Electronics", "name": "Headphones" }, ... ]
    """
    if not os.path.exists(CATEGORY_FILE):
        return []
    try:
        with open(CATEGORY_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            cleaned = []
            for item in data:
                if not isinstance(item, dict):
                    continue
                name = (item.get("name") or "").strip()
                ptype = (item.get("product_type") or "").strip()
                if not name:
                    continue
                cleaned.append({"product_type": ptype, "name": name})
            return cleaned
        return []
    except json.JSONDecodeError:
        return []


def save_product_categories(categories):
    """Persist product categories list to JSON."""
    if not isinstance(categories, list):
        categories = []
    with open(CATEGORY_FILE, "w", encoding="utf-8") as f:
        json.dump(categories, f, indent=2, ensure_ascii=False)


# =========================================
# ✅ PRODUCT MASTER DROPDOWN HELPERS
#    (Tax codes, UOM, Warehouse, Size, Color, Supplier)
# =========================================
def _load_simple_list(path, cleaner):
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, list):
            return []
        cleaned = []
        for item in data:
            if not isinstance(item, dict):
                continue
            cleaned_item = cleaner(item)
            if cleaned_item:
                cleaned.append(cleaned_item)
        return cleaned
    except json.JSONDecodeError:
        return []


def _save_simple_list(path, items):
    if not isinstance(items, list):
        items = []
    with open(path, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2, ensure_ascii=False)


def load_tax_codes():
    def cleaner(item):
        name = (item.get("code") or "").strip()
        if not name:
            return None
        return {
            "code": name,
            "percent": float(item.get("percent", 0)),
            "description": (item.get("description") or "").strip(),
        }

    return _load_simple_list(TAX_CODE_FILE, cleaner)


def save_tax_codes(items):
    _save_simple_list(TAX_CODE_FILE, items)


def load_uoms():
    def cleaner(item):
        name = (item.get("name") or "").strip()
        if not name:
            return None
        try:
            items = int(item.get("items", 0))
        except (TypeError, ValueError):
            items = 0
        return {
            "name": name,
            "items": items,
            "description": (item.get("description") or "").strip(),
        }

    return _load_simple_list(UOM_FILE, cleaner)


def save_uoms(items):
    _save_simple_list(UOM_FILE, items)


def load_warehouses():
    def cleaner(item):
        name = (item.get("name") or "").strip()
        if not name:
            return None
        return {
            "name": name,
            "location": (item.get("location") or "").strip(),
            "manager": (item.get("manager") or "").strip(),
            "contact": (item.get("contact") or "").strip(),
            "notes": (item.get("notes") or "").strip(),
        }

    return _load_simple_list(WAREHOUSE_FILE, cleaner)


def save_warehouses(items):
    _save_simple_list(WAREHOUSE_FILE, items)


def load_sizes():
    def cleaner(item):
        name = (item.get("name") or "").strip()
        if not name:
            return None
        return {"name": name}

    return _load_simple_list(SIZE_FILE, cleaner)


def save_sizes(items):
    _save_simple_list(SIZE_FILE, items)


def load_colors():
    def cleaner(item):
        name = (item.get("name") or "").strip()
        if not name:
            return None
        return {"name": name}

    return _load_simple_list(COLOR_FILE, cleaner)


def save_colors(items):
    _save_simple_list(COLOR_FILE, items)


def load_suppliers():
    def cleaner(item):
        name = (item.get("name") or "").strip()
        if not name:
            return None
        return {
            "name": name,
            "contact": (item.get("contact") or "").strip(),
            "phone": (item.get("phone") or "").strip(),
            "email": (item.get("email") or "").strip(),
            "address": (item.get("address") or "").strip(),
        }

    return _load_simple_list(SUPPLIER_FILE, cleaner)


def save_suppliers(items):
    _save_simple_list(SUPPLIER_FILE, items)


# =========================================
# ✅ ROLE HELPERS
# =========================================
def save_roles(roles):
    with open(ROLE_FILE, "w", encoding="utf-8") as f:
        json.dump(roles, f, indent=2, ensure_ascii=False)


# =========================================
# ✅ GLOBAL AFTER_REQUEST HEADERS
# =========================================
@app.after_request
def set_security_headers(resp):
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0, private"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


# =========================================
# MODULE INDEX (matches sidebar: Dashboard → Masters → CRM)
# =========================================
# 1. ROOT & AUTH         — /, /login, /signup, /logout, forgot-password, reset-password, OTP
# 2. DASHBOARD           — /dashboard
# 3. MASTERS — Manage Users   — /manage-users, /create-user, /update-user, /delete-user, /api/users
# 4. MASTERS — Department & Roles — /department-roles, /api/departments, /api/roles
# 5. MASTERS — Products  — /products, /products/create, /import, /api/products
# 6. MASTERS — Customer  — /customer, /import-customer, /addnew-customer, /api/customer, /api/customers
# 7. CRM — Enquiry List  — /enquiry-list, /api/enquiries (REST for Postman)
# 8. CRM — New Enquiry   — /new-enquiry, /save-enquiry, /add-product, /api/enquiry/…
# 9. UTILITY             — /profile, /search, /logout
# =========================================

# =========================================
# 1. ROOT & AUTH
# =========================================
@app.route("/")
def root():
    return redirect(url_for("login"))


# =========================================
# 2. DASHBOARD
# =========================================
@app.route("/dashboard")
def dashboard():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()

    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "dashboard.html",
        page="dashboard",
        title="Dashboard",
        user_email=user_email,
        user_name=user_name,
    )


# =========================================
# 1. ROOT & AUTH — Auth pages (GET)
# =========================================
@app.route("/signup")
def home():
    return render_template("signup.html")


@app.route("/login")
def login():
    message = request.args.get("message", "")

    # BUG_009: When the client explicitly asks for JSON (eg. Postman),
    # treat GET on /login as an invalid HTTP method for the JSON API.
    if wants_json():
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Method not allowed. Use POST /login for authentication.",
                }
            ),
            405,
        )

    return render_template("index.html", message=message)


@app.route("/forgot-password")
def forgot_password():
    return render_template("forgot-password.html")


@app.route("/check-your-mail")
def check_your_mail_page():
    email = request.args.get("email", "")
    return render_template("check-your-mail.html", email=email)


# =========================================
# 3. MASTERS — Manage Users
# =========================================
def _db_fetch_users_ordered(include_id: bool = False):
    """Return users in the same order as Manage Users table (latest first)."""
    conn = get_db_connection()
    cur = conn.cursor()
    if include_id:
        cur.execute(
            """
            SELECT id, name, email, phone, role, first_name, last_name, country_code,
                   contact_number, branch, department, reporting_to, available_branches, employee_id
            FROM users
            ORDER BY id DESC
            """
        )
        rows = cur.fetchall()
        users = []
        for r in rows:
            users.append(
                {
                    "id": r[0],
                    "name": r[1],
                    "email": r[2],
                    "phone": r[3],
                    "role": r[4],
                    "first_name": r[5],
                    "last_name": r[6],
                    "country_code": r[7],
                    "contact_number": r[8],
                    "branch": r[9],
                    "department": r[10],
                    "reporting_to": r[11],
                    "available_branches": str(r[12]) if r[12] is not None else "",
                    "employee_id": r[13],
                }
            )
    else:
        cur.execute("SELECT name, email, phone, role FROM users ORDER BY id DESC")
        rows = cur.fetchall()
        users = [{"name": r[0], "email": r[1], "phone": r[2], "role": r[3]} for r in rows]
    cur.close()
    conn.close()
    return users


def _db_get_user_by_email(email: str):
    """Get single DB user row by email (case-insensitive). Includes branch/department for RBAC."""
    conn = get_db_connection()
    cur = conn.cursor()
    row = None
    try:
        cur.execute(
            """
            SELECT id, name, email, phone, role, branch, department
            FROM users
            WHERE LOWER(email) = LOWER(%s)
            LIMIT 1
            """,
            ((email or "").strip(),),
        )
        row = cur.fetchone()
    except Exception:
        try:
            cur.execute(
                """
                SELECT id, name, email, phone, role
                FROM users
                WHERE LOWER(email) = LOWER(%s)
                LIMIT 1
                """,
                ((email or "").strip(),),
            )
            row = cur.fetchone()
        except Exception:
            row = None
    finally:
        cur.close()
        conn.close()
    if not row:
        return None
    if len(row) >= 7:
        return {
            "id": row[0],
            "name": row[1],
            "email": row[2],
            "phone": row[3],
            "role": row[4],
            "branch": row[5] or "",
            "department": row[6] or "",
        }
    return {
        "id": row[0],
        "name": row[1],
        "email": row[2],
        "phone": row[3],
        "role": row[4],
        "branch": "",
        "department": "",
    }


# --- RBAC: session + roles.json (matrix) + platform Admin / Super Admin ---
RBAC_MODULES = (
    "department_roles",
    "products",
    "customer",
    "new_enquiry",
    "quotation",
    "sales",
    "delivery",
    "invoice",
)


def _rbac_empty_perm():
    return {"full_access": False, "view": False, "create": False, "edit": False, "delete": False}


def _rbac_full_perm():
    return {"full_access": True, "view": True, "create": True, "edit": True, "delete": True}


def _rbac_admin_perm():
    """Admin policy: create/view/edit allowed, delete denied."""
    return {"full_access": False, "view": True, "create": True, "edit": True, "delete": False}


def normalize_menu_permissions(raw):
    """Normalize roles.json permission block (nested or flat checkbox keys from create-role UI)."""
    if not isinstance(raw, dict):
        return _rbac_empty_perm()
    if any(k in raw for k in ("full_access", "view", "create", "edit", "delete")):
        return {
            "full_access": bool(raw.get("full_access")),
            "view": bool(raw.get("view")),
            "create": bool(raw.get("create")),
            "edit": bool(raw.get("edit")),
            "delete": bool(raw.get("delete")),
        }
    fa = fv = fc = fe = fd = False
    for k, v in raw.items():
        if not v:
            continue
        ks = str(k).lower()
        if ks.endswith("_full") or ks == "full_access":
            fa = True
        elif ks.endswith("_view"):
            fv = True
        elif ks.endswith("_create"):
            fc = True
        elif ks.endswith("_edit"):
            fe = True
        elif ks.endswith("_delete"):
            fd = True
    if fa:
        return _rbac_full_perm()
    return {"full_access": False, "view": fv, "create": fc, "edit": fe, "delete": fd}


def get_current_user_profile():
    """
    Prefer PostgreSQL session + DB row (login uses DB). Fallback to users.json.
    Fixes RBAC when user exists only in DB or role differs from stale JSON.
    """
    email = session.get("user")
    if not email:
        return None
    role = session.get("role")
    name = "User"
    department = session.get("department")
    branch = session.get("branch")

    dbu = _db_get_user_by_email(email)
    if dbu:
        name = dbu.get("name") or "User"
        role = dbu.get("role") or "User"
        if department is None:
            department = dbu.get("department")
        if branch is None:
            branch = dbu.get("branch")
    else:
        role = "User" 
    return {
    "email": email,
    "name": name,
    "role": role.strip(),
    "department": (department or "").strip(),
    "branch": (branch or "").strip() or "Main Branch",
}


   
def get_effective_permissions_for_session():
    """Effective menu permissions: platform Admin/Super Admin = full; else roles.json matrix."""
    empty = {m: _rbac_empty_perm() for m in RBAC_MODULES}
    if not session.get("user"):
        return {"is_platform_admin": False, **empty}

    prof = get_current_user_profile()
    if not prof:
        return {"is_platform_admin": False, **empty}

    rn = (prof.get("role") or "").strip().lower().replace(" ", "").replace("_", "")
    if rn in ("superadmin", "admin"):
        full = {m: _rbac_full_perm() for m in RBAC_MODULES}
        full["is_platform_admin"] = True
        return full

    roles = get_roles_from_db()
    # roles = get_roles_from_db()
    dept = (prof.get("department") or "").strip().lower()
    branch = (prof.get("branch") or "Main Branch").strip().lower()
    role_name = (prof.get("role") or "").strip().lower()
    matched = None
    for r in roles:
        if not isinstance(r, dict):
            continue
        rd = (r.get("department") or "").strip().lower()
        rb = (r.get("branch") or "").strip().lower()
        rr = (r.get("role") or "").strip().lower()
        if rd == dept and rb == branch and rr == role_name:
            matched = r
            break

    out = {"is_platform_admin": False}
    perms = (matched or {}).get("permissions") or {}
    for m in RBAC_MODULES:
        out[m] = normalize_menu_permissions(perms.get(m) or {})
    return out


@app.context_processor
def inject_rbac():
    try:
        if session.get("user"):
            return {"rbac": get_effective_permissions_for_session()}
    except Exception:
        pass
    return {"rbac": {}}


@app.context_processor
def inject_profile_display_name():
    """
    Inject consistent profile name/email for the top-right dropdown.

    Some routes were passing `user_name` from `users.json` (can be stale).
    We always prefer the DB-backed `get_current_user_profile()` here.
    """
    email = session.get("user")
    if not email:
        return {"profile_user_name": "User", "profile_user_email": ""}

    prof = get_current_user_profile() or {}
    return {
        "profile_user_name": prof.get("name") or "User",
        "profile_user_email": email,
    }


def _db_sync_users_id_sequence(cur):
    """Fix out-of-sync users.id sequence (common after manual imports)."""
    cur.execute(
        """
        SELECT setval(
            pg_get_serial_sequence('users', 'id'),
            COALESCE((SELECT MAX(id) FROM users), 0) + 1,
            false
        )
        """
    )


@app.route("/manage-users")
def manage_users():
    user_email = session.get("user")
    if not user_email:
        if wants_json():
            return jsonify({"success": False, "message": "Session expired"}), 401
        return redirect(url_for("login", message="session_expired"))

    users = _db_fetch_users_ordered(include_id=False)

    user_name = "User"
    user_role = "User"

    current_email = (user_email or "").strip().lower()

    for u in users:
        if not isinstance(u, dict):
            continue

        u_email = (u.get("email") or "").strip().lower()
        if u_email == current_email:
            user_name = u.get("name") or "User"
            user_role = (u.get("role") or "User").strip()
            break

    print("DEBUG manage_users: email =", user_email, "role =", user_role)

    if wants_json():
        return jsonify({
            "success": True,
            "users": users,
            "total": len(users),
            "current_user": {"email": user_email, "name": user_name, "role": user_role}
        }), 200

    return render_template(
        "manage-users.html",
        users=users,
        title="Manage Users - Stackly",
        page="manage_users",
        section="masters",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
    )
# =========================================
# 4. MASTERS — Department & Roles
# =========================================
@app.route("/department-roles")
def department_roles():
    user_email = session.get("user")
    if not user_email:
        if wants_json():
            return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
        return redirect(url_for("login", message="session_expired"))

    # departments = load_departments()
    departments = get_departments_from_db()
    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"

    if wants_json():
        return jsonify(
            {
                "success": True,
                "departments": departments,
                "total": len(departments),
                "current_user": {"email": user_email, "name": user_name, "role": user_role},
                "permissions": get_effective_permissions_for_session(),
            }
        ), 200

    return render_template(
        "department-roles.html",
        title="Department & Roles - Stackly",
        page="department_roles",
        section="masters",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
        departments=departments,
    )



@app.route("/department-roles/create", methods=["GET", "POST"])
def create_department():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"

    branches_list = [
        {"id": "main_branch", "name": "Main Branch"},
        {"id": "branch_1", "name": "Branch 1"},
        {"id": "branch_2", "name": "Branch 2"},
    ]

    roles = get_roles_from_db()
    print("ROLES COUNT:", len(roles))

    if request.method == "POST":
        # -----------------------------------------
        #  ROLE-BASED ACCESS CHECK
        #  Only Super Admin and Admin can create departments
        # -----------------------------------------
        normalized_role = user_role.replace(" ", "").replace("_", "").lower()
        if normalized_role not in ["superadmin", "admin"]:
            error = "User cannot create new departments."
            return render_template(
                "create-department.html",
                title="Create Department - Stackly",
                page="department_roles",
                section="masters",
                user_email=user_email,
                user_name=user_name,
                user_role=user_role,
                error=error,
                branches=branches_list,
                roles=roles,
            )

        code = (request.form.get("code") or "").strip()
        name = (request.form.get("department_name") or "").strip()
        branch = (request.form.get("branch") or "").strip()
        desc = (request.form.get("description") or "").strip()

        if not code:
            error = "Department code is required."
            return render_template(
                "create-department.html",
                title="Create Department - Stackly",
                page="department_roles",
                section="masters",
                user_email=user_email,
                user_name=user_name,
                user_role=user_role,
                error=error,
                form={"code": code, "department_name": name, "branch": branch, "description": desc},
                branches=[],
                roles=roles,
            )

        if not name:
            error = "Department name is required."
            return render_template(
                "create-department.html",
                title="Create Department - Stackly",
                page="department_roles",
                section="masters",
                user_email=user_email,
                user_name=user_name,
                user_role=user_role,
                error=error,
                form={"code": code, "department_name": name, "branch": branch, "description": desc},
                branches=[],
                roles=roles,
            )

        departments = get_departments_from_db()

        # Check for duplicates (case-insensitive) - either code OR name should be unique
        for d in departments:
            existing_code = (d.get("code") or "").strip().lower()
            existing_name = (d.get("name") or "").strip().lower()
            new_code = code.lower()
            new_name = name.lower()
            
            if existing_code == new_code:
                error = "Department code already exists. Please use a different code."
                return render_template(
                    "create-department.html",
                    title="Create Department - Stackly",
                    page="department_roles",
                    section="masters",
                    user_email=user_email,
                    user_name=user_name,
                    user_role=user_role,
                    error=error,
                    form={"code": code, "department_name": name, "branch": branch, "description": desc},
                    branches=branches_list,
                roles=roles,
                )
            
            if existing_name == new_name:
                error = "Department name already exists. Please use a different name."
                return render_template(
                    "create-department.html",
                    title="Create Department - Stackly",
                    page="department_roles",
                    section="masters",
                    user_email=user_email,
                    user_name=user_name,
                    user_role=user_role,
                    error=error,
                    form={"code": code, "department_name": name, "branch": branch, "description": desc},
                    branches=branches_list,
                    roles=roles,
                )

        new_dept = {
            "code": code,
            "name": name,
            "branch": branch,
            "description": desc,
        }
        
        # 🔥 ADD DB INSERT
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO departments (code, name, branch, description)
            VALUES (%s, %s, %s, %s)
        """, (
            code,
            name,
            branch,
            desc
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Department has been created successfully", "success")
        return redirect(url_for("department_roles"))

    return render_template(
        "create-department.html",
        title="Create Department - Stackly",
        page="department_roles",
        section="masters",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
        branches=branches_list,
        roles=roles,
    )

# =========================================
# 4. MASTERS — Department & Roles — Edit/Delete (UI)
# =========================================
@app.route("/department-roles/edit", methods=["POST"])
def edit_department():
    # Session check (same as Edit Product / other edit modules)
    user_email = session.get("user")
    if not user_email:
        return jsonify(success=False, error="Session expired. Please login first."), 401

    try:
        data = request.get_json(silent=True) or {}
        # Original code identifies the row (before rename); legacy clients may send "id" with the old code
        original_code = (data.get("original_code") or data.get("id") or "").strip()
        code = (data.get("code") or "").strip()
        name = (data.get("name") or "").strip()
        description = data.get("description")

        if not original_code:
            return jsonify(success=False, error="Missing department identifier (original code)"), 400

        departments = get_departments_from_db()
        if not isinstance(departments, list):
            departments = []

        current = find_department_by_code(departments, original_code)
        if not current:
            return jsonify(success=False, error="Department not found"), 404

        # Check for duplicates (case-insensitive) - exclude current department
        new_code = code.lower()
        new_name = name.lower()

        for dept in departments:
            if dept is current:
                continue
            existing_code = (dept.get("code") or "").strip().lower()
            existing_name = (dept.get("name") or "").strip().lower()
            if existing_code == new_code:
                return jsonify(success=False, error="Department code already exists. Please use a different code."), 409
            if existing_name == new_name:
                return jsonify(success=False, error="Department name already exists. Please use a different name."), 409

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            UPDATE departments
            SET code = %s, name = %s, description = %s
            WHERE LOWER(code) = LOWER(%s)
        """, (
            code,
            name,
            description,
            original_code
        ))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify(success=True)
   
    except Exception as e:
        print("EDIT ERROR:", e)
        return jsonify(success=False, error=str(e)), 500

@app.route("/department-roles/delete", methods=["POST"])
def delete_department():
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "error": "session_expired"}), 401

    data = request.get_json(silent=True) or {}
    code_ref = (data.get("code") or data.get("id") or "").strip()

    if not code_ref:
        return jsonify({"success": False, "error": "missing_code"}), 400
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM departments WHERE LOWER(code) = LOWER(%s)", (code_ref,))

    if cur.rowcount == 0:
        cur.close()
        conn.close()
        return jsonify({"success": False, "error": "not_found"}), 404

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})

    

# =========================================
# =========================================
# 4. MASTERS — Department & Roles — APIs
# =========================================
@app.route("/api/me/permissions", methods=["GET"])
def api_me_permissions():
    """JSON: effective RBAC matrix for the logged-in user (session + roles.json)."""
    if not session.get("user"):
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    return jsonify(
        {
            "success": True,
            "permissions": get_effective_permissions_for_session(),
            "profile": get_current_user_profile(),
        }
    ), 200


@app.route("/api/departments", methods=["GET"])
def api_departments():
    """Get all departments - supports JSON response for Postman"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    departments = get_departments_from_db()
    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"
    perms = get_effective_permissions_for_session()

    return jsonify(
        {
            "success": True,
            "departments": [department_for_api(d) for d in departments if isinstance(d, dict)],
            "total": len(departments),
            "current_user": {
                "email": user_email,
                "name": user_name,
                "role": user_role,
            },
            "permissions": perms,
        }
    ), 200


@app.route("/api/departments/<path:dept_ref>", methods=["GET"])
def api_get_department(dept_ref):
    """Get single department by code (URL path, case-insensitive)."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    departments = get_departments_from_db()
    department = find_department_by_code(departments, dept_ref)
    
    if not department:
        return jsonify({"success": False, "message": "Department not found"}), 404
    
    return jsonify({
        "success": True,
        "department": department_for_api(department)
    }), 200



@app.route("/api/departments", methods=["POST"])
def api_create_department():
    """Create new department"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    # Check user role
    users = load_users()
    user_role = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_role = (u.get("role") or "User").strip()
            break
    
    normalized_role = user_role.replace(" ", "").replace("_", "").lower()
    if normalized_role not in ["superadmin", "admin"]:
        return jsonify({
            "success": False,
            "message": "User cannot create new departments."
        }), 403
    
    data = request.get_json() or {}
    code = (data.get("code") or "").strip()
    name = (data.get("name") or data.get("department_name") or "").strip()
    branch = (data.get("branch") or "").strip()
    description = (data.get("description") or "").strip()
    
    # Validation
    if not code:
        return jsonify({"success": False, "message": "Department code is required."}), 400
    
    if not name:
        return jsonify({"success": False, "message": "Department name is required."}), 400
    
    departments = get_departments_from_db()
    
    # Check for duplicates (case-insensitive)
    for d in departments:
        existing_code = (d.get("code") or "").strip().lower()
        existing_name = (d.get("name") or "").strip().lower()
        new_code = code.lower()
        new_name = name.lower()
        
        if existing_code == new_code:
            return jsonify({
                "success": False,
                "message": "Department code already exists. Please use a different code."
            }), 409
        
        if existing_name == new_name:
            return jsonify({
                "success": False,
                "message": "Department name already exists. Please use a different name."
            }), 409
    
    new_dept = {
        "code": code,
        "name": name,
        "branch": branch,
        "description": description,
    }
    departments.append(new_dept)
    # save_departments(departments)
    
    return jsonify({
        "success": True,
        "message": "Department created successfully",
        "department": department_for_api(new_dept)
    }), 201



@app.route("/api/departments/<path:dept_ref>", methods=["PUT"])
def api_update_department(dept_ref):
    """Update existing department"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    # Check user role
    users = load_users()
    user_role = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_role = (u.get("role") or "User").strip()
            break
    
    normalized_role = user_role.replace(" ", "").replace("_", "").lower()
    if normalized_role not in ["superadmin", "admin"]:
        return jsonify({
            "success": False,
            "message": "Only Super Admin or Admin can edit departments."
        }), 403
    
    data = request.get_json() or {}
    code = (data.get("code") or "").strip()
    name = (data.get("name") or "").strip()
    description = (data.get("description") or "").strip()
    
    if not dept_ref or not str(dept_ref).strip():
        return jsonify({"success": False, "message": "Department code is required in the URL."}), 400
    
    departments = load_departments()
    current = find_department_by_code(departments, dept_ref)
    if not current:
        return jsonify({"success": False, "message": "Department not found"}), 404
    
    # Check for duplicates (case-insensitive) - exclude current department
    merged_code = (code or current.get("code") or "").strip().lower()
    merged_name = (name or current.get("name") or "").strip().lower()

    for dept in departments:
        if dept is current:
            continue

        existing_code = (dept.get("code") or "").strip().lower()
        existing_name = (dept.get("name") or "").strip().lower()

        if existing_code == merged_code:
            return jsonify({
                "success": False,
                "message": "Department code already exists. Please use a different code."
            }), 409

        if existing_name == merged_name:
            return jsonify({
                "success": False,
                "message": "Department name already exists. Please use a different name."
            }), 409

    if code:
        current["code"] = code
    if name:
        current["name"] = name
    if description is not None:
        current["description"] = description
    
    save_departments(departments)
    return jsonify({
        "success": True,
        "message": "Department updated successfully",
        "department": department_for_api(current)
    }), 200


@app.route("/api/departments/<path:dept_ref>", methods=["DELETE"])
def api_delete_department(dept_ref):
    """Delete department by code (URL path segment)."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    # Check user role
    users = load_users()
    user_role = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_role = (u.get("role") or "User").strip()
            break
    
    normalized_role = user_role.replace(" ", "").replace("_", "").lower()
    if normalized_role not in ["superadmin", "admin"]:
        return jsonify({
            "success": False,
            "message": "Only Super Admin or Admin can delete departments."
        }), 403
    
    if not dept_ref or not str(dept_ref).strip():
        return jsonify({"success": False, "message": "Department code is required in the URL."}), 400
    
    departments = load_departments()
    before_count = len(departments)
    cref = str(dept_ref).strip().lower()
    
    departments = [d for d in departments if _dept_code_key(d) != cref]
    after_count = len(departments)
    
    if after_count == before_count:
        return jsonify({"success": False, "message": "Department not found"}), 404
    
    save_departments(departments)
    return jsonify({
        "success": True,
        "message": "Department deleted successfully"
    }), 200


# =========================================
# 4. MASTERS — Department & Roles — Role UI
# =========================================
@app.route("/department-role/create/new")
def department_new():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()

    user_name = "User"
    user_role = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            user_role = (u.get("role") or "User").strip()
            break

    return render_template(
        "create-role.html",
        page="department_roles",
        section="masters",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role
    )


# =========================================
# 4. MASTERS — Department & Roles — Save/Edit Role (UI)
# =========================================
@app.route("/save_role", methods=["POST"])
def save_role():
    data = request.get_json() or {}

    try:
        # -----------------------------------------
        #  ROLE-BASED ACCESS CHECK
        #  Only Super Admin and Admin can create roles
        # -----------------------------------------
        user_email = session.get("user")
        if not user_email:
            return jsonify({"status": "error", "message": "session_expired"}), 401

        users = load_users()
        user_role = "User"
        for u in users:
            if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
                user_role = (u.get("role") or "User").strip().lower()
                break

        # Normalize role for comparison
        normalized_role = user_role.replace(" ", "").replace("_", "").lower()
        if normalized_role not in ["superadmin", "admin"]:
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "User cannot create roles.",
                    }
                ),
                403,
            )

        # -----------------------------------------
        #  VALIDATION: Description max 50 characters
        # -----------------------------------------
        description = (data.get("description") or "").strip()
        if description and len(description) > 50:
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Description must not exceed 50 characters.",
                    }
                ),
                400,
            )

        roles = get_roles_from_db()        

        # -----------------------------------------
        #  DUPLICATE CHECK
        #  - Combination of department + branch + role must be unique
        # -----------------------------------------
        new_dept   = (data.get("department") or "").strip().lower()
        new_branch = (data.get("branch") or "").strip().lower()
        new_role   = (data.get("role") or "").strip().lower()

        for r in roles:
            dept   = (r.get("department") or "").strip().lower()
            branch = (r.get("branch") or "").strip().lower()
            role   = (r.get("role") or "").strip().lower()

            if dept == new_dept and branch == new_branch and role == new_role:
                # Duplicate found → do NOT save
                return (
                    jsonify(
                        {
                            "status": "error",
                            "message": "This combination of Department, Branch and Role already exists.",
                        }
                    ),
                    409,
                )

        department = (data.get("department") or "").strip()
        branch = (data.get("branch") or "").strip()
        role = (data.get("role") or "").strip()

        if not department or not branch or not role:
            return jsonify({
                "status": "error",
                "message": "Department, Branch and Role are required"
            }), 400
                
        # No duplicate → append and save
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO roles (department_name, branch, role_name, description, permissions)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            data.get("department"),
            data.get("branch"),
            data.get("role"),
            data.get("description"),
            json.dumps(data.get("permissions", {}))
        ))

        conn.commit()
        cur.close()
        conn.close()
        # save_roles(roles)             # ✅ use same saver
        return jsonify({"status": "success"})
    except Exception as e:
        print("❌ data save error:", e)
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/department-roles/create/edit", methods=["POST"])
def edit_role():
    try:
        data = request.get_json() or {}

        old_role = (data.get("old_role") or "").strip()
        new_role = (data.get("role") or "").strip()
        description = (data.get("description") or "").strip()
        new_department = (data.get("department") or "").strip()

        if not old_role or not new_role:
            return jsonify(success=False, error="Missing role data")
        
        # Validate description max 50 characters
        if description and len(description) > 50:
            return jsonify(success=False, error="Description must not exceed 50 characters.")

        roles = get_roles_from_db()
        
        # -----------------------------------------
        #  DUPLICATE CHECK
        #  - Combination of Role + Department must be unique (case-insensitive)
        #  - Exclude the current role being edited
        # -----------------------------------------
        new_role_lower = new_role.lower()
        new_dept_lower = new_department.lower() if new_department else ""
        
        for r in roles:
            existing_role = (r.get("role") or "").strip()
            existing_dept = (r.get("department") or "").strip()
            
            # Skip the current role being edited
            if existing_role == old_role:
                continue
            
            # Check for duplicate combination (case-insensitive)
            if existing_role.lower() == new_role_lower and existing_dept.lower() == new_dept_lower:
                return jsonify(
                    success=False,
                    error="This combination of Role and Department already exists."
                )
        
        # No duplicate found → proceed with update
        updated = False
        for r in roles:
            if (r.get("role") or "").strip() == old_role:
                r["role"] = new_role
                r["description"] = description

                # Update department field if provided
                if new_department:
                    r["department"] = new_department

                updated = True
                break

        if not updated:
            return jsonify(success=False, error="Role not found")

        # save_roles(roles)
        return jsonify(success=True)

    except Exception as e:
        print("EDIT ERROR:", e)
        return jsonify(success=False, error=str(e))


@app.route("/department-roles/create/delete", methods=["POST"])
def delete_role():
    data = request.get_json(silent=True) or {}
    print("DELETE DATA:", data)

    description = data.get("description")

    if not description:
        return jsonify(success=False, error="missing_description"), 400

    roles = get_roles_from_db()
    before = len(roles)

    roles = [
        r for r in roles
        if r.get("description", "").strip().lower()
        != description.strip().lower()
    ]

    if len(roles) == before:
        return jsonify(success=False, error="not_found"), 404

    # save_roles(roles)
    return jsonify(success=True)


# =========================================
# 4. MASTERS — Department & Roles — APIs (continued)
# =========================================
@app.route("/api/roles", methods=["GET"])
def api_roles():
    """Get all roles - supports JSON response for Postman"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"
    roles = get_roles_from_db()
    
    return jsonify({
        "success": True,
        "roles": roles,
        "total": len(roles),
        "current_user": {
            "email": user_email,
            "name": user_name,
            "role": user_role
        }
    }), 200



@app.route("/api/roles/<int:role_index>", methods=["GET"])
def api_get_role(role_index):
    """Get single role by index (0-based)"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    roles = get_roles_from_db()
    
    if role_index < 0 or role_index >= len(roles):
        return jsonify({"success": False, "message": "Role index out of range"}), 404
    
    return jsonify({
        "success": True,
        "role": roles[role_index],
        "index": role_index
    }), 200


@app.route("/api/roles", methods=["POST"])
def api_create_role():
    """Create new role"""
    
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    prof = get_current_user_profile() or {}
    user_role = (prof.get("role") or "User").strip()
    print("FINAL ROLE IN API:", user_role)
    normalized_role = user_role.replace(" ", "").replace("_", "").lower()


    data = request.get_json() or {}
    department = (data.get("department") or "").strip()
    branch = (data.get("branch") or "").strip()
    role_name = (data.get("role") or "").strip()
    description = (data.get("description") or "").strip()
    
    # Validation
    if not department:
        return jsonify({"success": False, "message": "Department is required."}), 400
    
    if not branch:
        return jsonify({"success": False, "message": "Branch is required."}), 400
    
    if not role_name:
        return jsonify({"success": False, "message": "Role is required."}), 400
    
    if not description:
        return jsonify({"success": False, "message": "Description is required."}), 400
    
    if len(description) > 50:
        return jsonify({
            "success": False,
            "message": "Description must not exceed 50 characters."
        }), 400
    
    roles = get_roles_from_db()
    
    # Check for duplicates (case-insensitive) - combination of department + branch + role
    new_dept = department.lower()
    new_branch = branch.lower()
    new_role = role_name.lower()
    
    for r in roles:
        dept = (r.get("department") or "").strip().lower()
        br = (r.get("branch") or "").strip().lower()
        role = (r.get("role") or "").strip().lower()
        
        if dept == new_dept and br == new_branch and role == new_role:
            return jsonify({
                "success": False,
                "message": "This combination of Department, Branch and Role already exists."
            }), 409
    
    new_role_data = {
        "department": department,
        "branch": branch,
        "role": role_name,
        "description": description
    }
    
    # Add permissions if provided
    if "permissions" in data:
        new_role_data["permissions"] = data["permissions"]
    
    permissions = data.get("permissions", {})
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO roles (department_name, branch, role_name, description, permissions)
        VALUES (%s, %s, %s, %s, %s)
    """, (
        department,   # ✅ FIXED
        branch,
        role_name,
        description,
        json.dumps(permissions)
    ))

    conn.commit()
    cur.close()
    conn.close()
    
    return jsonify({
        "success": True,
        "message": "Role created successfully",
        "role": new_role_data
    }), 201


@app.route("/api/roles/<int:role_index>", methods=["PUT"])
def api_update_role(role_index):
    """Update existing role by index"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
     # ✅ DB role check (NO JSON)
    prof = get_current_user_profile() or {}
    user_role = prof.get("role", "User")
    
    normalized_role = user_role.replace(" ", "").replace("_", "").lower()
    if normalized_role not in ["superadmin", "admin"]:
        return jsonify({
            "success": False,
            "message": "Only Super Admin or Admin can edit roles."
        }), 403
    
    data = request.get_json() or {}
    new_role_name = (data.get("role") or "").strip()
    description = (data.get("description") or "").strip()
    new_department = (data.get("department") or "").strip()
    branch = (data.get("branch") or "").strip()
    permissions = data.get("permissions", {})
    conn = get_db_connection()
    cur = conn.cursor()

    # 🔥 get role ID using index
    cur.execute("""
        SELECT r.id, d.id
        FROM roles r
        JOIN departments d ON r.department_name = d.name
        ORDER BY r.id
    """)
    rows = cur.fetchall()
    if role_index >= len(rows):
        cur.close()
        conn.close()
        return jsonify({"success": False, "message": "Role not found"}), 404

    role_id = rows[role_index][0]
     # 🔥 get department_id if changed
    dept_id = None
    if new_department:
        cur.execute("SELECT id FROM departments WHERE name = %s", (new_department,))
        dept = cur.fetchone()
        if not dept:
            cur.close()
            conn.close()
            return jsonify({"success": False, "message": "Department not found"}), 400
        dept_id = dept[0]

    # 🔥 UPDATE QUERY
    cur.execute("""
        UPDATE roles
        SET 
            role_name = COALESCE(%s, role_name),
            description = COALESCE(%s, description),
            branch = COALESCE(%s, branch),
            department_id = COALESCE(%s, department_id),
            permissions = COALESCE(%s, permissions)
        WHERE id = %s
    """, (
        new_role_name if new_role_name else None,
        description if description else None,
        branch if branch else None,
        dept_id,
        json.dumps(permissions) if permissions else None,
        role_id
    ))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        "success": True,
        "message": "Role updated successfully"
    }), 200

@app.route("/api/roles/<int:role_index>", methods=["DELETE"])
def api_delete_role(role_index):
    """Delete role by index"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    prof = get_current_user_profile() or {}

    user_role = (prof.get("role") or "User") \
        .strip() \
        .replace(" ", "") \
        .replace("_", "") \
        .lower()

    print("DELETE ROLE CHECK:", user_role)

    if user_role not in ["superadmin", "admin"]:

    
        return jsonify({
            "success": False,
            "message": "Only Super Admin or Admin can delete roles."
        }), 403

    # ✅ SAME SOURCE AS UI
    roles = get_roles_from_db()

    if role_index < 0 or role_index >= len(roles):
        return jsonify({"success": False, "message": "Role not found"}), 404

    role_id = roles[role_index]["id"]   # 🔥 KEY FIX

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM roles WHERE id = %s", (role_id,))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        "success": True,
        "message": "Role deleted successfully"
    }), 200
    


# =========================================
# 9. UTILITY — Profile
# =========================================
@app.route("/profile")
def profile():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()

    user_name = "User"
    mobile = ""

    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").strip().lower() == user_email.lower():
            user_name = u.get("name", "User")
            mobile = u.get("phone", "")
            break

    return render_template(
        "profile.html",
        user_email=user_email,
        user_name=user_name,
        mobile=mobile,
        page="profile",
    )


# =========================================
# 1. ROOT & AUTH — Check Email (AJAX)
# =========================================
@app.route("/check-email", methods=["POST"])
def check_email():
    data = request.get_json()
    email = (data.get("email") or "").strip().lower()

    if not email:
        return jsonify({"status": "error", "message": "Email is required."}), 400

    users = load_users()
    exists = False

    if isinstance(users, list):
        for u in users:
            if isinstance(u, dict) and (u.get("email") or "").strip().lower() == email:
                exists = True
                break
    elif isinstance(users, dict) and email in users:
        exists = True

    if not exists:
        return jsonify({"status": "error", "message": "Email not registered."}), 404

    return jsonify({"status": "ok"}), 200


# =========================================
# 1. ROOT & AUTH — Forgot Password (AJAX)
# =========================================
@app.route("/send-reset-link", methods=["POST"])
def send_reset_link():
    data = request.get_json()
    email = (data.get("email") or "").strip().lower()

    if not email:
        return jsonify({"status": "error", "message": "Email is required"}), 400

    try:
        users = load_users()
        exists = False
        if isinstance(users, list):
            for u in users:
                if not isinstance(u, dict):
                    continue
                if (u.get("email") or "").strip().lower() == email:
                    exists = True
                    break
        elif isinstance(users, dict):
            for u in users.values():
                if not isinstance(u, dict):
                    continue
                if (u.get("email") or "").strip().lower() == email:
                    exists = True
                    break

        if not exists:
            return jsonify({"status": "error", "message": "Email not registered."}), 400

        now = time.time()
        locked_until = RESET_LOCK.get(email)
        if locked_until and now < locked_until:
            remaining = int(locked_until - now)
            return jsonify({
                "status": "error",
                "message": f"This email is locked. Try again after {remaining} seconds."
            }), 429

        count = RESET_SEND_COUNT.get(email, 0)
        if count >= MAX_RESET_SENDS:
            RESET_LOCK[email] = now + LOCKOUT_DURATION
            RESET_SEND_COUNT[email] = 0
            return jsonify({
                "status": "error",
                "message": "Reset link already sent 5 times. Email is locked for 2 minutes."
            }), 429

        RESET_SEND_COUNT[email] = count + 1

        token = str(uuid.uuid4())
        RESET_TOKENS[token] = email
        reset_link = url_for("reset_password_page", token=token, _external=True)

        subject = "Reset Your Password - Stackly POS"
        body = (
            "Hi,\n\n"
            "Click the link below to reset your password:\n\n"
            f"{reset_link}\n\n"
            "If you did not request this, please ignore this email.\n\n"
            "- Stackly Team"
        )

        send_email(email, subject, body)
        return jsonify({"status": "ok"}), 200

    except Exception as e:
        print("DEBUG send-reset-link error:", e)
        return jsonify({"status": "error", "message": "Server error while sending reset link."}), 500


# =========================================
# 1. ROOT & AUTH — Reset Password
# =========================================
@app.route("/reset-password")
def reset_password_page():
    token = request.args.get("token")
    if not token or token not in RESET_TOKENS:
        return "Invalid or expired reset link.", 400
    return render_template("reset-password.html", token=token)


@app.route("/reset-password", methods=["POST"])
def reset_password_submit():
    try:
        data = request.get_json() or {}
        token = data.get("token")
        new_password = (data.get("password") or "").strip()

        if not token or not new_password:
            return jsonify({"status": "error", "message": "Token and password are required."}), 400

        email = RESET_TOKENS.get(token)
        if not email:
            return jsonify({"status": "error", "message": "Invalid or expired token."}), 400

        email_key = email.strip().lower()
        users = load_users()
        updated = False

        for u in users:
            if not isinstance(u, dict):
                continue
            if (u.get("email") or "").strip().lower() == email_key:
                u["password"] = new_password
                updated = True
                break

        if not updated:
            return jsonify({"status": "error", "message": "User not found."}), 400

        save_users(users)
        RESET_TOKENS.pop(token, None)

        print("✅ Password reset for:", email_key)
        return jsonify({"status": "ok"}), 200

    except Exception as e:
        print("❌ Reset password error:", e)
        return jsonify({"status": "error", "message": "Server error while updating password."}), 500



# =========================================
# 3. MASTERS — Manage Users — Create User
# =========================================
@app.route("/create-user", methods=["GET", "POST"])
def create_user():
    user_email = session.get("user")
    if not user_email:
        # Check if JSON request
        if request.is_json or request.content_type == "application/json":
            return jsonify({"success": False, "message": "Session expired"}), 401
        return redirect(url_for("login", message="session_expired"))

    # ✅ DB USER FETCH
    

    # ✅ GET CURRENT USER FROM DB
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT name, role FROM users WHERE LOWER(email) = LOWER(%s)
    """, (user_email,))

    current_user = cursor.fetchone()

    cursor.close()
    conn.close()

    if current_user:
        user_name, user_role = current_user
    else:
        user_name = "User"
        user_role = "User"
    if request.method == "GET" and wants_json():
        return jsonify({
            "success": True,
            "page": "create-user",
            "current_user": {
                "email": user_email,
                "name": user_name,
                "role": user_role,
            },
        }), 200

    if request.method == "GET":
        return render_template(
            "create-user.html",
            title="Create User - Stackly",
            page="manage_users",
            section="masters",
            user_email=user_email,
            user_name=user_name,
            user_role=user_role,
            departments=get_departments_from_db(),
            roles=get_roles_from_db(),
        )

    # -----------------------------------------
    #  ROLE-BASED ACCESS CHECK
    #  Only Super Admin and Admin can create branch users
    # -----------------------------------------
    normalized_role = user_role.replace(" ", "").replace("_", "").lower()
    if normalized_role not in ["superadmin", "admin"]:
        error_message = "Create new branch user is restricted for your credentials."
        if request.is_json or request.content_type == "application/json":
            return jsonify({"success": False, "message": error_message}), 403
        else:
            flash(error_message, "error")
            return redirect(url_for("create_user"))

    # Determine if request is JSON or form data
    is_json_request = request.is_json or request.content_type == "application/json"
    
    if is_json_request:
        data = request.get_json(silent=True) or {}
        first_name = (data.get("first_name") or "").strip()
        last_name = (data.get("last_name") or "").strip()
        email = (data.get("email") or "").strip()
        country_code = (data.get("country_code") or "").strip()
        contact_number = (data.get("contact_number") or "").strip()
        branch = (data.get("branch") or "").strip()
        department = (data.get("department") or "").strip()
        role = (data.get("role") or "").strip()
        reporting_to = (data.get("reporting_to") or "").strip()
        available_branches = (data.get("available_branches") or "").strip()
        employee_id = (data.get("employee_id") or "").strip()
        new_password = (data.get("password") or "").strip()
    else:
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        email = request.form.get("email", "").strip()
        country_code = request.form.get("country_code", "").strip()
        contact_number = request.form.get("contact_number", "").strip()
        branch = request.form.get("branch", "").strip()
        department = request.form.get("department", "").strip()
        role = request.form.get("role", "").strip()
        reporting_to = request.form.get("reporting_to", "").strip()
        available_branches = request.form.get("available_branches", "").strip()
        employee_id = request.form.get("employee_id", "").strip()
        new_password = request.form.get("password", "").strip()

    # Validation errors list
    errors = []

    # Validate First Name
    if not first_name:
        errors.append("First Name is required")
    elif len(first_name) < 3:
        errors.append("First Name must be at least 3 characters")
    elif not NAME_REGEX.match(first_name):
        errors.append("First Name should contain only letters and spaces (3-20 characters)")

    # Validate Last Name
    if not last_name:
        errors.append("Last Name is required")
    elif len(last_name) < 3:
        errors.append("Last Name must be at least 3 characters")
    elif not NAME_REGEX.match(last_name):
        errors.append("Last Name should contain only letters and spaces (3-20 characters)")

    # Validate Email
    if not email:
        errors.append("Email is required")
    elif len(email) > MAX_EMAIL_LENGTH:
        errors.append(f"Email is too long (max {MAX_EMAIL_LENGTH} characters)")
    elif not EMAIL_REGEX.match(email):
        errors.append("Enter a valid email address")

    # Validate Country Code
    valid_country_codes = ["+91", "+971", "+974", "+966", "+94", "+880", "+977", "+1", "+44", "+61"]
    phone_rules = {
        "+91": 10,   # India
        "+971": 9,   # United Arab Emirates
        "+974": 8,   # Qatar
        "+966": 9,   # Saudi Arabia
        "+94": 9,    # Sri Lanka
        "+880": 10,  # Bangladesh
        "+977": 10,  # Nepal
        "+1": 10,    # United States
        "+44": 10,   # United Kingdom (mobile)
        "+61": 9     # Australia
    }
    
    is_valid_country_code = country_code in valid_country_codes
    
    if not country_code:
        errors.append("Country code is required")
    elif not is_valid_country_code:
        errors.append(f"Invalid country code. Valid codes are: {', '.join(valid_country_codes)}")

    # Validate Contact Number
    if not contact_number:
        errors.append("Contact Number is required")
    elif not re.match(r"^\d+$", contact_number):
        errors.append("Contact Number must contain digits only")
    elif is_valid_country_code and country_code in phone_rules:
        # Only validate length if country code is valid
        required_phone_len = phone_rules[country_code]
        if len(contact_number) != required_phone_len:
            errors.append(f"Contact Number must be exactly {required_phone_len} digits for {country_code}")

    # Validate Branch
    if not branch:
        errors.append("Branch is required")

    # Validate Department
    if not department:
        errors.append("Department is required")

    # Validate Role
    if not role:
        errors.append("Role is required")

    # Validate Reporting To
    if not reporting_to:
        errors.append("Reporting To is required")
    elif len(reporting_to) < 3:
        errors.append("Reporting To must be at least 3 characters")
    elif not re.match(r"^[A-Za-z.\-\s]{3,40}$", reporting_to):
        errors.append("Reporting To may contain letters, dots, hyphens and spaces (3-40 characters)")

    # Validate Available Branches (digits only per memory)
    if not available_branches:
        errors.append("Available Branches is required")
    elif not re.match(r"^\d+$", available_branches):
        errors.append("Available Branches must contain only digits")

    # Validate Employee ID
    if not employee_id:
        errors.append("Employee ID is required")
    elif not re.match(r"^[A-Za-z0-9\-]{1,20}$", employee_id):
        errors.append("Employee ID may have letters, numbers and '-' (max 20 characters)")
    # 🔥 DB DUPLICATE CHECK
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT email, contact_number, employee_id FROM users")
    db_users = cursor.fetchall()

    for u in db_users:
        db_email = (u[0] or "").lower()
        db_contact = u[1]
        db_emp_id = u[2]

        if email.lower() == db_email:
            errors.append("Email already exists")

        if contact_number == db_contact:
            errors.append("Contact number already exists")

        if employee_id == db_emp_id:
            errors.append("Employee ID already exists")

    cursor.close()
    conn.close()
   
    # Return errors if any
    if errors:
        if is_json_request:
            return jsonify({"success": False, "message": "; ".join(errors), "errors": errors}), 400
        else:
            for error in errors:
                flash(error, "error")
            return redirect(url_for("create_user"))

    # Create new user (no persisted id; password required in file — default if not supplied)
    full_name = (first_name + " " + last_name).strip()
    full_phone = f"{country_code}{contact_number}" if country_code and contact_number else contact_number
    stored_password = new_password or DEFAULT_BRANCH_USER_PASSWORD






    new_user = {
        "name": full_name,
        "phone": full_phone,
        "first_name": first_name,
        "last_name": last_name,
        "email": email,
        "country_code": country_code,
        "contact_number": contact_number,
        "branch": branch,
        "department": department,
        "role": role,
        "reporting_to": reporting_to,
        "available_branches": available_branches,
        "employee_id": employee_id,
        "password": stored_password,
    }

    # data = request.get_json() if request.is_json else request.form

    conn = get_db_connection()
    cursor = conn.cursor()

    print("DEBUG NAME:", full_name)
    print("DEBUG PHONE:", full_phone)

    insert_sql = """
    INSERT INTO users (
        name, phone, first_name, last_name, email,
        country_code, contact_number, branch, department,
        role, reporting_to, available_branches, employee_id, password
    )
    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """
    insert_vals = (
        full_name,
        full_phone,
        first_name,
        last_name,
        email,
        country_code,
        contact_number,
        branch,
        department,
        role,
        reporting_to,
        int(available_branches),
        employee_id,
        stored_password,
    )
    try:
        cursor.execute(insert_sql, insert_vals)
        conn.commit()
    except psycopg2.errors.UniqueViolation as e:
        # If users.id sequence is behind, sync once and retry.
        conn.rollback()
        if "users_pkey" in str(e):
            _db_sync_users_id_sequence(cursor)
            conn.commit()
            cursor.execute(insert_sql, insert_vals)
            conn.commit()
        else:
            raise
    except Exception as e:
        conn.rollback()
        if is_json_request:
            return jsonify({"success": False, "message": f"Failed to create user: {e}"}), 500
        flash("Failed to create user. Please try again.", "error")
        return redirect(url_for("create_user"))
    finally:
        cursor.close()
        conn.close()




    # Return appropriate response
    if is_json_request:
        return jsonify({
            "success": True,
            "message": "User created successfully",
            "user": {
                "name": new_user["name"],
                "phone": new_user["phone"],
                "first_name": new_user["first_name"],
                "last_name": new_user["last_name"],
                "email": new_user["email"],
                "country_code": new_user["country_code"],
                "contact_number": new_user["contact_number"],
                "branch": new_user["branch"],
                "department": new_user["department"],
                "role": new_user["role"],
                "reporting_to": new_user["reporting_to"],
                "available_branches": new_user["available_branches"],
                "employee_id": new_user["employee_id"],
            },
        }), 201
    else:
        flash("User created successfully", "success")
        return redirect(url_for("manage_users"))

def normalize_role(role: str) -> str:
    return (role or "").strip().lower().replace(" ", "").replace("_", "")# =========================================

# 3. MASTERS — Manage Users — Update User
# =========================================
@app.route("/update-user", methods=["POST"])
def update_user():
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired"}), 401

    data = request.get_json(silent=True) or {}
    try:
        idx = int(data.get("index", -1))
    except (TypeError, ValueError):
        idx = -1

    if idx < 0:
        return jsonify({"success": False, "message": "Invalid index"}), 400

    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip()
    phone = (data.get("phone") or "").strip()
    role = (data.get("role") or "").strip() or "Admin"

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # one lightweight role check query
        cur.execute(
            "SELECT role FROM users WHERE LOWER(email)=LOWER(%s) LIMIT 1",
            ((user_email or "").strip(),),
        )
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "message": "Current user not found"}), 403
        if normalize_role(row[0]) not in ["superadmin", "admin"]:
            return jsonify({"success": False, "message": "Only Super Admin / Admin can edit users."}), 403

        # map current table index -> DB id
        cur.execute("SELECT id FROM users ORDER BY id DESC OFFSET %s LIMIT 1", (idx,))
        target = cur.fetchone()
        if not target:
            return jsonify({"success": False, "message": "User index out of range"}), 400
        user_id = target[0]

        cur.execute(
            """
            UPDATE users
            SET name=%s, email=%s, phone=%s, role=%s
            WHERE id=%s
            """,
            (name, email, phone, role, user_id),
        )
        conn.commit()
        return jsonify({"success": True, "message": "User updated"}), 200
    finally:
        cur.close()
        conn.close()


# =========================================
# 5. MASTERS — Products
# =========================================
@app.route("/products")
def products():
    user_email = session.get("user")
    if not user_email:
        if wants_json():
            return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
        return redirect(url_for("login"))

    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"

    if wants_json():
        products_list = load_products()
        return jsonify(
            {
                "success": True,
                "products": products_list,
                "total": len(products_list),
                "current_user": {"email": user_email, "name": user_name, "role": user_role},
                "permissions": get_effective_permissions_for_session(),
            }
        ), 200

    return render_template(
        "products.html",
        title="Product Master - Stackly",
        page="products",
        section="masters",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
    )


# =========================================
# ✅ API: PRODUCT CATEGORIES (PERSISTENT)
# =========================================
@app.route("/api/product-categories", methods=["GET", "POST"])
def api_product_categories():
    """
    GET  /api/product-categories?type=Electronics
         → { success, categories: ["Headphones", ...] }
    POST /api/product-categories
         JSON: { "name": "Headphones", "product_type": "Electronics" }
    """
    if request.method == "GET":
        product_type = (request.args.get("type") or "").strip()
        all_cats = load_product_categories()

        if product_type:
            # When a product type is specified, return categories
            # saved for that type PLUS any "global" categories
            # that were saved without a product_type.
            pt_norm = product_type.strip().lower()
            names = []
            for c in all_cats:
                ptype = (c.get("product_type") or "").strip().lower()
                if ptype in ("", pt_norm):
                    names.append(c.get("name") or "")
        else:
            # If no product_type is specified, return all names
            names = [c.get("name") or "" for c in all_cats]

        # remove duplicates (case-insensitive) while preserving order
        seen = set()
        unique = []
        for n in names:
            name_clean = (n or "").strip()
            if not name_clean:
                continue
            key = name_clean.lower()
            if key in seen:
                continue
            seen.add(key)
            unique.append(name_clean)

        return jsonify({"success": True, "categories": unique}), 200

    # POST → save new category
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    product_type = (data.get("product_type") or "").strip()

    if not name:
        return jsonify({"success": False, "message": "Category name is required."}), 400
    if len(name) < 3 or len(name) > 50:
        return jsonify({"success": False, "message": "Category Name Should contain atleast 3 characters."}), 400
    # Only alphabets and spaces allowed
    if not re.fullmatch(r"[A-Za-z\s]+", name):
        return jsonify({"success": False, "message": "Category name can contain only letters and spaces."}), 400

    # optional product_type – but we still store it for filtering
    all_cats = load_product_categories()
    name_lower = name.lower()
    type_lower = product_type.lower()

    for c in all_cats:
        if (c.get("name") or "").strip().lower() == name_lower and \
           (c.get("product_type") or "").strip().lower() == type_lower:
            return jsonify({"success": False, "message": "Category already exists for this product type."}), 409

    all_cats.append({"product_type": product_type, "name": name})
    save_product_categories(all_cats)

    return jsonify({"success": True, "message": "Category saved successfully."}), 201


# =========================================
# ✅ API: PRODUCT MASTER DROPDOWNS
#    Tax codes, UOM, Warehouse, Size, Color, Supplier
# =========================================
def _require_login_json():
    user_email = session.get("user")
    if not user_email:
        return None, jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    return user_email, None, None


@app.route("/api/product-tax-codes", methods=["GET", "POST"])
def api_product_tax_codes():
    if request.method == "GET":
        codes = load_tax_codes()
        return jsonify({"success": True, "items": codes}), 200

    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    data = request.get_json(silent=True) or {}
    code = (data.get("code") or "").strip()
    description = (data.get("description") or "").strip()
    try:
        percent = float(data.get("percent", 0))
    except (TypeError, ValueError):
        percent = 0

    if not code:
        return jsonify({"success": False, "message": "Tax name is required."}), 400

    if percent < 1 or percent > 100:
        return jsonify({"success": False, "message": "Tax percentage must be between 1 and 100."}), 400

    # Extract pure tax name (before "(xx%)") for validation & duplicate check
    base_name = code.split("(")[0].strip()

    # Tax name: only alphabets and spaces, at least 3 characters
    if len(base_name) < 3:
        return jsonify({"success": False, "message": "Tax Name should contain atleast 3 characters."}), 400
    if not re.fullmatch(r"[A-Za-z\s]+", base_name):
        return jsonify({"success": False, "message": "Tax name can contain only letters and spaces."}), 400

    items = load_tax_codes()
    key = base_name.lower()
    for item in items:
        existing = (item.get("code") or "").strip()
        existing_base = existing.split("(")[0].strip().lower()
        if existing_base == key:
            # same tax name already present → treat as duplicate name
            return jsonify({"success": False, "message": "This tax name already exists."}), 409

    items.append({"code": code, "percent": percent, "description": description})
    save_tax_codes(items)
    return jsonify({"success": True, "message": "Tax code saved."}), 201


@app.route("/api/product-uoms", methods=["GET", "POST"])
def api_product_uoms():
    if request.method == "GET":
        uoms = load_uoms()
        return jsonify({"success": True, "items": uoms}), 200

    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    try:
        items_val = int(data.get("items", 0))
    except (TypeError, ValueError):
        items_val = 0
    description = (data.get("description") or "").strip()

    if not name:
        return jsonify({"success": False, "message": "UOM name is required."}), 400
    if len(name) < 3:
        return jsonify({"success": False, "message": "UOM Name should contain atleast 3 characters."}), 400
    if len(name) > 50:
        return jsonify({"success": False, "message": "UOM Name should not exceed 50 characters."}), 400

    items = load_uoms()
    key = name.strip().lower()
    for item in items:
        if (item.get("name") or "").strip().lower() == key:
            # Duplicate UOM name not allowed
            return jsonify({"success": False, "message": "UOM name already exists."}), 409

    items.append({"name": name, "items": items_val, "description": description})
    save_uoms(items)
    return jsonify({"success": True, "message": "UOM saved."}), 201


@app.route("/api/product-warehouses", methods=["GET", "POST"])
def api_product_warehouses():
    if request.method == "GET":
        warehouses = load_warehouses()
        return jsonify({"success": True, "items": warehouses}), 200

    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    location = (data.get("location") or "").strip()
    manager = (data.get("manager") or "").strip()
    contact = (data.get("contact") or "").strip()
    notes = (data.get("notes") or "").strip()

    if not name:
        return jsonify({"success": False, "message": "Warehouse name is required."}), 400
    if len(name) < 3:
        return jsonify({"success": False, "message": "Warehouse Name should contain atleast 3 characters."}), 400
    if len(notes) > 50:
        return jsonify({"success": False, "message": "Notes must be 50 characters or less."}), 400

    items = load_warehouses()
    key = name.strip().lower()
    for item in items:
        if (item.get("name") or "").strip().lower() == key:
            # Duplicate warehouse name not allowed
            return jsonify({"success": False, "message": "Warehouse name already exists."}), 409

    items.append(
        {
            "name": name,
            "location": location,
            "manager": manager,
            "contact": contact,
            "notes": notes,
        }
    )
    save_warehouses(items)
    return jsonify({"success": True, "message": "Warehouse saved."}), 201


@app.route("/api/product-sizes", methods=["GET", "POST"])
def api_product_sizes():
    if request.method == "GET":
        sizes = load_sizes()
        return jsonify({"success": True, "items": sizes}), 200

    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()

    if not name:
        return jsonify({"success": False, "message": "Size name is required."}), 400
    if len(name) < 3:
        return jsonify({"success": False, "message": "Size Name should contain atleast 3 characters."}), 400

    items = load_sizes()
    key = name.strip().lower()
    for item in items:
        if (item.get("name") or "").strip().lower() == key:
            return jsonify({"success": True, "message": "Size already exists."}), 200

    items.append({"name": name})
    save_sizes(items)
    return jsonify({"success": True, "message": "Size saved."}), 201


@app.route("/api/product-colors", methods=["GET", "POST"])
def api_product_colors():
    if request.method == "GET":
        colors = load_colors()
        return jsonify({"success": True, "items": colors}), 200

    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()

    if not name:
        return jsonify({"success": False, "message": "Color name is required."}), 400
    if len(name) < 3:
        return jsonify({"success": False, "message": "Color Name should contain atleast 3 characters."}), 400

    items = load_colors()
    key = name.strip().lower()
    for item in items:
        if (item.get("name") or "").strip().lower() == key:
            return jsonify({"success": True, "message": "Color already exists."}), 200

    items.append({"name": name})
    save_colors(items)
    return jsonify({"success": True, "message": "Color saved."}), 201


@app.route("/api/product-suppliers", methods=["GET", "POST"])
def api_product_suppliers():
    if request.method == "GET":
        suppliers = load_suppliers()
        return jsonify({"success": True, "items": suppliers}), 200

    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    contact = (data.get("contact") or "").strip()
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip()
    address = (data.get("address") or "").strip()

    if not name:
        return jsonify({"success": False, "message": "Supplier name is required."}), 400
    if len(name) < 3:
        return jsonify({"success": False, "message": "Supplier Name should contain atleast 3 characters."}), 400
    if not re.fullmatch(r"[A-Za-z\s]+", name):
        return jsonify({"success": False, "message": "Supplier Name should contain atleast 3 characters."}), 400

    # Contact person: apply same rules as Supplier Name
    if not contact or len(contact.strip()) < 3 or not re.fullmatch(r"[A-Za-z\s]+", contact.strip()):
        return jsonify({"success": False, "message": "Contact Person Name should contain atleast 3 characters."}), 400

    items = load_suppliers()
    key = name.strip().lower()
    for item in items:
        if (item.get("name") or "").strip().lower() == key:
            # Duplicate supplier name not allowed
            return jsonify({"success": False, "message": "Supplier name already exists."}), 409

    items.append(
        {
            "name": name,
            "contact": contact,
            "phone": phone,
            "email": email,
            "address": address,
        }
    )
    save_suppliers(items)
    return jsonify({"success": True, "message": "Supplier saved."}), 201

# =========================
# API: GET SINGLE PRODUCT (Supports HTML & JSON)
# =========================
@app.route("/api/products/<product_id>", methods=["GET"])
def api_get_product(product_id):
    """
    GET /api/products/<product_id>
    
    Returns a single product by ID
    Supports both JSON and HTML responses
    """
    # BUG_001 / BUG_006: Require login for product APIs and return JSON 401
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    products = load_products()
    p = next((x for x in products if x.get("product_id") == str(product_id)), None)
    
    if not p:
        error_response = {
            "success": False,
            "message": "Product not found",
            "error": f"Product with ID '{product_id}' does not exist"
        }
        if wants_json():
            return jsonify(error_response), 404
        else:
            return jsonify(error_response), 404
    
    response_data = {
        "success": True,
        "data": p,
        "message": "Product retrieved successfully"
    }
    
    if wants_json():
        return jsonify(response_data), 200
    else:
        return jsonify(response_data), 200

# =========================
# API: CREATE PRODUCT (Supports HTML & JSON)
# =========================
@app.route("/api/products", methods=["POST"])
def api_create_product():
    """
    POST /api/products
    Content-Type: application/json
    
    Request Body (JSON):
    {
        "product_name": "Product Name",
        "type": "Physical",
        "category": "Electronics",
        "status": "Active",
        "stock_level": 100,
        "price": 99.99,
        ... (other optional fields)
    }
    
    Returns created product with generated ID
    """
    # BUG_001 / BUG_006: Require login for product APIs
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    if not request.is_json:
        error_response = {
            "success": False,
            "message": "Content-Type must be application/json",
            "error": "Invalid request format"
        }
        if wants_json():
            return jsonify(error_response), 400
        else:
            return jsonify(error_response), 400
    
    data = request.get_json() or {}
    
    # Validation
    product_name = (data.get("product_name") or "").strip()
    if not product_name:
        error_response = {
            "success": False,
            "message": "Product name is required",
            "error": "Validation failed"
        }
        if wants_json():
            return jsonify(error_response), 400
        else:
            return jsonify(error_response), 400
    
    # Generate product ID
    product_id = generate_product_id()
    
    # Build product object
    product = {
        "product_id": str(product_id),
        "product_name": product_name,
        "type": (data.get("type") or "").strip(),
        "category": (data.get("category") or "").strip(),
        "status": (data.get("status") or "Active").strip(),
        "stock_level": int(data.get("stock_level", 0)),
        "price": float(data.get("price", 0.0)),
        "description": (data.get("description") or "").strip(),
        "sub_category": (data.get("sub_category") or "").strip(),
        "unit_price": (data.get("unit_price") or "").strip(),
        "discount": (data.get("discount") or "").strip(),
        "tax_code": (data.get("tax_code") or "").strip(),
        "quantity": (data.get("quantity") or "").strip(),
        "uom": (data.get("uom") or "").strip(),
        "reorder_level": (data.get("reorder_level") or "").strip(),
        "warehouse": (data.get("warehouse") or "").strip(),
        "size": (data.get("size") or "").strip(),
        "color": (data.get("color") or "").strip(),
        "weight": (data.get("weight") or "").strip(),
        "specifications": (data.get("specifications") or "").strip(),
        "related_products": (data.get("related_products") or "").strip(),
        "supplier": (data.get("supplier") or "").strip(),
        "product_usage": (data.get("product_usage") or "").strip(),
        "image": (data.get("image") or "").strip(),
    }
    
    # Save product
    products = load_products()
    products.append(product)
    save_products(products)
    
    response_data = {
        "success": True,
        "message": "Product created successfully",
        "data": product
    }
    
    if wants_json():
        return jsonify(response_data), 201
    else:
        return jsonify(response_data), 201


# =========================
# API: UPDATE PRODUCT (PUT - Full Update) (Supports HTML & JSON)
# =========================
@app.route("/api/products/<product_id>", methods=["PUT"])
def api_update_product(product_id):
    """
    PUT /api/products/<product_id>
    Content-Type: application/json
    
    Full update - replaces entire product
    """
    # BUG_001 / BUG_006: Require login for product APIs
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    if not request.is_json:
        error_response = {
            "success": False,
            "message": "Content-Type must be application/json",
            "error": "Invalid request format"
        }
        if wants_json():
            return jsonify(error_response), 400
        else:
            return jsonify(error_response), 400
    
    data = request.get_json(silent=True) or {}

    # BUG_005: Validate payload and reject obviously invalid values
    errors = []
    product_name = (data.get("product_name") or "").strip()
    if not product_name:
        errors.append("Product name is required.")

    if "stock_level" in data:
        try:
            stock_val = int(data.get("stock_level"))
            if stock_val < 0:
                errors.append("Stock level must be 0 or greater.")
        except (TypeError, ValueError):
            errors.append("Stock level must be a whole number.")

    if "price" in data:
        try:
            price_val = float(data.get("price"))
            if price_val <= 0:
                errors.append("Price must be greater than 0.")
        except (TypeError, ValueError):
            errors.append("Price must be a valid number.")

    if errors:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Validation failed",
                    "errors": errors,
                }
            ),
            400,
        )
    products = load_products()
    updated = False
    product = None

    for p in products:
        if str(p.get("product_id")) == str(product_id):
            # Full update - replace all fields
            p["product_name"] = (data.get("product_name") or p.get("product_name") or "").strip()
            p["type"] = (data.get("type") or p.get("type") or "").strip()
            p["category"] = (data.get("category") or p.get("category") or "").strip()
            p["status"] = (data.get("status") or p.get("status") or "Active").strip()

            try:
                p["stock_level"] = int(data.get("stock_level", p.get("stock_level", 0)))
            except:
                p["stock_level"] = 0

            try:
                p["price"] = float(data.get("price", p.get("price", 0)))
            except:
                p["price"] = 0.0

            # Update optional fields if provided
            if "description" in data:
                p["description"] = (data.get("description") or "").strip()
            if "sub_category" in data:
                p["sub_category"] = (data.get("sub_category") or "").strip()
            if "tax_code" in data:
                p["tax_code"] = (data.get("tax_code") or "").strip()
            if "supplier" in data:
                p["supplier"] = (data.get("supplier") or "").strip()

            product = p
            updated = True
            break

    if not updated:
        error_response = {
            "success": False,
            "message": "Product not found",
            "error": f"Product with ID '{product_id}' does not exist"
        }
        if wants_json():
            return jsonify(error_response), 404
        else:
            return jsonify(error_response), 404

    # -------------------- DUPLICATE VALIDATION (exclude current product) -----------------------------
    # Normalize values for comparison (case-insensitive for text, exact for numbers)
    updated_product_name = (product.get("product_name") or "").strip().lower()
    updated_type = (product.get("type") or "").strip().lower()
    updated_category = (product.get("category") or "").strip().lower()
    updated_status = (product.get("status") or "").strip().lower()
    updated_stock_level = product.get("stock_level", 0)
    updated_price = product.get("price", 0.0)
    
    # Check 1: Duplicate product name (case-insensitive, exclude current product)
    for existing in products:
        if str(existing.get("product_id")) == str(product_id):
            continue  # Skip the product being updated
        existing_name = (existing.get("product_name") or "").strip().lower()
        if existing_name == updated_product_name:
            error_response = {
                "success": False,
                "message": f"Product with name '{product.get('product_name')}' already exists. Please use a different product name."
            }
            if wants_json():
                return jsonify(error_response), 409
            else:
                return jsonify(error_response), 409
    
    # Check 2: Duplicate combination (exclude current product)
    for existing in products:
        if str(existing.get("product_id")) == str(product_id):
            continue  # Skip the product being updated
        existing_name = (existing.get("product_name") or "").strip().lower()
        existing_type = (existing.get("type") or "").strip().lower()
        existing_category = (existing.get("category") or "").strip().lower()
        existing_status = (existing.get("status") or "").strip().lower()
        existing_stock_level = existing.get("stock_level", 0)
        existing_price = existing.get("price", 0.0)
        
        # Compare all fields (case-insensitive for text, exact for numbers)
        if (existing_name == updated_product_name and
            existing_type == updated_type and
            existing_category == updated_category and
            existing_status == updated_status and
            existing_stock_level == updated_stock_level and
            abs(existing_price - updated_price) < 0.01):  # Float comparison with tolerance
            error_response = {
                "success": False,
                "message": f"A product with the same combination (Name: '{product.get('product_name')}', Type: '{product.get('type')}', Category: '{product.get('category')}', Status: '{product.get('status')}', Stock Level: {updated_stock_level}, Price: {updated_price}) already exists."
            }
            if wants_json():
                return jsonify(error_response), 409
            else:
                return jsonify(error_response), 409

    save_products(products)
    
    response_data = {
        "success": True,
        "message": "Product updated successfully",
        "data": product
    }
    
    if wants_json():
        return jsonify(response_data), 200
    else:
        return jsonify(response_data), 200


# =========================
# API: PARTIAL UPDATE PRODUCT (PATCH) (Supports HTML & JSON)
# =========================
@app.route("/api/products/<product_id>", methods=["PATCH"])
def api_patch_product(product_id):
    """
    PATCH /api/products/<product_id>
    Content-Type: application/json
    
    Partial update - only updates provided fields
    """
    # BUG_001 / BUG_006: Require login for product APIs
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    if not request.is_json:
        error_response = {
            "success": False,
            "message": "Content-Type must be application/json",
            "error": "Invalid request format"
        }
        if wants_json():
            return jsonify(error_response), 400
        else:
            return jsonify(error_response), 400
    
    data = request.get_json(silent=True) or {}

    # BUG_005: Validate any fields that are provided
    errors = []
    if "product_name" in data:
        name = (data.get("product_name") or "").strip()
        if not name:
            errors.append("Product name cannot be blank.")
    if "stock_level" in data:
        try:
            stock_val = int(data.get("stock_level"))
            if stock_val < 0:
                errors.append("Stock level must be 0 or greater.")
        except (TypeError, ValueError):
            errors.append("Stock level must be a whole number.")
    if "price" in data:
        try:
            price_val = float(data.get("price"))
            if price_val <= 0:
                errors.append("Price must be greater than 0.")
        except (TypeError, ValueError):
            errors.append("Price must be a valid number.")

    if errors:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Validation failed",
                    "errors": errors,
                }
            ),
            400,
        )
    products = load_products()
    updated = False
    product = None

    for p in products:
        if str(p.get("product_id")) == str(product_id):
            # Partial update - only update provided fields
            if "product_name" in data:
                p["product_name"] = (data.get("product_name") or "").strip()
            if "type" in data:
                p["type"] = (data.get("type") or "").strip()
            if "category" in data:
                p["category"] = (data.get("category") or "").strip()
            if "status" in data:
                p["status"] = (data.get("status") or "Active").strip()
            if "stock_level" in data:
                try:
                    p["stock_level"] = int(data.get("stock_level", 0))
                except:
                    p["stock_level"] = 0
            if "price" in data:
                try:
                    p["price"] = float(data.get("price", 0))
                except:
                    p["price"] = 0.0
            if "description" in data:
                p["description"] = (data.get("description") or "").strip()
            if "sub_category" in data:
                p["sub_category"] = (data.get("sub_category") or "").strip()
            if "tax_code" in data:
                p["tax_code"] = (data.get("tax_code") or "").strip()
            if "supplier" in data:
                p["supplier"] = (data.get("supplier") or "").strip()

            product = p
            updated = True
            break

    if not updated:
        error_response = {
            "success": False,
            "message": "Product not found",
            "error": f"Product with ID '{product_id}' does not exist"
        }
        if wants_json():
            return jsonify(error_response), 404
        else:
            return jsonify(error_response), 404

    # -------------------- DUPLICATE VALIDATION (exclude current product) -----------------------------
    # Normalize values for comparison (case-insensitive for text, exact for numbers)
    updated_product_name = (product.get("product_name") or "").strip().lower()
    updated_type = (product.get("type") or "").strip().lower()
    updated_category = (product.get("category") or "").strip().lower()
    updated_status = (product.get("status") or "").strip().lower()
    updated_stock_level = product.get("stock_level", 0)
    updated_price = product.get("price", 0.0)
    
    # Check 1: Duplicate product name (case-insensitive, exclude current product)
    for existing in products:
        if str(existing.get("product_id")) == str(product_id):
            continue  # Skip the product being updated
        existing_name = (existing.get("product_name") or "").strip().lower()
        if existing_name == updated_product_name:
            error_response = {
                "success": False,
                "message": f"Product with name '{product.get('product_name')}' already exists. Please use a different product name."
            }
            if wants_json():
                return jsonify(error_response), 409
            else:
                return jsonify(error_response), 409
    
    # Check 2: Duplicate combination (exclude current product)
    for existing in products:
        if str(existing.get("product_id")) == str(product_id):
            continue  # Skip the product being updated
        existing_name = (existing.get("product_name") or "").strip().lower()
        existing_type = (existing.get("type") or "").strip().lower()
        existing_category = (existing.get("category") or "").strip().lower()
        existing_status = (existing.get("status") or "").strip().lower()
        existing_stock_level = existing.get("stock_level", 0)
        existing_price = existing.get("price", 0.0)
        
        # Compare all fields (case-insensitive for text, exact for numbers)
        if (existing_name == updated_product_name and
            existing_type == updated_type and
            existing_category == updated_category and
            existing_status == updated_status and
            existing_stock_level == updated_stock_level and
            abs(existing_price - updated_price) < 0.01):  # Float comparison with tolerance
            error_response = {
                "success": False,
                "message": f"A product with the same combination (Name: '{product.get('product_name')}', Type: '{product.get('type')}', Category: '{product.get('category')}', Status: '{product.get('status')}', Stock Level: {updated_stock_level}, Price: {updated_price}) already exists."
            }
            if wants_json():
                return jsonify(error_response), 409
            else:
                return jsonify(error_response), 409

    save_products(products)
    
    response_data = {
        "success": True,
        "message": "Product updated successfully",
        "data": product
    }
    
    if wants_json():
        return jsonify(response_data), 200
    else:
        return jsonify(response_data), 200


# =========================================
# 5. MASTERS — Products (continued)
# =========================================
@app.route("/products/create")
def create_new_product_page():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "create-new-product.html",
        title="Create Product - Stackly",
        page="products",
        section="masters",
        user_email=user_email,
        user_name=user_name,
    )


@app.route("/download-template")
def download_template():
    # 1. Create empty dataframe
    df = pd.DataFrame(columns=[
        "Product ID",
        "Product Name",
        "Type",
        "Category",
        "Status",
        "Stock Level",
        "Price"
    ])

    # 2. Save to Excel in memory
    output = io.BytesIO()
    df.to_excel(output, index=False, sheet_name="Products")
    output.seek(0)

    # 3. Load workbook to add validations
    wb = load_workbook(output)
    ws = wb.active

    # ---------- DROPDOWN VALIDATIONS ----------
    type_validation = DataValidation(
        type="list",
        formula1='"Physical,Digital"',
        allow_blank=False
    )

    category_validation = DataValidation(
        type="list",
        formula1='"Electronics,Clothing,Food,Furniture"',
        allow_blank=False
    )

    status_validation = DataValidation(
        type="list",
        formula1='"Active,Inactive"',
        allow_blank=False
    )

    # ---------- NUMBER VALIDATIONS ----------
    stock_validation = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Stock Level",
        error="Stock Level must be a whole number (0 or greater)."
    )

    price_validation = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Price",
        error="Price must be a number greater than 0."
    )

    product_id_validation = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Product ID",
        error="Product ID must contain ONLY numbers (no letters or special characters)."
    )

    product_name_validation = DataValidation(
        type="custom",
        # Custom validation formulas in openpyxl should NOT include a leading "="
        formula1=(
            'AND(B2<>"",'
            'NOT(OR('
            'ISNUMBER(SEARCH("0",B2)),'
            'ISNUMBER(SEARCH("1",B2)),'
            'ISNUMBER(SEARCH("2",B2)),'
            'ISNUMBER(SEARCH("3",B2)),'
            'ISNUMBER(SEARCH("4",B2)),'
            'ISNUMBER(SEARCH("5",B2)),'
            'ISNUMBER(SEARCH("6",B2)),'
            'ISNUMBER(SEARCH("7",B2)),'
            'ISNUMBER(SEARCH("8",B2)),'
            'ISNUMBER(SEARCH("9",B2))'
            '))'
            ')'
        ),
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Product Name",
        error="Only alphabets and spaces are allowed. Numbers are not permitted."
    )

    # Add validations to worksheet
    ws.add_data_validation(product_id_validation)
    ws.add_data_validation(type_validation)
    ws.add_data_validation(category_validation)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(stock_validation)
    ws.add_data_validation(price_validation)
    ws.add_data_validation(product_name_validation)

    # Apply validation to columns (row 2 to 1000)
    product_id_validation.add("A2:A1000")
    product_name_validation.add("B2:B1000")
    type_validation.add("C2:C1000")
    category_validation.add("D2:D1000")
    status_validation.add("E2:E1000")
    stock_validation.add("F2:F1000")
    price_validation.add("G2:G1000")

    # 4. Save workbook again
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    # 5. Download file
    return send_file(
        final_output,
        as_attachment=True,
        download_name="Product_Import_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/download-customer-template")
def download_customer_template():

    # 1. Empty dataframe
    df = pd.DataFrame(columns=[
        "Customer ID",
        "Name",
        "Company",
        "Customer Type",
        "Email",
        "Status",
        "Credit Limit",
        "City"
    ])

    # 2. Save to memory
    output = io.BytesIO()
    df.to_excel(output, index=False, sheet_name="Customers")
    output.seek(0)

    # 3. Load workbook
    wb = load_workbook(output)
    ws = wb.active

    # ---------------- DROPDOWNS ----------------
    customer_type_validation = DataValidation(
        type="list",
        formula1='"Retail,Wholesale,Corporate,Online,Distributor"',
        allow_blank=False
    )

    status_validation = DataValidation(
        type="list",
        formula1='"Active,Inactive"',
        allow_blank=False
    )

    # ---------------- CUSTOMER ID VALIDATION (NUMBERS ONLY) ----------------
    # Match the simple numeric rule used for Product ID:
    # Allow: Whole number, greater than 0
    customer_id_validation = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Customer ID",
        error="Customer ID must contain ONLY numbers greater than 0 (no letters or special characters)."
    )





    

    credit_limit_validation = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Credit Limit",
        error="Credit Limit must be 0 or greater."
    )

    # ---------------- TEXT VALIDATIONS ----------------
    # Name validation (Customer template) – mirror the Product Name rule:
    # =AND(B2<>"",NOT(OR(ISNUMBER(SEARCH("0",B2)),...,ISNUMBER(SEARCH("9",B2)))))
    # NOTE: For openpyxl custom validations, formula1 should NOT start with "=".
    name_validation = DataValidation(
        type="custom",
        formula1=(
            'AND(B2<>"",'
            'NOT(OR('
            'ISNUMBER(SEARCH("0",B2)),'
            'ISNUMBER(SEARCH("1",B2)),'
            'ISNUMBER(SEARCH("2",B2)),'
            'ISNUMBER(SEARCH("3",B2)),'
            'ISNUMBER(SEARCH("4",B2)),'
            'ISNUMBER(SEARCH("5",B2)),'
            'ISNUMBER(SEARCH("6",B2)),'
            'ISNUMBER(SEARCH("7",B2)),'
            'ISNUMBER(SEARCH("8",B2)),'
            'ISNUMBER(SEARCH("9",B2))'
            '))'
            ')'
        ),
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Name",
        error="Name must contain only alphabets (letters) and spaces. Numbers and special characters are not allowed."
    )

    city_validation = DataValidation(
        type="custom",
        formula1='=AND(H2<>"",NOT(ISNUMBER(SEARCH("0",H2))))',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid City",
        error="City must contain only letters."
    )

    email_validation = DataValidation(
        type="custom",
        formula1='=AND(E2<>"",ISNUMBER(SEARCH("@",E2)),ISNUMBER(SEARCH(".",E2)))',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Email",
        error="Enter a valid email address."
    )

    # Add validations
    ws.add_data_validation(customer_id_validation)
    ws.add_data_validation(name_validation)
    ws.add_data_validation(customer_type_validation)
    ws.add_data_validation(email_validation)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(credit_limit_validation)
    ws.add_data_validation(city_validation)

    # Apply validations (Row 2–1000)
    customer_id_validation.add("A2:A1000")
    name_validation.add("B2:B1000")
    customer_type_validation.add("D2:D1000")
    email_validation.add("E2:E1000")
    status_validation.add("F2:F1000")
    credit_limit_validation.add("G2:G1000")
    city_validation.add("H2:H1000")

    # Save final file
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return send_file(
        final_output,
        as_attachment=True,
        download_name="Customer_Import_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/import", methods=["GET", "POST"])
def import_products():
    user_email = session.get("user")
    if not user_email:
        if request.method == "POST":
            return jsonify({
                "status": "error",
                "message": "Session expired. Please login first."
            }), 401
        return redirect(url_for("login", message="session_expired"))

    if request.method == "POST":
        file = request.files.get("file")

        if not file:
            return jsonify({
                "status": "error",
                "message": "No file uploaded"
            })

        # Save uploaded file
        upload_folder = "uploads"
        os.makedirs(upload_folder, exist_ok=True)
        file.save(os.path.join(upload_folder, file.filename))

        # Dummy response (replace later)
        return jsonify({
            "status": "success",
            "valid": 120,
            "invalid": 5,
            "skipped": 2,
            "errors": [
                'Missing "UOM" in Row 10,12,13',
                'Invalid GST Rate in Row 18'
            ]
        })

    # GET → open import page
    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "import-product.html",
        title="Import Products - Stackly",
        page="products",
        section="masters",
        user_email=user_email,
        user_name=user_name,
    )


@app.get("/import-product")
def import_product_metadata():
    """Small JSON endpoint so /import page has a named Fetch/XHR entry."""
    user_email = session.get("user")
    if not user_email:
        return jsonify(
            {"success": False, "message": "Session expired. Please login first."}
        ), 401

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return jsonify(
        {
            "success": True,
            "page": "import-product",
            "current_user": {"email": user_email, "name": user_name},
        }
    ), 200


def _product_row_signature_from_excel(row) -> tuple:
    """
    Canonical 6-tuple for product duplicate detection (import + validation).
    Text fields compared case-insensitively; stock/price normalized so Excel 12 vs 12.0 match.
    """
    def _s(v):
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except (TypeError, ValueError):
            pass
        return str(v).strip()

    name = _s(row.get("Product Name"))
    t = _s(row.get("Type"))
    cat = _s(row.get("Category"))
    st = _s(row.get("Status"))
    stock_raw = row.get("Stock Level")
    price_raw = row.get("Price")
    try:
        stock = int(float(stock_raw)) if not pd.isna(stock_raw) else 0
    except (TypeError, ValueError):
        stock = 0
    try:
        price = float(price_raw) if not pd.isna(price_raw) else 0.0
    except (TypeError, ValueError):
        price = 0.0
    price_r = round(price + 1e-12, 2)
    return (name.lower(), t.lower(), cat.lower(), st.lower(), stock, price_r)


def _product_signature_from_stored(p: dict) -> tuple:
    """Same tuple as _product_row_signature_from_excel for products in product.json."""
    try:
        stock = int(float(p.get("stock_level", 0) or 0))
    except (TypeError, ValueError):
        stock = 0
    try:
        price = float(p.get("price", 0) or 0)
    except (TypeError, ValueError):
        price = 0.0
    price_r = round(price + 1e-12, 2)
    return (
        str(p.get("product_name") or "").strip().lower(),
        str(p.get("type") or "").strip().lower(),
        str(p.get("category") or "").strip().lower(),
        str(p.get("status") or "").strip().lower(),
        stock,
        price_r,
    )


@app.route("/upload", methods=["POST"])
def upload_file():
    file = request.files.get("file")

    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    # BUG_003 / BUG_004: validate file extension early (expect Excel template)
    filename_lower = file.filename.lower()
    if not (filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls")):
        return (
            jsonify(
                {
                    "error": "Invalid file format. Please upload the provided Excel template (.xlsx or .xls)."
                }
            ),
            400,
        )

    try:
        df = pd.read_excel(file)
    except Exception:
        return jsonify({"error": "Invalid Excel file"}), 400

    # Check if file is empty (no data rows, only headers)
    if df.empty or len(df) == 0:
        return jsonify({
            "error": "No data found",
            "message": "The uploaded file contains no data. Please ensure the file has at least one row of product data."
        }), 400

    # Ensure required columns exist in the uploaded template
    required_columns = [
        "Product ID",
        "Product Name",
        "Type",
        "Category",
        "Status",
        "Stock Level",
        "Price",
    ]
    missing_cols = [c for c in required_columns if c not in df.columns]
    if missing_cols:
        return (
            jsonify(
                {
                    "error": "Invalid template",
                    "message": "The uploaded Excel file does not match the required product import template.",
                    "missing_columns": missing_cols,
                }
            ),
            400,
        )

    valid_rows = 0
    invalid_rows = 0
    skipped_rows = 0
    error_details = []
    skipped_row_numbers = []  # Track which rows were skipped (completely blank)
    
    
    # Load existing products to check against database
    existing_products = load_products()
    existing_product_ids = {str(p.get("product_id", "")) for p in existing_products if p.get("product_id")}
    existing_signatures = {_product_signature_from_stored(p) for p in existing_products}
    
    # Track Product IDs for uniqueness validation within uploaded file
    seen_product_ids = {}  # key: Product ID (as string), value: first row number
    
    # Track seen row combinations for duplicate detection (excluding Product ID)
    seen_rows = {}  # signature -> first Excel row number

    # Helper to check blank (NaN or empty/whitespace)
    def is_blank(val):
        if pd.isna(val):
            return True
        return str(val).strip() == ""
    
    # Helper to normalize value for comparison
    def normalize_value(val):
        if pd.isna(val):
            return ""
        return str(val).strip()

    for index, row in df.iterrows():
        errors = []

        # Treat rows where ALL A–G columns are blank as skipped
        if all(
            is_blank(row.get(col_name))
            for col_name in ["Product ID", "Product Name", "Type", "Category", "Status", "Stock Level", "Price"]
        ):
            skipped_rows += 1
            skipped_row_numbers.append(index + 2)  # Store row number (Excel row, +2 for header)
            # Don't validate or count as valid/invalid; just skip
            continue

        # --- Product ID (optional - will be auto-generated if blank, must be unique if provided) ---
        pid_raw = row.get("Product ID")
        if not is_blank(pid_raw):
            # Only validate if Product ID is provided
            pid_str = str(pid_raw).strip()
            
            # Check if Product ID already has "P" prefix (e.g., "P124")
            if pid_str.upper().startswith("P") and len(pid_str) > 1:
                # Extract numeric part after "P"
                numeric_part = pid_str[1:].strip()
                try:
                    pid_num = float(numeric_part)
                    if not pid_num.is_integer() or int(pid_num) <= 0:
                        errors.append("Product ID must be a valid number after 'P' prefix")
                    else:
                        # Normalize to uppercase (P124)
                        pid_str = f"P{int(pid_num)}"
                except (ValueError, TypeError):
                    errors.append("Product ID must be a valid number after 'P' prefix")
            else:
                # Try to convert to number (handles both "12" and "12.0" from Excel)
                try:
                    pid_num = float(pid_str)
                    # Check if it's a whole number (no decimal part)
                    if not pid_num.is_integer():
                        errors.append("Product ID must be a whole number")
                    elif int(pid_num) <= 0:
                        errors.append("Product ID must be greater than 0")
                    else:
                        # Prepend "P" to numeric Product ID (e.g., 124 becomes P124)
                        pid_str = f"P{int(pid_num)}"
                except (ValueError, TypeError):
                    # If conversion fails, it's not a valid number
                    errors.append("Product ID must be a whole number")
                    pid_str = None
            
            # Validate uniqueness if Product ID is valid
            if pid_str and not any("Product ID" in err for err in errors):
                # Check if Product ID already exists in the uploaded file
                if pid_str in seen_product_ids:
                    first_row = seen_product_ids[pid_str]
                    errors.append(f"Duplicate Product ID: Product ID {pid_str} already exists in row {first_row}")
                # Check if Product ID already exists in the database
                elif pid_str in existing_product_ids:
                    errors.append(f"Duplicate Product ID: Product ID {pid_str} already exists in the system")
                else:
                    seen_product_ids[pid_str] = index + 2  # Store the row number (Excel row, +2 for header)
            # Note: If Product ID is blank, it will be auto-generated during import (P101, P102, etc.)

        # --- Product Name (mandatory, alphabets + spaces, min length 3) ---
        pname_raw = row.get("Product Name")
        if is_blank(pname_raw):
            errors.append("Product Name is required")
        else:
            product_name = str(pname_raw).strip()
            if not re.fullmatch(r"^[A-Za-z ]+$", product_name):
                errors.append("Product Name must contain ONLY letters and spaces")
            elif len(product_name) < 3:
                errors.append("Product Name must be at least 3 characters")

        # --- Type (mandatory) ---
        type_raw = row.get("Type")
        if is_blank(type_raw):
            errors.append("Type is required")

        # --- Category (mandatory) ---
        category_raw = row.get("Category")
        if is_blank(category_raw):
            errors.append("Category is required")

        # --- Status (mandatory) ---
        status_raw = row.get("Status")
        if is_blank(status_raw):
            errors.append("Status is required")

        # --- Stock Level (mandatory, whole number >= 0) ---
        stock_raw = row.get("Stock Level")
        if is_blank(stock_raw):
            errors.append("Stock Level is required")
        else:
            try:
                stock_num = float(stock_raw)
                if not stock_num.is_integer():
                    errors.append("Stock Level must be a whole number")
                elif stock_num < 0:
                    errors.append("Stock Level must be 0 or greater")
            except Exception:
                errors.append("Stock Level must be a number")

        # --- Price (mandatory, number > 0) ---
        price_raw = row.get("Price")
        if is_blank(price_raw):
            errors.append("Price is required")
        else:
            try:
                price_num = float(price_raw)
                if price_num <= 0:
                    errors.append("Price must be greater than 0")
            except Exception:
                errors.append("Price must be a valid number")

        # --- Duplicate Row Check (same 6-field signature as save_product; normalized numbers) ---
        sig = _product_row_signature_from_excel(row)
        if sig in seen_rows:
            first_row = seen_rows[sig]
            errors.append(
                f"Duplicate row: This combination of Product Name, Type, Category, Status, Stock Level, and Price is identical to row {first_row}"
            )
        elif sig in existing_signatures:
            errors.append(
                "Duplicate product: this combination already exists in the system (same Name, Type, Category, Status, Stock Level, and Price)."
            )
        else:
            if any(sig[:4]) or sig[4] != 0 or sig[5] != 0.0:
                seen_rows[sig] = index + 2

        if errors:
            invalid_rows += 1
            error_details.append({
                "row": index + 2,  # +2 because header is row 1
                "errors": errors
            })
        else:
            valid_rows += 1

    return jsonify({
        "total_rows": len(df),
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "skipped_rows": skipped_rows,
        "skipped_row_numbers": skipped_row_numbers,  # List of row numbers that were skipped
        "error_details": error_details
    })

@app.route("/import-products-validated", methods=["POST"])
def import_products_validated():
    file = request.files.get("file")

    if not file:
        return jsonify({"success": False, "message": "No file uploaded"}), 400

    # BUG_003 / BUG_004: enforce correct Excel template here as well
    filename_lower = file.filename.lower()
    if not (filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls")):
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Invalid file format. Please upload the product Excel template (.xlsx or .xls).",
                }
            ),
            400,
        )

    try:
        df = pd.read_excel(file)
    except Exception:
        return jsonify({"success": False, "message": "Invalid Excel file"}), 400

    products = load_products()
    
    # Get existing product IDs to determine next sequence number
    existing_ids = {str(p.get("product_id", "")).strip() for p in products if p.get("product_id")}
    existing_sigs = {_product_signature_from_stored(p) for p in products}
    batch_sigs = set()
    skipped_duplicates = 0
    
    # Find the maximum numeric value from existing product IDs
    max_num = 0
    for pid in existing_ids:
        # Extract numeric part from formats like "P101", "101", "P-101", etc.
        match = re.search(r"(\d+)$", pid)
        if match:
            max_num = max(max_num, int(match.group(1)))
    
    added = 0
    current_sequence = 0

    for _, row in df.iterrows():
        errors = []

        # Product Name validation (same rules as /upload)
        raw_name = row.get("Product Name")
        if pd.isna(raw_name):
            errors.append("Product Name is required")
        else:
            product_name = str(raw_name).strip()
            if not re.fullmatch(r"^[A-Za-z ]+$", product_name):
                errors.append("Product Name must contain ONLY letters and spaces")
            elif len(product_name) < 3:
                errors.append("Product Name must be at least 3 characters")

        if errors:
            # Skip invalid rows, we only import validated rows
            continue

        # Skip duplicate rows (same 6-field signature as /upload and save_product)
        sig = _product_row_signature_from_excel(row)
        if sig in batch_sigs or sig in existing_sigs:
            skipped_duplicates += 1
            continue

        # Basic extraction of remaining fields
        raw_id = row.get("Product ID")
        product_id = "" if pd.isna(raw_id) else str(raw_id).strip()

        # Auto-generate Product ID if blank
        if not product_id:
            # Generate next Product ID based on existing products + already added in this batch
            current_sequence += 1
            product_id = f"P{max_num + current_sequence}"
        else:
            # If Product ID is provided, normalize it to "P###" format
            # Check if Product ID already has "P" prefix (case-insensitive)
            if product_id.upper().startswith("P") and len(product_id) > 1:
                # Extract numeric part after "P"
                numeric_part = product_id[1:].strip()
                try:
                    pid_num = float(numeric_part)
                    if pid_num.is_integer() and int(pid_num) > 0:
                        # Normalize to "P###" format (e.g., "P124", "p124" -> "P124")
                        product_id = f"P{int(pid_num)}"
                        num_val = int(pid_num)
                        max_num = max(max_num, num_val)
                    else:
                        # Invalid format, keep as-is but try to extract numeric part for max_num
                        match = re.search(r"(\d+)$", product_id)
                        if match:
                            num_val = int(match.group(1))
                            max_num = max(max_num, num_val)
                except (ValueError, TypeError):
                    # Invalid format, keep as-is but try to extract numeric part for max_num
                    match = re.search(r"(\d+)$", product_id)
                    if match:
                        num_val = int(match.group(1))
                        max_num = max(max_num, num_val)
            else:
                # Product ID doesn't start with "P" - check if it's purely numeric
                try:
                    # Try to convert to number (handles both "124" and "124.0" from Excel)
                    pid_num = float(product_id)
                    if pid_num.is_integer() and int(pid_num) > 0:
                        # Prepend "P" to numeric Product ID (e.g., 124 becomes P124)
                        product_id = f"P{int(pid_num)}"
                        num_val = int(pid_num)
                        max_num = max(max_num, num_val)
                    else:
                        # Not a valid positive integer, keep as-is but try to extract numeric part
                        match = re.search(r"(\d+)$", product_id)
                        if match:
                            num_val = int(match.group(1))
                            max_num = max(max_num, num_val)
                except (ValueError, TypeError):
                    # If conversion fails, it might be alphanumeric, extract numeric part if present
                    match = re.search(r"(\d+)$", product_id)
                    if match:
                        num_val = int(match.group(1))
                        max_num = max(max_num, num_val)
                    # If no numeric part found, keep as-is

        type_val = "" if pd.isna(row.get("Type")) else str(row.get("Type")).strip()
        category_val = "" if pd.isna(row.get("Category")) else str(row.get("Category")).strip()
        status_val = "" if pd.isna(row.get("Status")) else str(row.get("Status")).strip() or "Active"

        stock_raw = row.get("Stock Level")
        price_raw = row.get("Price")

        try:
            stock_level = int(stock_raw) if not pd.isna(stock_raw) else 0
        except Exception:
            stock_level = 0

        try:
            price = float(price_raw) if not pd.isna(price_raw) else 0.0
        except Exception:
            price = 0.0

        item = {
            "product_id": product_id,
            "product_name": product_name,
            "type": type_val,
            "category": category_val,
            "status": status_val,
            "stock_level": stock_level,
            "price": price,
        }

        products.append(item)
        added += 1
        batch_sigs.add(sig)
        existing_sigs.add(sig)

    save_products(products)

    msg = f"Successfully imported {added} product(s)"
    if skipped_duplicates:
        msg += f" ({skipped_duplicates} duplicate row(s) skipped)"
    return jsonify(
        {
            "success": True,
            "added": added,
            "skipped_duplicates": skipped_duplicates,
            "message": msg,
        }
    )

  
@app.route("/save-product", methods=["POST"])
def save_product():
    try:
        product_id = generate_product_id()

        # ------- safe converters so empty / bad values won't crash ------
        def to_int(val, default=0):
            try:
                return int(val)
            except (TypeError, ValueError):
                return default

        def to_float(val, default=0.0):
            try:
                return float(val)
            except (TypeError, ValueError):
                return default

        # -------------------- IMAGE HANDLING ----------------------------
        image = request.files.get("product_image")
        image_filename = None

        if image and image.filename:
            image_filename = secure_filename(image.filename)

            # ensure upload folder exists
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)

            image_path = os.path.join(UPLOAD_FOLDER, image_filename)
            image.save(image_path)

        # -------------------- PRODUCT OBJECT ----------------------------
        form = request.form

        product = {
            "product_id": str(product_id),
            "product_name": (form.get("product_name") or "").strip(),
            "type": (form.get("product_type") or "").strip(),
            "category": (form.get("category") or "").strip(),
            "status": (form.get("status") or "active").strip(),

            # fields used in table
            "stock_level": to_int(form.get("stock_level")),          
            "price": to_float(form.get("unit_price")),               

            # extra fields (for future use)
            "description": (form.get("description") or "").strip(),
            "sub_category": (form.get("sub_category") or "").strip(),
            "unit_price": (form.get("unit_price") or "").strip(),
            "discount": (form.get("discount") or "").strip(),
            "tax_code": (form.get("tax_code") or "").strip(),
            "quantity": (form.get("quantity") or "").strip(),
            "uom": (form.get("uom") or "").strip(),
            "reorder_level": (form.get("reorder_level") or "").strip(),
            "warehouse": (form.get("warehouse") or "").strip(),
            "size": (form.get("size") or "").strip(),
            "color": (form.get("color") or "").strip(),
            "weight": (form.get("weight") or "").strip(),
            "specifications": (form.get("specifications") or "").strip(),
            "related_products": (form.get("related_products") or "").strip(),
            "supplier": (form.get("supplier") or "").strip(),
            "product_usage": (form.get("product_usage") or "").strip(),
            "image": image_filename,
        }

        # -------------------- DUPLICATE VALIDATION -----------------------------
        products = load_products()
        
        # Normalize values for comparison (case-insensitive for text, exact for numbers)
        new_product_name = (product.get("product_name") or "").strip().lower()
        new_type = (product.get("type") or "").strip().lower()
        new_category = (product.get("category") or "").strip().lower()
        new_status = (product.get("status") or "").strip().lower()
        new_stock_level = product.get("stock_level", 0)
        new_price = product.get("price", 0.0)
        
        # Check 1: Duplicate product name (case-insensitive)
        for existing in products:
            existing_name = (existing.get("product_name") or "").strip().lower()
            if existing_name == new_product_name:
                return jsonify(
                    success=False, 
                    message=f"Product with name '{product.get('product_name')}' already exists. Please use a different product name."
                ), 409
        
        # Check 2: Duplicate combination (product_name + type + category + status + stock_level + price)
        for existing in products:
            existing_name = (existing.get("product_name") or "").strip().lower()
            existing_type = (existing.get("type") or "").strip().lower()
            existing_category = (existing.get("category") or "").strip().lower()
            existing_status = (existing.get("status") or "").strip().lower()
            existing_stock_level = existing.get("stock_level", 0)
            existing_price = existing.get("price", 0.0)
            
            # Compare all fields (case-insensitive for text, exact for numbers)
            if (existing_name == new_product_name and
                existing_type == new_type and
                existing_category == new_category and
                existing_status == new_status and
                existing_stock_level == new_stock_level and
                abs(existing_price - new_price) < 0.01):  # Float comparison with tolerance
                return jsonify(
                    success=False,
                    message=f"A product with the same combination (Name: '{product.get('product_name')}', Type: '{product.get('type')}', Category: '{product.get('category')}', Status: '{product.get('status')}', Stock Level: {new_stock_level}, Price: {new_price}) already exists."
                ), 409
        
        # -------------------- SAVE TO JSON -----------------------------
        products.append(product)
        save_products(products)

        return jsonify(success=True, product_id=product_id)

    except Exception as e:
        
        print("ERROR in /save-product:", e)
        return jsonify(success=False, message="Internal server error"), 500


# =========================================
# 6. MASTERS — Customer
# =========================================
@app.route("/customer")
def customer():
    user_email = session.get("user")
    if not user_email:
        if wants_json():
            return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
        return redirect(url_for("login", message="session_expired"))

    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"

    if wants_json():
        #============================from here  
        # customers_list = load_customer()
        customers_list = get_customers_from_db()
        #====================replace above
        return jsonify(
            {
                "success": True,
                "customers": customers_list,
                "total": len(customers_list),
                "current_user": {"email": user_email, "name": user_name, "role": user_role},
                "permissions": get_effective_permissions_for_session(),
            }
        ), 200

    return render_template(
        "customer.html",
        page="customer",
        section="masters",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
    )


@app.route("/import-customer", methods=["GET", "POST"])
def import_customer():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    # JSON metadata for Fetch/XHR on Import Customers page
    if request.method == "GET" and wants_json():
        return jsonify(
            {
                "success": True,
                "page": "import-customer",
                "current_user": {"email": user_email, "name": user_name},
            }
        ), 200

    return render_template(
        "import-customer.html",
        title="Import Customers - Stackly",
        page="customer",
        section="masters",
        user_email=user_email,
        user_name=user_name,
    )

# def load_customer():
#     if not os.path.exists(CUSTOMER_FILE):
#         return []
#     try:
#         with open(CUSTOMER_FILE, "r", encoding="utf-8") as f:
#             data = json.load(f)
#             return data if isinstance(data, list) else []
#     except Exception as e:
#         print(f"Error loading customer data: {e}")
#         return []


# def save_customer(customer):
#     """Save customer data to JSON file."""
#     with open(CUSTOMER_FILE, "w", encoding="utf-8") as f:
#         json.dump(customer, f, indent=2, ensure_ascii=False)


# Excel template + CSV import: canonical labels (case-insensitive match)
_CUSTOMER_TYPE_ALLOWED = (
    "Retail",
    "Wholesale",
    "Corporate",
    "Online",
    "Distributor",
    "Individual",
    "Business",
    "Organization",
)
_CUSTOMER_TYPE_BY_LOWER = {t.lower(): t for t in _CUSTOMER_TYPE_ALLOWED}


def normalize_customer_type(value):
    """Return canonical customer type, or None if empty/invalid. Accepts any casing (e.g. DISTRIBUTOR -> Distributor)."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    s = str(value).strip()
    if not s:
        return None
    return _CUSTOMER_TYPE_BY_LOWER.get(s.lower())


def _customer_row_signature_from_excel(row) -> tuple:
    """
    Canonical 7-tuple for customer duplicate detection (same logical row as upload duplicate check).
    Excludes Customer ID. Credit limit normalized like product price (Excel 100 vs 100.0).
    """
    def _s(v):
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except (TypeError, ValueError):
            pass
        return str(v).strip()

    name = _s(row.get("Name")).lower()
    company = _s(row.get("Company")).lower()
    ct_raw = row.get("Customer Type")
    if _s(ct_raw) == "":
        ctype_key = ""
    else:
        cn = normalize_customer_type(ct_raw)
        ctype_key = (cn or _s(ct_raw)).lower()

    email = _s(row.get("Email")).lower()
    status = _s(row.get("Status")).lower()
    cr = row.get("Credit Limit")
    try:
        credit = float(cr) if not pd.isna(cr) else 0.0
    except (TypeError, ValueError):
        credit = 0.0
    credit_r = round(credit + 1e-12, 2)
    city = _s(row.get("City")).lower()
    return (name, company, ctype_key, email, status, credit_r, city)


def _customer_signature_from_stored(c: dict) -> tuple:
    """Same tuple as _customer_row_signature_from_excel for rows in customer.json."""
    cr_raw = c.get("credit_limit", "")
    try:
        credit = float(str(cr_raw).replace(",", "").strip() or 0)
    except (TypeError, ValueError):
        credit = 0.0
    credit_r = round(credit + 1e-12, 2)
    ct = normalize_customer_type(c.get("customer_type") or c.get("company_type"))
    if ct is None:
        ctype_key = str(c.get("customer_type") or c.get("company_type") or "").strip().lower()
    else:
        ctype_key = ct.lower()
    return (
        str(c.get("name") or "").strip().lower(),
        str(c.get("company") or "").strip().lower(),
        ctype_key,
        str(c.get("email") or "").strip().lower(),
        str(c.get("status") or "").strip().lower(),
        credit_r,
        str(c.get("city") or "").strip().lower(),
    )


@app.route("/import-customers-validated", methods=["POST"])
def import_customers_validated():
    user_email = session.get("user")
    if not user_email:
        return jsonify(success=False, message="Session expired. Please login."), 401

    file = request.files.get("file")
    if not file or file.filename.strip() == "":
        return jsonify(success=False, message="No file uploaded"), 400

    # ---- Read file (csv/xlsx) ----
    try:
        if file.filename.lower().endswith(".csv"):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
    except Exception as e:
        return jsonify(success=False, message=f"Invalid file: {e}"), 400
    

    

    # ---- Required columns ----
    required_columns = [
        "Customer ID", "Name", "Company",
        "Customer Type", "Email", "Status",
        "Credit Limit", "City"
    ]
    
    
    for col in required_columns:
        if col not in df.columns:
            return jsonify(success=False, message=f"Missing column: {col}"), 400
#============================================from here
    # customers = load_customer()
    customers = get_customers_from_db()
    #============================================upto here replace
    # Normalize Customer IDs for comparison (lowercase for case-insensitive matching)
    existing_ids = {str(c.get("customer_id", "")).strip().lower() for c in customers if c.get("customer_id")}

    existing_emails = {
        str(c.get("email", "")).strip().lower()
        for c in customers
        if str(c.get("email", "")).strip() != ""
    }
    existing_sigs = {_customer_signature_from_stored(c) for c in customers}
    batch_sigs = set()
    skipped_duplicates = 0

    added = 0
    updated = 0
    skipped = 0
    errors = []
    seen_emails = {}
    seen_customer_ids_in_batch = {}  # Track Customer IDs in this import batch

    def is_valid_email(v):
        s = str(v or "").strip().lower()
        if not re.fullmatch(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$", s):
            return False
        allowed_tlds = {"com", "in", "net", "org", "co.in"}  # Match upload-customer route
        tld = s.rsplit(".", 1)[-1] if "." in s else ""
        return tld in allowed_tlds

    # ✅ Script detection
    def has_script(val):
        if pd.isna(val):
            return False
        s = str(val).lower()
        patterns = [
            "<script", "</script", "javascript:",
            "onerror=", "onload=", "<img", "<svg", "<iframe"
        ]
        return any(p in s for p in patterns)

    # ✅ SQL injection detection
    def has_sql_injection(val):
        if pd.isna(val):
            return False
        s = str(val).lower()
        patterns = [
            r"\bor\s+1\s*=\s*1\b",
            r"\bunion\s+select\b",
            r"\bdrop\s+table\b",
            r"\bdelete\s+from\b",
            r"\binsert\s+into\b",
            r"\bupdate\s+\w+\s+set\b",
            r"--",
            r";"
        ]
        return any(re.search(p, s) for p in patterns)

    def is_blank(v):
        return pd.isna(v) or str(v).strip() == ""
    
    # Find the maximum numeric value from existing customer IDs (same pattern as product import)
    max_num = 0
    for x in existing_ids:
        s = str(x).strip()
        # Extract numeric part from formats like "C101", "101", "C-101", etc.
        match = re.search(r"(\d+)$", s)
        if match:
            max_num = max(max_num, int(match.group(1)))
    
    current_sequence = 0
    
    def get_next_customer_id():
        # Generate next Customer ID based on existing customers + already added in this batch
        nonlocal current_sequence, max_num
        current_sequence += 1
        return f"C{max_num + current_sequence}"



    for idx, row in df.iterrows():
        row_no = idx + 2
        try:
            # ✅ Skip fully empty row
            if all(is_blank(row[col]) for col in required_columns):
                skipped += 1
                continue

            # ✅ Block Script/HTML injection
            if any(has_script(row.get(col, "")) for col in required_columns):
                skipped += 1
                errors.append(f"Row {row_no}: Script/HTML content detected")
                continue

            # ✅ Block SQL injection patterns
            if any(has_sql_injection(row.get(col, "")) for col in required_columns):
                skipped += 1
                errors.append(f"Row {row_no}: SQL injection pattern detected")
                continue

            # Extract fields first
            name = str(row.get("Name", "")).strip()
            company = str(row.get("Company", "")).strip()
            ctype_raw = row.get("Customer Type")
            ctype = ""
            email = str(row.get("Email", "")).strip().lower()
            status = str(row.get("Status", "")).strip()
            # Keep raw value for correct NaN/blank detection (must match /upload-customer logic)
            credit_val_raw = row.get("Credit Limit")
            city = str(row.get("City", "")).strip()
            
            # ---- Comprehensive validation during import (same as upload phase) ----
            # Apply same validation rules as /upload-customer to ensure only valid rows are imported
            validation_errors = []
            
            # Name validation (mandatory, letters + spaces, min length 3, max 40)
            if is_blank(name):
                validation_errors.append("Name is required")
            else:
                if not re.fullmatch(r"^[A-Za-z ]+$", name):
                    validation_errors.append("Name must contain ONLY letters and spaces")
                elif len(name) < 3:
                    validation_errors.append("Name must be at least 3 characters")
                elif len(name) > 40:
                    validation_errors.append("Name must not exceed 40 characters")
            
            # Company validation (mandatory, min 3, max 50)
            if is_blank(company):
                validation_errors.append("Company is required")
            else:
                if len(company) < 3:
                    validation_errors.append("Company must be at least 3 characters")
                elif len(company) > 50:
                    validation_errors.append("Company must not exceed 50 characters")
                elif not re.fullmatch(r"^[A-Za-z0-9 &.,'()\/-]+$", company):
                    validation_errors.append("Company contains invalid characters")
            
            # Customer Type validation (mandatory; case-insensitive vs Excel/CSV)
            if is_blank(ctype_raw):
                validation_errors.append("Customer Type is required")
            else:
                ctype_canon = normalize_customer_type(ctype_raw)
                if ctype_canon is None:
                    validation_errors.append(
                        f"Customer Type must be one of: {', '.join(_CUSTOMER_TYPE_ALLOWED)}"
                    )
                else:
                    ctype = ctype_canon
            
            # Email validation (mandatory, valid format, unique)
            if is_blank(email):
                validation_errors.append("Email is required")
            elif not is_valid_email(email):
                validation_errors.append("Invalid email format")
            elif len(email) > 50:
                validation_errors.append("Email must not exceed 50 characters")
            else:
                # Check for duplicate Email in uploaded file
                if email in seen_emails:
                    validation_errors.append(f"Duplicate Email: Email already exists in row {seen_emails[email]}")
                # Check if Email already exists in the database
                elif email in existing_emails:
                    validation_errors.append("Duplicate Email: Email already exists in the system")
                else:
                    seen_emails[email] = row_no
            
            # Status validation (mandatory)
            if is_blank(status):
                validation_errors.append("Status is required")
            elif status not in ["Active", "Inactive"]:
                validation_errors.append("Status must be 'Active' or 'Inactive'")
            
            # Credit Limit validation (mandatory, number >= 0, max 10,000,000)
            # Use the SAME logic as /upload-customer so that a row considered invalid
            # during validation is also treated as invalid during import.
            if is_blank(credit_val_raw):
                validation_errors.append("Credit Limit is required")
            else:
                try:
                    credit_limit_num = float(credit_val_raw)
                    if credit_limit_num < 0:
                        validation_errors.append("Credit Limit must be 0 or greater")
                    elif credit_limit_num > 10000000:
                        validation_errors.append("Credit Limit must not exceed 10,000,000")
                except (ValueError, TypeError):
                    validation_errors.append("Credit Limit must be a valid number")
            
            # City validation (mandatory, letters + spaces, min length 3, max 40)
            if is_blank(city):
                validation_errors.append("City is required")
            else:
                if not re.fullmatch(r"^[A-Za-z ]+$", city):
                    validation_errors.append("City must contain ONLY letters and spaces")
                elif len(city) < 3:
                    validation_errors.append("City must be at least 3 characters")
                elif len(city) > 40:
                    validation_errors.append("City must not exceed 40 characters")
            
            # Now that validation has passed, normalize Credit Limit for storage
            credit_raw = "" if pd.isna(credit_val_raw) else str(credit_val_raw).strip()

            # ---------------- Customer ID validation (add to validation_errors) ----------------
            cid_raw = row.get("Customer ID")
            customer_id = None
            customer_id_lower = None
            is_existing_id = False  # Track if this ID exists in original database
            
            if is_blank(cid_raw):
                # Auto-generate Customer ID if blank (same as product import)
                customer_id = get_next_customer_id()
                customer_id_lower = customer_id.lower()
            else:
                # If Customer ID is provided, normalize it (same pattern as product import)
                cid_str = str(cid_raw).strip()
                
                # Check if Customer ID already has "C" prefix (e.g., "C124")
                if cid_str.upper().startswith("C") and len(cid_str) > 1:
                    # Extract numeric part after "C"
                    numeric_part = cid_str[1:].strip()
                    try:
                        cid_num = float(numeric_part)
                        if cid_num.is_integer() and int(cid_num) > 0:
                            # Normalize to uppercase (C124)
                            customer_id = f"C{int(cid_num)}"
                        else:
                            validation_errors.append("Customer ID must be a valid number after 'C' prefix")
                    except (ValueError, TypeError):
                        validation_errors.append("Customer ID must be a valid number after 'C' prefix")
                else:
                    # Try to convert to number (handles both "12" and "12.0" from Excel)
                    try:
                        cid_num = float(cid_str)
                        # Check if it's a whole number (no decimal part)
                        if cid_num.is_integer() and int(cid_num) > 0:
                            # Prepend "C" to numeric Customer ID (e.g., 124 becomes C124)
                            customer_id = f"C{int(cid_num)}"
                        else:
                            validation_errors.append("Customer ID must be a whole number greater than 0")
                    except (ValueError, TypeError):
                        validation_errors.append("Customer ID must contain only numbers (digits 0-9).")
                
                # Only proceed with Customer ID checks if Customer ID is valid
                if customer_id:
                    # Update max_num if this Customer ID has a higher number
                    match = re.search(r"(\d+)$", customer_id)
                    if match:
                        num_val = int(match.group(1))
                        max_num = max(max_num, num_val)
                    
                    customer_id_lower = customer_id.lower()
                    
                    # Check for duplicate Customer ID in this import batch
                    if customer_id_lower in seen_customer_ids_in_batch:
                        validation_errors.append(f"Duplicate Customer ID: Customer ID '{customer_id}' already exists in row {seen_customer_ids_in_batch[customer_id_lower]}")
                    else:
                        seen_customer_ids_in_batch[customer_id_lower] = row_no
                    
                    # Check if Customer ID already exists in ORIGINAL database (before we start adding)
                    is_existing_id = customer_id_lower in existing_ids
            
            # Skip row if validation errors exist (only import valid rows)
            if validation_errors:
                skipped += 1
                errors.append(f"Row {row_no}: " + ", ".join(validation_errors))
                continue
            
            # At this point, all validations passed, so customer_id and customer_id_lower should be set
            if not customer_id or not customer_id_lower:
                # This shouldn't happen, but safety check
                skipped += 1
                errors.append(f"Row {row_no}: Customer ID validation failed")
                continue

            sig = _customer_row_signature_from_excel(row)
            if sig in batch_sigs:
                skipped += 1
                skipped_duplicates += 1
                continue
            if is_existing_id:
                c_match = next(
                    (c for c in customers if str(c.get("customer_id", "")).strip().lower() == customer_id_lower),
                    None,
                )
                my_sig = _customer_signature_from_stored(c_match) if c_match else None
                if sig in existing_sigs and my_sig is not None and sig != my_sig:
                    skipped += 1
                    skipped_duplicates += 1
                    continue
            else:
                if sig in existing_sigs:
                    skipped += 1
                    skipped_duplicates += 1
                    continue
            
            # Update existing_ids to track IDs in this batch (for duplicate prevention)
            if customer_id_lower not in existing_ids:
                existing_ids.add(customer_id_lower)

            # ---- Update if duplicate ID exists in original database ----
            if is_existing_id:
                c_match = next(
                    (c for c in customers if str(c.get("customer_id", "")).strip().lower() == customer_id_lower),
                    None,
                )
                old_sig = _customer_signature_from_stored(c_match) if c_match else None
                if old_sig is not None:
                    existing_sigs.discard(old_sig)
                for c in customers:
                    if str(c.get("customer_id", "")).strip().lower() == customer_id_lower:
                        c["name"] = name
                        c["company"] = company
                        c["customer_type"] = ctype
                        c["company_type"] = ctype
                        c["status"] = status
                        c["email"] = email
                        c["credit_limit"] = credit_raw
                        c["city"] = city
                        break
                existing_sigs.add(sig)
                batch_sigs.add(sig)
                updated += 1
                continue

            # ---- Add new customer ----
            customers.append({
                "customer_id": customer_id,
                "name": name,
                "company": company,
                "customer_type": ctype,
                "company_type": ctype,
                "status": status,
                "email": email,
                "credit_limit": credit_raw,
                "city": city,
                "sales_rep": ""
            })
            existing_ids.add(customer_id_lower)
            email_key = email.strip().lower()
            existing_emails.add(email_key)
            existing_sigs.add(sig)
            batch_sigs.add(sig)
            added += 1

        except Exception as e:
            errors.append(f"Row {row_no}: {e}")

    # save_customer(customers)

    return jsonify(
        success=True,
        added=added,
        updated=updated,
        skipped=skipped,
        skipped_duplicates=skipped_duplicates,
        error_details=errors,
    )

# =========================================
# ✅ API — Get All Customer(JSON)
# =========================================
@app.route("/api/customer", methods=["GET"])
def api_customer():
    """GET /api/customer — requires login (same as /api/products)."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    try:
        #==================================from here
        # customers = load_customer()
        customers = get_customers_from_db()
        #======================================upto replace
        
        # Ensure we return an array, even if empty
        if not isinstance(customers, list):
            customers = []
        
        # Get query parameters
        q = (request.args.get("q") or "").strip().lower()
        status = (request.args.get("status") or "").strip()
        ctype = (request.args.get("type") or "").strip()
        sales_rep = (request.args.get("sales_rep") or "").strip()
        
        page = int(request.args.get("page") or 1)
        page_size = int(request.args.get("page_size") or 10)
        
        # Filter function
        def match(c):
            # Search filter
            if q:
                hay = " ".join([
                    str(c.get("customer_id", "")),
                    str(c.get("name", "")),
                    str(c.get("company", "")),
                ]).lower()
                if q not in hay:
                    return False
            
            # Status filter
            if status:
                c_status = str(c.get("status") or "").lower()
                if c_status != status.lower():
                    return False
            
            # Customer Type filter
            if ctype:
                c_customer_type = str(c.get("customer_type") or c.get("company_type") or "").lower()
                if c_customer_type != ctype.lower():
                    return False
            
            # Sales Rep filter
            if sales_rep:
                c_rep = str(c.get("sales_rep") or "").lower()
                if c_rep != sales_rep.lower():
                    return False
            
            return True
        
        # Apply filters
        filtered = [c for c in customers if match(c)]
        
        # Extract meta data for dropdowns (from all customers, not filtered)
        statuses = sorted({str(c.get("status", "")) for c in customers if c.get("status")})
        types = sorted({str(c.get("customer_type") or c.get("company_type", "")) for c in customers if (c.get("customer_type") or c.get("company_type"))})
        sales_reps = sorted({str(c.get("sales_rep", "")) for c in customers if c.get("sales_rep")})
        
        # Pagination
        total_items = len(filtered)
        total_pages = max(1, (total_items + page_size - 1) // page_size)
        page = max(1, min(page, total_pages))
        
        start = (page - 1) * page_size
        end = start + page_size
        items = filtered[start:end]
        
        # Build response (same structure as /api/products)
        response_data = {
            "success": True,
            "data": {
                "items": items,
                "page": page,
                "total_pages": total_pages,
                "total_items": total_items,
                "meta": {
                    "statuses": statuses,
                    "types": types,
                    "sales_reps": sales_reps
                }
            }
        }

        response = jsonify(response_data)
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        response.headers['Access-Control-Allow-Origin'] = '*'
        return response
        
    except Exception as e:
        print(f"Error in /api/customer: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e),
            "data": {
                "items": [],
                "page": 1,
                "total_pages": 1,
                "total_items": 0,
                "meta": {
                    "statuses": [],
                    "types": [],
                    "sales_reps": []
                }
            }
        }), 500


@app.route("/api/customer/<customer_id>", methods=["GET"])
def api_get_customer(customer_id):
    """
    GET /api/customer/<customer_id>
    
    Returns a single customer by ID (same pattern as /api/products/<product_id>)
    """
    try:
        #===================================from here
        # customers = load_customer()
        customers = get_customers_from_db()
        #=======================================upto here replace
        customer = next((c for c in customers if str(c.get("customer_id")) == str(customer_id)), None)
        
        if not customer:
            return jsonify({
                "success": False,
                "message": "Customer not found"
            }), 404
        
        return jsonify({
            "success": True,
            "data": customer,
            "message": "Customer retrieved successfully"
        }), 200
        
    except Exception as e:
        print(f"Error in /api/customer/<customer_id>: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


# =========================================
# ✅ API: UPDATE CUSTOMER (PUT /api/customer/<id>) — JSON for Postman
# =========================================
@app.route("/api/customer/<customer_id>", methods=["PUT"])
def api_update_customer(customer_id):
    """Update customer by ID. Requires JSON body. Same pattern as PUT /api/products/<id>."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    if not request.is_json:
        return jsonify({"success": False, "message": "Content-Type must be application/json"}), 400

    data = request.get_json(silent=True) or {}
    if not data:
        return jsonify({"success": False, "message": "JSON body required"}), 400

    name = (data.get("name") or data.get("Name") or "").strip()
    if not name:
        return jsonify({"success": False, "message": "Customer name is required"}), 400

    try:
        # customers = load_customer()
        # if not isinstance(customers, list):
        #     customers = []

        # email = (data.get("email") or data.get("Email") or "").strip().lower()

        # for cust in customers:
        #     if str(cust.get("customer_id")) != str(customer_id) and (cust.get("email") or "").strip().lower() == email and email:
        #         return jsonify({"success": False, "message": "Duplicate email already exists."}), 409

        # found = False
        # for cust in customers:
        #     if str(cust.get("customer_id")) == str(customer_id):
        #         cust["name"] = (data.get("name") or data.get("Name") or cust.get("name", "")).strip()
        #         cust["company"] = (data.get("company") or data.get("Company") or cust.get("company", "")).strip()
        #         cust["customer_type"] = (data.get("customer_type") or data.get("Customer Type") or cust.get("customer_type", "")).strip()
        #         cust["company_type"] = cust["customer_type"]
        #         cust["email"] = (data.get("email") or data.get("Email") or cust.get("email", "")).strip().lower()
        #         cust["credit_limit"] = str(data.get("credit_limit") or data.get("Credit Limit") or cust.get("credit_limit", "")).strip()
        #         cust["status"] = (data.get("status") or data.get("Status") or cust.get("status", "")).strip()
        #         cust["city"] = (data.get("city") or data.get("City") or cust.get("city", "")).strip()
        #         found = True
        #         break

        # if not found:
        #     return jsonify({"success": False, "message": "Customer not found"}), 404

        # save_customer(customers)
        # updated = next((c for c in customers if str(c.get("customer_id")) == str(customer_id)), None)
        # return jsonify({"success": True, "message": "Customer updated", "customer": updated}), 200
        conn = get_db_connection()
        cur = conn.cursor()

        # Check customer exists
        cur.execute("SELECT customer_id FROM customers WHERE customer_id = %s", (customer_id,))
        if not cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"success": False, "message": "Customer not found"}), 404

        # Update customer
        cur.execute("""
            UPDATE customers
            SET name=%s,
                company=%s,
                customer_type=%s,
                company_type=%s,
                email=%s,
                credit_limit=%s,
                status=%s,
                city=%s
            WHERE customer_id=%s
        """, (
            (data.get("name") or "").strip(),
            (data.get("company") or "").strip(),
            (data.get("customer_type") or "").strip(),
            (data.get("customer_type") or "").strip(),  # same as company_type
            (data.get("email") or "").strip().lower(),
            float(data.get("credit_limit") or 0),
            (data.get("status") or "").strip(),
            (data.get("city") or "").strip(),
            customer_id
        ))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({
            "success": True,
            "message": "Customer updated successfully"
        }), 200
    #======================================upto here=============================
    except Exception as e:
        print(f"Error in api_update_customer: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Server error: {str(e)}"}), 500


# =========================================
# ✅ API: DELETE CUSTOMER (DELETE /api/customer/<id>) — JSON for Postman
# =========================================
@app.route("/api/customer/<customer_id>", methods=["DELETE"])
def api_delete_customer(customer_id):
    """Delete customer by ID."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    # customers = load_customer()
    # new_list = [c for c in customers if str(c.get("customer_id")) != str(customer_id)]

    # if len(new_list) == len(customers):
    #     return jsonify({"success": False, "message": "Customer not found"}), 404

    # save_customer(new_list)
    # return jsonify({"success": True, "message": "Customer deleted successfully"}), 200
    conn = get_db_connection()
    cur = conn.cursor()

    # Check exists
    cur.execute("SELECT customer_id FROM customers WHERE customer_id = %s", (customer_id,))
    if not cur.fetchone():
        cur.close()
        conn.close()
        return jsonify({"success": False, "message": "Customer not found"}), 404

    # Delete
    cur.execute("DELETE FROM customers WHERE customer_id = %s", (customer_id,))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        "success": True,
        "message": "Customer deleted successfully"
    }), 200
##############################################upto here

@app.route("/update-customer/<customer_id>", methods=["POST"])
def update_customer(customer_id):
    # Require login (same as Edit Product / api_update_customer)
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    #======================from here
    # customer = load_customer()
    customers = get_customers_from_db()
    #============================upto replace
    data = request.get_json(silent=True) or {}

    name = (data.get("name") or "").strip()
    company = (data.get("company") or "").strip()
    # accept both keys, but prefer "customer_type"
    customer_type = (
        (data.get("customer_type") or "").strip()
        or (data.get("company_type") or "").strip()
    )
    email = (data.get("email") or "").strip().lower()
    credit_limit = (data.get("credit_limit") or "").strip()
    status = (data.get("status") or "").strip()
    city = (data.get("city") or "").strip()

 # ✅ DUPLICATE EMAIL CHECK (exclude current customer)
    for cust in customer:
        if str(cust.get("customer_id")) != str(customer_id):
            existing_email = (cust.get("email") or "").strip().lower()
            if email and existing_email == email:
                return jsonify({
                    "success": False,
                    "message": "Duplicate email! This email already exists."
                }), 409

    # find the matching customer
    found = False
    for cust in customer:
        if str(cust.get("customer_id")) == str(customer_id):
            cust["name"] = name
            cust["company"] = company
            # update both keys so data is consistent
            cust["customer_type"] = customer_type
            cust["company_type"] = customer_type
            cust["email"] = email
            cust["credit_limit"] = credit_limit
            cust["status"] = status
            cust["city"] = city
            found = True
            break

    if not found:
        return jsonify({"success": False, "message": "Customer not found"}), 404

    # save_customer(customer)
    return jsonify({"success": True, "message": "Customer updated"}), 200

#==============================================from here
# =========================================
# ✅ DELETE CUSTOMER (POST)
# =========================================
@app.route("/delete-customer/<cust_id>", methods=["POST"])
def delete_customer(cust_id):
   conn = get_db_connection()
   cur = conn.cursor()

   cur.execute("SELECT customer_id FROM customers WHERE customer_id=%s", (cust_id,))
   if not cur.fetchone():
       cur.close()
       conn.close()
       return jsonify({"ok": False, "message": "Customer not found"}), 404

   cur.execute("DELETE FROM customers WHERE customer_id=%s", (cust_id,))

   conn.commit()
   cur.close()
   conn.close()

   return jsonify({"ok": True, "message": "Customer deleted"})
#==============================================upto here fully replace

# =========================================
# 6. MASTERS — Customer — Add New Customer
# =========================================
@app.route("/addnew-customer")
def addnew_customer():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "customer-addnew-customer.html",
        title="Add New Customer - Stackly",
        page="customer",
        section="masters",
        user_email=user_email,
        user_name=user_name,
    )


# =========================================
# ✅ API — Get All Customers (Add New Customer)
# =========================================
@app.route("/api/customers", methods=["GET"])
def get_customers():
    # Load customers from customer.json (same file used for Customer Master)
    customers = get_customers_from_db()
    return jsonify(customers)


# =========================================
# ✅ API — Save Customer (Add New Customer)
# =========================================
@app.route("/api/customers", methods=["POST"])
def create_customer():
    data = request.get_json()

    if not data:
        return jsonify({"error": "No data received"}), 400

    gst_id = (data.get("gstNumber") or "").strip().upper()

    try:
        ######################################################from here cust########################
        # Load existing customers from customer.json
        customers = get_customers_from_db()
######################################################upto here cust########################
        # DUPLICATE GST CHECK
        if gst_id:
            for c in customers:
                if (c.get("gstNumber") or "").strip().upper() == gst_id:
                    return jsonify({
                        "error": "GST/Tax ID already exists"
                    }), 409  # Conflict

        # Generate customer ID in format C101, C102, etc.
        customer_id = generate_customer_id_for_master()
        
        # Transform form data to match customer.json format
        first_name = (data.get("firstName") or "").strip()
        last_name = (data.get("lastName") or "").strip()
        full_name = (first_name + " " + last_name).strip()
        
        # For company field: use name if customer type is Individual, otherwise leave empty
        # (The form doesn't have a company field, so we'll use name as company for consistency)
        company_name = full_name if data.get("customerType") == "Individual" else (data.get("company") or "")
        
        customer_data = {
            "customer_id": customer_id,
            "name": full_name or "Unknown",
            "company": company_name or full_name,
            "customer_type": data.get("customerType") or "",
            "status": data.get("customerStatus") or "Active",
            "email": (data.get("email") or "").strip().lower(),
            "credit_limit": str(data.get("creditLimit") or "0"),
            "city": data.get("city") or "",
            "sales_rep": (data.get("salesRep") if data.get("salesRep") != "custom" else data.get("salesRepCustom")) or "",
            "company_type": data.get("customerType") or "",
            "phone": data.get("phoneNumber") or "",
            "gstNumber": gst_id,
            "address": data.get("address") or "",
            "street": data.get("street") or "",
            "state": data.get("state") or "",
            "zipCode": data.get("zipCode") or "",
            "country": data.get("country") or "",
            "billingAddress": data.get("billingAddress") or "",
            "shippingAddress": data.get("shippingAddress") or "",
            "paymentTerms": (data.get("paymentTerms") if data.get("paymentTerms") != "custom" else data.get("paymentTermsCustom")) or "",
            "creditTerm": (data.get("creditTerm") if data.get("creditTerm") != "custom" else data.get("creditTermCustom")) or "",
            "availableLimit": str(data.get("availableLimit") or "0")
        }
######################################################from here cust########################
        # ✅ SAVE CUSTOMER to customer.json
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO customers (
                customer_id, name, company, customer_type,
                status, email, credit_limit, city,
                sales_rep, company_type
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            customer_id,
            customer_data["name"],
            customer_data["company"],
            customer_data["customer_type"],
            customer_data["status"],
            customer_data["email"],
            float(customer_data["credit_limit"] or 0),
            customer_data["city"],
            customer_data["sales_rep"],
            customer_data["company_type"]
        ))

        conn.commit()
        cur.close()
        conn.close()
        

        return jsonify({
            "message": "Customer saved successfully",
            "customerId": customer_id
        }), 201
    
    #############################upto this cust################################3

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Error saving customer: {e}")
        return jsonify({"error": str(e)}), 500


# =========================================
# ✅ ID GENERATION
# =========================================
def generate_customer_id():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT customer_id FROM customers ORDER BY id DESC LIMIT 1")
    row = cur.fetchone()

    cur.close()
    conn.close()

    if row and row[0]:
        try:
            num = int(row[0].replace("C", ""))
            return f"C{str(num + 1).zfill(3)}"   # ✅ zero padding
        except:
            return "C001"
    else:
        return "C001"


# =========================================from here to full code replace with this
# ✅ ID GENERATION FOR CUSTOMER MASTER (C101, C102, etc.)
# =========================================
def generate_customer_id_for_master():
    customers = get_customers_from_db()

    ids = []
    for c in customers:
        cust_id = c.get("customer_id")   # ✅ DEFINE HERE

        if cust_id and str(cust_id).startswith('C'):
            try:
                num = int(str(cust_id).replace("C", ""))
                ids.append(num)
            except:
                pass

    if ids:
        new_id = max(ids) + 1
    else:
        new_id = 1

    return f"C{str(new_id).zfill(3)}"


@app.route('/api/customers/new-id', methods=['GET'])
def get_new_customer_id():
    return jsonify({"customerId": generate_customer_id()})


@app.route('/api/customers/master-id', methods=['GET'])
def get_master_customer_id():
    return jsonify({"customerId": generate_customer_id_for_master()})


@app.route("/upload-customer", methods=["POST"])
def upload_customer_file():
    file = request.files.get("file")

    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        if file.filename.endswith(".csv"):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
    except Exception:
        return jsonify({"error": "Invalid Excel/CSV file"}), 400

    # Check if file is empty (no data rows, only headers)
    if df.empty or len(df) == 0:
        return jsonify({
            "error": "No data found",
            "message": "The uploaded file contains no data. Please ensure the file has at least one row of customer data."
        }), 400

    valid_rows = 0
    invalid_rows = 0
    skipped_rows = 0
    error_details = []
    skipped_row_numbers = []  # Track which rows were skipped (completely blank)
    
    # Load existing customers to check against database
    #================================from here
    # existing_customers = load_customer()
    existing_customers = get_customers_from_db()
    #==============upto replace


    # Normalize Customer IDs for comparison (same pattern as product import)
    existing_customer_ids = {str(c.get("customer_id", "")).strip().lower() for c in existing_customers if c.get("customer_id")}
    
    # Track Customer IDs for uniqueness validation within uploaded file
    seen_customer_ids = {}  # key: Customer ID (as string, lowercase), value: first row number
    
    # Track Emails for uniqueness validation within uploaded file
    seen_emails = {}  # key: Email (as string, lowercase), value: first row number
    
    # Load existing emails from database for uniqueness check
    existing_emails = {
        str(c.get("email", "")).strip().lower()
        for c in existing_customers
        if str(c.get("email", "")).strip() != ""
    }
    existing_customer_signatures = {_customer_signature_from_stored(c) for c in existing_customers}
    
    # Track seen row combinations for duplicate detection (excluding Customer ID)
    seen_rows = {}  # signature -> first Excel row number

    # Helper to check blank (NaN or empty/whitespace)
    def is_blank(val):
        if pd.isna(val):
            return True
        return str(val).strip() == ""
    
    # Helper to normalize value for comparison
    def normalize_value(val):
        if pd.isna(val):
            return ""
        return str(val).strip()
    
    

    # Security validation helpers
    def has_script(v):
        if pd.isna(v):
            return False
        s = str(v).lower()
        patterns = [
            "<script", "</script", "javascript:", "onerror=", "onload=",
            "<img", "<svg", "<iframe"
        ]
        return any(p in s for p in patterns)
    
    def has_sql_injection(v):
        if pd.isna(v):
            return False
        s = str(v).lower()
        patterns = [
            r"\bor\s+1\s*=\s*1\b",      # or 1=1
            r"\bunion\s+select\b",      # union select
            r"\bdrop\s+table\b",        # drop table
            r"\bdelete\s+from\b",       # delete from
            r"\binsert\s+into\b",       # insert into
            r"\bupdate\s+\w+\s+set\b",  # update x set
            r"--",                      # SQL comment
            r";"                        # multiple statements
        ]
        return any(re.search(p, s) for p in patterns)

    # Required columns for validation
    required_columns = [
        "Customer ID", "Name", "Company",
        "Customer Type", "Email", "Status",
        "Credit Limit", "City"
    ]

    # Check if all required columns exist
    for col in required_columns:
        if col not in df.columns:
            return jsonify({
                "total_rows": len(df),
                "valid_rows": 0,
                "invalid_rows": len(df),
                "skipped_rows": 0,
                "skipped_row_numbers": [],
                "error_details": [{"row": 0, "errors": [f"Missing column: {col}"]}]
            }), 400

    # ---------------- ROW VALIDATION ----------------
    for index, row in df.iterrows():
        errors = []
        row_no = index + 2  # Excel row number (header is row 1)

        # Treat rows where ALL required columns are blank as skipped
        if all(
            is_blank(row.get(col_name))
            for col_name in required_columns
        ):
            skipped_rows += 1
            skipped_row_numbers.append(row_no)  # Store row number (Excel row, +2 for header)
            # Don't validate or count as valid/invalid; just skip
            continue

        # Block Script/HTML injection in ANY column
        if any(has_script(row.get(col)) for col in required_columns):
            invalid_rows += 1
            error_details.append({
                "row": row_no,
                "errors": ["Script/HTML content detected"]
            })
            continue
        
        # Block SQL injection patterns in ANY column
        if any(has_sql_injection(row.get(col)) for col in required_columns):
            invalid_rows += 1
            error_details.append({
                "row": row_no,
                "errors": ["SQL injection pattern detected"]
            })
            continue

                
        # ---------------- Customer ID (optional - will be auto-generated if blank, must be unique if provided) ----------------
        # Same validation pattern as Product ID in product import
        cid_raw = row.get("Customer ID")
        customer_id = None
        
        if not is_blank(cid_raw):
            # Only validate if Customer ID is provided
            cid_str = str(cid_raw).strip()
            
            # Check if Customer ID already has "C" prefix (e.g., "C124")
            if cid_str.upper().startswith("C") and len(cid_str) > 1:
                # Extract numeric part after "C"
                numeric_part = cid_str[1:].strip()
                try:
                    cid_num = float(numeric_part)
                    if not cid_num.is_integer() or int(cid_num) <= 0:
                        errors.append("Customer ID must be a valid number after 'C' prefix")
                    else:
                        # Normalize to uppercase (C124)
                        cid_str = f"C{int(cid_num)}"
                except (ValueError, TypeError):
                    errors.append("Customer ID must be a valid number after 'C' prefix")
            else:
                # Try to convert to number (handles both "12" and "12.0" from Excel)
                try:
                    cid_num = float(cid_str)
                    # Check if it's a whole number (no decimal part)
                    if not cid_num.is_integer():
                        errors.append("Customer ID must be a whole number")
                    elif int(cid_num) <= 0:
                        errors.append("Customer ID must be greater than 0")
                    else:
                        # Prepend "C" to numeric Customer ID (e.g., 124 becomes C124)
                        cid_str = f"C{int(cid_num)}"
                except (ValueError, TypeError):
                    # If conversion fails, it's not a valid number
                    errors.append("Customer ID must be a whole number")
                    cid_str = None
            
            # Validate uniqueness if Customer ID is valid
            if cid_str and not any("Customer ID" in err for err in errors):
                # Normalize to lowercase for comparison
                cid_normalized = cid_str.lower()
                
                # Check if Customer ID already exists in the uploaded file
                if cid_normalized in seen_customer_ids:
                    first_row = seen_customer_ids[cid_normalized]
                    errors.append(f"Duplicate Customer ID: Customer ID {cid_str} already exists in row {first_row}")
                # Check if Customer ID already exists in the database
                elif cid_normalized in existing_customer_ids:
                    errors.append(f"Duplicate Customer ID: Customer ID {cid_str} already exists in the system")
                else:
                    seen_customer_ids[cid_normalized] = row_no  # Store the row number (Excel row, +2 for header)
                    customer_id = cid_str
            # Note: If Customer ID is blank, it will be auto-generated during import (C101, C102, etc.)





        # ---------------- Name (mandatory, letters + spaces, min length 3) ----------------
        name_raw = row.get("Name")
        if is_blank(name_raw):
            errors.append("Name is required")
        else:
            name = str(name_raw).strip()
            if not re.fullmatch(r"^[A-Za-z ]+$", name):
                errors.append("Name must contain ONLY letters and spaces")
            elif len(name) < 3:
                errors.append("Name must be at least 3 characters")
            elif len(name) > 40:
                errors.append("Name must not exceed 40 characters")

        # ---------------- Company (mandatory) ----------------
        company_raw = row.get("Company")
        if is_blank(company_raw):
            errors.append("Company is required")
        else:
            company = str(company_raw).strip()
            if len(company) < 3:
                errors.append("Company must be at least 3 characters")
            elif len(company) > 50:
                errors.append("Company must not exceed 50 characters")
            # Allow letters, numbers, spaces, and common business symbols
            if not re.fullmatch(r"^[A-Za-z0-9 &.,'()\/-]+$", company):
                errors.append("Company contains invalid characters")

        # ---------------- Customer Type (mandatory; case-insensitive vs Excel/CSV) ----------------
        customer_type_raw = row.get("Customer Type")
        if is_blank(customer_type_raw):
            errors.append("Customer Type is required")
        else:
            customer_type_canon = normalize_customer_type(customer_type_raw)
            if customer_type_canon is None:
                errors.append(
                    f"Customer Type must be one of: {', '.join(_CUSTOMER_TYPE_ALLOWED)}"
                )

        # ---------------- Email (mandatory, valid format, must be unique) ----------------
        email_raw = row.get("Email")
        if is_blank(email_raw):
            errors.append("Email is required")
        else:
            email_str = str(email_raw).strip().lower()
            
            # Basic email format check
            basic_ok = re.fullmatch(
                r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$",
                email_str
            )
            
            # Allowed TLDs (matching product import pattern)
            allowed_tlds = {"com", "in", "net", "org", "co.in"}
            tld = email_str.rsplit(".", 1)[-1] if "." in email_str else ""
            
            if not basic_ok:
                errors.append("Invalid email format")
            elif tld not in allowed_tlds:
                errors.append(f"Email domain must end with one of: {', '.join(allowed_tlds)}")
            elif len(email_str) > 50:
                errors.append("Email must not exceed 50 characters")
            else:
                # Check for duplicate Email in uploaded file (same pattern as Product ID uniqueness)
                if email_str in seen_emails:
                    first_row = seen_emails[email_str]
                    errors.append(f"Duplicate Email: Email already exists in row {first_row}")
                # Check if Email already exists in the database
                elif email_str in existing_emails:
                    errors.append(f"Duplicate Email: Email already exists in the system")
                else:
                    seen_emails[email_str] = row_no  # Store the row number (Excel row, +2 for header)

        # ---------------- Status (mandatory) ----------------
        status_raw = row.get("Status")
        if is_blank(status_raw):
            errors.append("Status is required")
        else:
            status = str(status_raw).strip()
            if status not in ["Active", "Inactive"]:
                errors.append("Status must be 'Active' or 'Inactive'")

        # ---------------- Credit Limit (mandatory, number >= 0) ----------------
        credit_limit_raw = row.get("Credit Limit")
        if is_blank(credit_limit_raw):
            errors.append("Credit Limit is required")
        else:
            try:
                credit_limit_num = float(credit_limit_raw)
                if credit_limit_num < 0:
                    errors.append("Credit Limit must be 0 or greater")
                elif credit_limit_num > 10000000:
                    errors.append("Credit Limit must not exceed 10,000,000")
            except (ValueError, TypeError):
                errors.append("Credit Limit must be a valid number")

        # ---------------- City (mandatory, letters + spaces, min length 3) ----------------
        city_raw = row.get("City")
        if is_blank(city_raw):
            errors.append("City is required")
        else:
            city = str(city_raw).strip()
            if not re.fullmatch(r"^[A-Za-z ]+$", city):
                errors.append("City must contain ONLY letters and spaces")
            elif len(city) < 3:
                errors.append("City must be at least 3 characters")
            elif len(city) > 40:
                errors.append("City must not exceed 40 characters")

        # --- Duplicate row / duplicate-in-system (normalized signature; credit limit like product price) ---
        sig = _customer_row_signature_from_excel(row)
        if sig in seen_rows:
            first_row = seen_rows[sig]
            errors.append(
                f"Duplicate row: This combination of Name, Company, Customer Type, Email, Status, Credit Limit, and City is identical to row {first_row}"
            )
        elif sig in existing_customer_signatures:
            errors.append(
                "Duplicate customer: this combination already exists in the system (same Name, Company, Type, Email, Status, Credit Limit, and City)."
            )
        else:
            if sig != ("", "", "", "", "", 0.0, ""):
                seen_rows[sig] = row_no

        if errors:
            invalid_rows += 1
            error_details.append({
                "row": row_no,  # +2 because header is row 1
                "errors": errors
            })
        else:
            valid_rows += 1

    return jsonify({
        "total_rows": len(df),
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "skipped_rows": skipped_rows,
        "skipped_row_numbers": skipped_row_numbers,  # List of row numbers that were skipped
        "error_details": error_details
    })

# =========================================
# ✅ CUSTOM DROPDOWNS
# =========================================
CUSTOM_DROPDOWNS = os.path.join(app.root_path, "custom_dropdowns.json")


@app.route("/api/custom-dropdowns", methods=["GET"])
def get_custom_dropdowns():
    if not os.path.exists(CUSTOM_DROPDOWNS):
        return jsonify({
            "paymentTerms": [],
            "creditTerms": [],
            "salesReps": []
        })

    with open(CUSTOM_DROPDOWNS, "r", encoding="utf-8") as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError:
            data = {}

    return jsonify({
        "paymentTerms": data.get("paymentTerms", []),
        "creditTerms": data.get("creditTerms", []),
        "salesReps": data.get("salesReps", [])
    })


@app.route("/api/custom-dropdowns", methods=["POST"])
def save_custom_dropdown():
    data = request.get_json()
    field = data.get("field")
    value = data.get("value")

    if not field or not value:
        return jsonify({"error": "Invalid data"}), 400

    # default structure
    dropdowns = {
        "paymentTerms": [],
        "creditTerms": [],
        "salesReps": []
    }

    # load existing data safely
    if os.path.exists(CUSTOM_DROPDOWNS):
        with open(CUSTOM_DROPDOWNS, "r", encoding="utf-8") as f:
            try:
                existing = json.load(f)
                dropdowns.update(existing)  # 🔥 merge safely
            except json.JSONDecodeError:
                pass

    # ensure key exists (important)
    if field not in dropdowns:
        dropdowns[field] = []

    # avoid duplicates (case-insensitive)
    if value.lower() not in [v.lower() for v in dropdowns[field]]:
        dropdowns[field].append(value)

    with open(CUSTOM_DROPDOWNS, "w", encoding="utf-8") as f:
        json.dump(dropdowns, f, indent=2, ensure_ascii=False)

    return jsonify({"success": True})


@app.route("/crm")
def crm():
    return render_template("crm.html", page="crm")


# =========================================
# 1. ROOT & AUTH — OTP APIs
# =========================================
@app.route("/send_otp", methods=["POST"])
def send_otp():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()

    if not email:
        return jsonify(success=False, message="Email is required"), 400

    if len(email) > MAX_EMAIL_LENGTH:
        return jsonify(success=False, message="Email is too long. Max 50 characters."), 400

    if not EMAIL_REGEX.match(email):
        return jsonify(success=False, message="Enter a valid email address like name@gmail.com or name@yahoo.com"), 400

    # =========================================
    # BUG_008 — Simple per‑email OTP rate limit
    # Return HTTP 429 if too many OTPs are requested in a short window.
    # =========================================
    now = time.time()
    history = OTP_SEND_COUNT.get(email, [])
    # Keep only recent timestamps inside the window
    history = [ts for ts in history if now - ts <= OTP_WINDOW_SECONDS]
    if len(history) >= MAX_OTP_SENDS:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Too many OTP requests. Please wait a few minutes before trying again.",
                }
            ),
            429,
        )
    history.append(now)
    OTP_SEND_COUNT[email] = history

    otp = generate_otp()
    print("DEBUG OTP for", email, "=", otp)

    save_otp_in_db(email, otp)

    try:
        send_otp_email(email, otp)
    except Exception as e:
        print("Error sending OTP:", e)
        return jsonify(success=False, message="Error sending OTP. Try again."), 500

    return jsonify(success=True, message="OTP sent successfully!")


@app.route("/verify_otp", methods=["POST"])
def verify_otp():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()
    otp = (data.get("otp") or "").strip()

    if verify_otp_in_db(email, otp):
        return jsonify({"success": True, "message": "OTP verified successfully!"}), 200
    return jsonify({"success": False, "message": "Invalid or expired OTP"}), 400


# =========================================
# 1. ROOT & AUTH — Signup API
# =========================================
@app.route("/signup", methods=["POST"])
def signup():
    data = request.get_json() or {}

    name = (data.get("name") or "").strip()
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip().lower()
    password = (data.get("password") or "").strip()

    missing = []
    if not name:
        missing.append("Name")
    if not phone:
        missing.append("Phone number")
    if not email:
        missing.append("Email")
    if not password:
        missing.append("Password")

    if missing:
        if len(missing) == 1:
            msg = f"⚠️ {missing[0]} is required"
        else:
            msg = "⚠️ " + ", ".join(missing[:-1]) + f" and {missing[-1]} are required"
        return jsonify({"success": False, "message": msg}), 400

    if not NAME_REGEX.match(name):
        return jsonify({"success": False, "message": "⚠️ Name must be 3–20 letters only"}), 400

    if not re.match(r"^\+\d{8,15}$", phone):
       return jsonify({"success": False, "message": "Enter valid phone with country code like +91XXXXXXXXXX"}), 400

    if len(email) > MAX_EMAIL_LENGTH:
        return jsonify({"success": False, "message": "⚠️ Email is too long (max 50 characters)"}), 400

    if not EMAIL_REGEX.match(email):
        return jsonify({"success": False, "message": "⚠️ Enter a valid email address (like name@gmail.com or name@outlook.com)"}), 400

    if not is_email_otp_verified(email):
        return jsonify({
            "success": False,
            "message": "⚠️ Please verify OTP for this email before signing up."
        }), 400

    users = load_users()
    if any((u.get("email") or "").strip().lower() == email for u in users):
        return jsonify({"success": False, "message": "⚠️ User already exists"}), 409

    users.append({
        "name": name,
        "phone": phone,
        "email": email,
        "password": password,
        "role": "User",   # ✅ assigned automatically
    })
    save_users(users)

    otps = load_otps()
    otps.pop(email, None)
    save_otps(otps)

    send_email(email, "Welcome!", f"Hello {name}, your account has been created successfully!")

    return jsonify({"success": True, "message": "🎉 Signup successful!"}), 200


# =========================================
# 1. ROOT & AUTH — Login API
# =========================================
@app.route("/login", methods=["POST"])
def login_post():
    """
    Login endpoint (JSON API).
    Handles:
      - basic validation
      - account lockout after repeated failures
      - session creation on success
    """
    try:
        # Safely get JSON data
        if not request.is_json:
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "Invalid request format. Expected JSON.",
                    }
                ),
                400,
            )

        data = request.get_json(silent=True) or {}
        email = (data.get("email") or "").strip().lower()
        password = (data.get("password") or "").strip()
        remember_me = data.get("rememberMe", False)

        # Validate input
        if not email:
            return jsonify({"success": False, "message": "Email is required"}), 400
        if not password:
            return jsonify({"success": False, "message": "Password is required"}), 400

        # Load users and failed attempts with error handling
        
        # ✅ DB LOGIN
        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            db_user = None
            try:
                cursor.execute(
                    """
                    SELECT name, role, password, branch, department FROM users
                    WHERE email = %s
                    LIMIT 1
                    """,
                    (email,),
                )
                db_user = cursor.fetchone()
            except Exception:
                cursor.execute(
                    """
                    SELECT name, role, password FROM users
                    WHERE email = %s
                    LIMIT 1
                    """,
                    (email,),
                )
                db_user = cursor.fetchone()

            if not db_user:
                try:
                    cursor.execute(
                        """
                        SELECT name, role, password, branch, department FROM users
                        WHERE LOWER(email) = LOWER(%s)
                        LIMIT 1
                        """,
                        (email,),
                    )
                    db_user = cursor.fetchone()
                except Exception:
                    cursor.execute(
                        """
                        SELECT name, role, password FROM users
                        WHERE LOWER(email) = LOWER(%s)
                        LIMIT 1
                        """,
                        (email,),
                    )
                    db_user = cursor.fetchone()

            cursor.close()
            conn.close()

        except Exception as e:
            print("❌ DB error:", e)
            return jsonify({"success": False, "message": "Database error"}), 500


        # ❌ User not found
        if not db_user:
            return jsonify({"success": False, "message": "User not found"}), 404

        db_name = db_user[0]
        db_role = db_user[1]
        db_password = db_user[2]
        db_branch = db_user[3] if len(db_user) > 3 else ""
        db_department = db_user[4] if len(db_user) > 4 else ""

        # ❌ Password wrong
        if db_password != password:
            return jsonify({"success": False, "message": "Incorrect password"}), 401

        # ✅ Login success (store branch/department for roles.json RBAC matching)
        session.permanent = bool(remember_me)
        session["user"] = email
        session["role"] = db_role
        session["branch"] = (db_branch or "").strip() or "Main Branch"
        session["department"] = (db_department or "").strip()
        session["last_active"] = time.time()

        return jsonify({"success": True, "message": "Login successful"}), 200

        # Check if account is locked
        info = failed_attempts.get(email, {})
        if "locked_until" in info and time.time() < info["locked_until"]:
            remaining = int(info["locked_until"] - time.time())
            return (
                jsonify(
                    {
                        "success": False,
                        "message": f"Account locked. Try again in {remaining}s.",
                    }
                ),
                403,
            )

        # Verify password
        if user.get("password") != password:
            info.setdefault("count", 0)
            info["count"] += 1

            if info["count"] >= LOCKOUT_THRESHOLD:
                info["locked_until"] = time.time() + LOCKOUT_DURATION
                failed_attempts[email] = info
                try:
                    save_failed_attempts(failed_attempts)
                except Exception as e:  # pragma: no cover - defensive
                    print(f"❌ Error saving failed attempts: {e}")
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": f"Too many failed attempts. Locked for {LOCKOUT_DURATION//60} min.",
                        }
                    ),
                    403,
                )

            failed_attempts[email] = info
            try:
                save_failed_attempts(failed_attempts)
            except Exception as e:  # pragma: no cover - defensive
                print(f"❌ Error saving failed attempts: {e}")

            remaining = LOCKOUT_THRESHOLD - info["count"]
            return (
                jsonify(
                    {
                        "success": False,
                        "message": f"Incorrect password. {remaining} attempts left.",
                    }
                ),
                401,
            )

        # Clear failed attempts on successful login
        if email in failed_attempts:
            failed_attempts.pop(email, None)
            try:
                save_failed_attempts(failed_attempts)
            except Exception as e:  # pragma: no cover - defensive
                print(f"❌ Error saving failed attempts: {e}")

        # Set session
        try:
            session.permanent = bool(remember_me)
            session["user"] = email
            session["remember_me"] = bool(
                remember_me
            )  # Store remember_me flag in session
            session["role"] = user.get("role", "User")
            session["last_active"] = time.time()
            print("✅ Login success, session active")
            return (
                jsonify({"success": True, "message": "Login successful"}),
                200,
            )
        except Exception as e:  # pragma: no cover - defensive
            print(f"❌ Error setting session: {e}")
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "Server error. Please try again later.",
                    }
                ),
                500,
            )

    except Exception as e:  # pragma: no cover - defensive
        print(f"❌ Unexpected error in login_post: {e}")
        import traceback

        traceback.print_exc()
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Server error. Please try again later.",
                }
            ),
            500,
        )


# =========================
# API: LIST + FILTER + PAGINATION (Supports HTML & JSON)
# =========================
@app.route("/api/products", methods=["GET"])
def api_products():
    """
    GET /api/products
    Query Parameters:
        - q: Search query
        - type: Filter by product type
        - category: Filter by category
        - status: Filter by status
        - stock: Filter by stock level (out/low/ok)
        - page: Page number (default: 1)
        - page_size: Items per page (default: 10)
        - format: Force format (json/html)
    
    Returns JSON or HTML based on Accept header
    """
    # BUG_001 / BUG_006: Require login for product list API
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    products = load_products()

    q = (request.args.get("q") or "").strip().lower()
    ptype = (request.args.get("type") or "").strip()
    cat = (request.args.get("category") or "").strip()
    status = (request.args.get("status") or "").strip()
    brand = (request.args.get("brand") or "").strip()
    stock = (request.args.get("stock") or "").strip()

    # BUG_002: Robust validation for pagination parameters
    raw_page = request.args.get("page", "1")
    raw_page_size = request.args.get("page_size", "10")
    try:
        page = int(raw_page)
        page_size = int(raw_page_size)
    except (TypeError, ValueError):
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Invalid pagination parameters. 'page' and 'page_size' must be integers.",
                }
            ),
            400,
        )

    if page <= 0 or page_size <= 0:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Invalid pagination parameters. 'page' and 'page_size' must be greater than 0.",
                }
            ),
            400,
        )

    # Optional upper bound to avoid huge pages
    if page_size > 1000:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Invalid pagination parameters. 'page_size' is too large.",
                }
            ),
            400,
        )

    # filter
    def match(p):
        if q:
            hay = " ".join([
                str(p.get("product_id","")),
                str(p.get("product_name","")),
                str(p.get("type","")),
                str(p.get("category","")),
                str(p.get("status","")),
            ]).lower()

            if q not in hay:
                return False
        if ptype and p.get("type") != ptype:
            return False
        if cat and p.get("category") != cat:
            return False
        if status and p.get("status") != status:
            return False
        if brand and p.get("brand") != brand:
            return False

        # stock buckets
        level = int(p.get("stock_level") or 0)
        if stock == "out" and level != 0:
            return False
        if stock == "low" and not (1 <= level <= 5):
            return False
        if stock == "ok" and level <= 5:
            return False

        return True

    filtered = [p for p in products if match(p)]

    # meta for dropdown - extract unique values from database
    types = sorted({p.get("type","") for p in products if p.get("type")})
    categories = sorted({p.get("category","") for p in products if p.get("category")})
    statuses = sorted({p.get("status","") for p in products if p.get("status")})
    brands = sorted({p.get("brand","") for p in products if p.get("brand")})

    # pagination
    total_items = len(filtered)
    total_pages = max(1, (total_items + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))

    start = (page - 1) * page_size
    end = start + page_size
    items = filtered[start:end]

    response_data = {
        "success": True,
        "data": {
        "items": items,
        "page": page,
        "total_pages": total_pages,
        "total_items": total_items,
            "meta": {
                "types": types,
                "categories": categories,
                "statuses": statuses,
                "brands": brands
            }
        }
    }

    if wants_json():
        return jsonify(response_data), 200
    else:
        # For HTML, redirect to products page or return JSON anyway
        return jsonify(response_data), 200



# =========================
# API: DELETE PRODUCT (Supports HTML & JSON)
# =========================
@app.route("/api/products/<product_id>", methods=["DELETE"])
def api_delete_product(product_id):
    """
    DELETE /api/products/<product_id>
    
    Deletes a product by ID
    """
    # BUG_001 / BUG_006: Require login for delete API
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    products = load_products()
    before = len(products)
    deleted_product = next((p for p in products if str(p.get("product_id")) == str(product_id)), None)
    
    products = [p for p in products if str(p.get("product_id")) != str(product_id)]
    save_products(products)

    if len(products) == before:
        error_response = {
            "success": False,
            "message": "Product not found",
            "error": f"Product with ID '{product_id}' does not exist"
        }
        if wants_json():
            return jsonify(error_response), 404
        else:
            return jsonify(error_response), 404
    
    response_data = {
        "success": True,
        "message": "Product deleted successfully",
        "data": deleted_product
    }
    
    if wants_json():
        return jsonify(response_data), 200
    else:
        return jsonify(response_data), 200
# =========================
# API: IMPORT CSV
# CSV columns: product_id,product_name,type,category,status,stock_level,price
# =========================
@app.route("/api/products/import", methods=["POST"])
def api_import_products():
    # BUG_001 / BUG_006: Require login for import API
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    if "file" not in request.files:
        return jsonify({"message": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith(".csv"):
        return jsonify({"message": "Only CSV supported"}), 400

    products = load_products()
    existing_ids = {p.get("product_id") for p in products}

    decoded = file.stream.read().decode("utf-8", errors="ignore").splitlines()
    reader = csv.DictReader(decoded)

    added = 0
    for row in reader:
        pid = (row.get("product_id") or "").strip()
        if not pid or pid in existing_ids:
            continue

        item = {
            "product_id": pid,
            "product_name": (row.get("product_name") or "").strip(),
            "type": (row.get("type") or "").strip(),
            "category": (row.get("category") or "").strip(),
            "status": (row.get("status") or "Active").strip(),
            "stock_level": int((row.get("stock_level") or 0)),
            "price": float((row.get("price") or 0))
        }
        products.append(item)
        existing_ids.add(pid)
        added += 1

    save_products(products)
    return jsonify({"message": f"Import done ✅ Added {added} products"})


# =========================================
# ✅ SESSION TIMEOUT FUNCTION
# =========================================
def check_session_timeout():
    if "user" not in session:
        return False

    # If "Remember Me" is checked, skip inactivity timeout check
    remember_me = session.get("remember_me", False)
    if remember_me:
        # Still update last_active for tracking, but don't expire session
        session["last_active"] = time.time()
        return True

    # For normal sessions, check inactivity timeout
    last_active = session.get("last_active", 0)
    now = time.time()

    if now - last_active > INACTIVITY_TIMEOUT:
        session.clear()
        return False

    session["last_active"] = now
    return True


# =========================================
# 9. UTILITY — Logout
# =========================================
@app.route("/logout")
def logout():
    session.pop("user", None)
    session.pop("last_active", None)
    session.pop("remember_me", None)
    return redirect(url_for("login", message="logged_out"))


# =========================================
# 9. UTILITY — Global Search
# =========================================
@app.route("/search")
def global_search():
    q = (request.args.get("q") or "").strip().lower()
    results = []

    if not q:
        return jsonify({"results": []})

    users = load_users()
    for u in users:
        if not isinstance(u, dict):
            continue

        name = (u.get("name") or "").strip()
        email = (u.get("email") or "").strip()
        phone = (u.get("phone") or "").strip()

        if (q in name.lower() or q in email.lower() or q in phone):
            results.append({
                "type": "User",
                "label": f"{name} - {email}",
                "url": "/manage-users",
            })

    menu_items = [
        ("Dashboard", "/dashboard"),
        ("Masters", "/manage-users"),
        ("Manage Users", "/manage-users"),
        ("Products", "/products"),
        ("Customer", "/customer"),
        ("Department Role", "/department-role"),
        ("CRM", "/crm"),
        ("Enquiry List", "/crm"),
        ("Quotation Module", "/crm"),
        ("Sales", "/crm"),
        ("Delivery Note Module", "/crm"),
        ("Invoice Module", "/crm"),
        ("Delivery Note Return", "/crm"),
        ("Invoice Return Module", "/crm"),
    ]

    for label, url in menu_items:
        if q in label.lower():
            results.append({
                "type": "Menu",
                "label": label,
                "url": url,
            })

    return jsonify({"results": results})


# =========================================
# 3. MASTERS — Manage Users — API
# =========================================
@app.route("/api/users", methods=["GET"])
def api_get_users():
    """Get all users as JSON - for Postman/API testing"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    users = _db_fetch_users_ordered(include_id=False)

    user_name = "User"
    user_role = "User"

    current_email = (user_email or "").strip().lower()

    for u in users:
        if not isinstance(u, dict):
            continue
        u_email = (u.get("email") or "").strip().lower()
        if u_email == current_email:
            user_name = u.get("name") or "User"
            user_role = (u.get("role") or "User").strip()
            break

    return jsonify({
        "success": True,
        "users": [user_public_dict(u) for u in users if isinstance(u, dict)],
        "total": len(users),
        "current_user": {
            "email": user_email,
            "name": user_name,
            "role": user_role
        }
    }), 200


# 3. MASTERS — Manage Users — API (continued)
@app.route("/api/users/<int:user_index>", methods=["GET"])
def api_get_user(user_index):
    """Get a single user by index as JSON - for Postman/API testing"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    if user_index < 0:
        return jsonify({"success": False, "message": "User index out of range"}), 404

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT name, email, phone, role, first_name, last_name, country_code,
                   contact_number, branch, department, reporting_to, available_branches, employee_id
            FROM users
            ORDER BY id DESC
            OFFSET %s LIMIT 1
            """,
            (user_index,),
        )
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "message": "User index out of range"}), 404
        user = {
            "name": row[0],
            "email": row[1],
            "phone": row[2],
            "role": row[3],
            "first_name": row[4],
            "last_name": row[5],
            "country_code": row[6],
            "contact_number": row[7],
            "branch": row[8],
            "department": row[9],
            "reporting_to": row[10],
            "available_branches": str(row[11]) if row[11] is not None else "",
            "employee_id": row[12],
        }
    finally:
        cur.close()
        conn.close()

    return jsonify({
        "success": True,
        "user": user_public_dict(user),
        "index": user_index
    }), 200


# =========================================
# ✅ API: CREATE USER (POST /api/users) — JSON for Postman
# =========================================
@app.route("/api/users", methods=["POST"])
def api_create_user():
    """Create new user. Requires JSON body. Use ?format=json or Accept: application/json."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    prof = get_current_user_profile() or {}
    user_role = prof.get("role") or "User"

    if normalize_role(user_role) not in ["superadmin", "admin"]:
        return jsonify({"success": False, "message": "Only Super Admin/Admin can create users."}), 403

    data = request.get_json(silent=True)
    if not data:
        return jsonify({"success": False, "message": "JSON body required"}), 400

    users = load_users()

    first_name = (data.get("first_name") or "").strip()
    last_name = (data.get("last_name") or "").strip()
    email = (data.get("email") or "").strip()
    country_code = (data.get("country_code") or "").strip()
    contact_number = (data.get("contact_number") or "").strip()
    branch = (data.get("branch") or "").strip()
    department = (data.get("department") or "").strip()
    role = (data.get("role") or "").strip()
    reporting_to = (data.get("reporting_to") or "").strip()
    available_branches = (data.get("available_branches") or "").strip()
    employee_id = (data.get("employee_id") or "").strip()

    errors = []
    if not first_name or len(first_name) < 3:
        errors.append("First Name required (min 3 chars)")
    if not last_name or len(last_name) < 3:
        errors.append("Last Name required (min 3 chars)")
    if not email:
        errors.append("Email required")
    elif not EMAIL_REGEX.match(email):
        errors.append("Invalid email format")
    if not branch:
        errors.append("Branch required")
    if not department:
        errors.append("Department required")
    if not role:
        errors.append("Role required")
    if not reporting_to:
        errors.append("Reporting To required")
    if not available_branches:
        errors.append("Available Branches required")
    if not employee_id:
        errors.append("Employee ID required")

    for u in users:
        if isinstance(u, dict):
            if (u.get("email") or "").strip().lower() == email.lower():
                errors.append("Email already exists")
                break
            if (u.get("employee_id") or "") == employee_id:
                errors.append("Employee ID already exists")
                break

    if errors:
        return jsonify({"success": False, "message": "; ".join(errors), "errors": errors}), 400

    full_name = (first_name + " " + last_name).strip()
    full_phone = f"{country_code}{contact_number}" if country_code and contact_number else contact_number
    api_password = (data.get("password") or "").strip() or DEFAULT_BRANCH_USER_PASSWORD

    new_user = {
        "name": full_name,
        "phone": full_phone,
        "first_name": first_name,
        "last_name": last_name,
        "email": email,
        "country_code": country_code,
        "contact_number": contact_number,
        "branch": branch,
        "department": department,
        "role": role,
        "reporting_to": reporting_to,
        "available_branches": available_branches,
        "employee_id": employee_id,
        "password": api_password,
    }
    users.append(new_user)
    save_users(users)

    safe_user = {k: v for k, v in new_user.items() if k != "password"}
    return jsonify({
        "success": True,
        "message": "User created successfully",
        "user": safe_user,
    }), 201


# =========================================
# ✅ API: UPDATE USER (PUT /api/users/<index>) — JSON for Postman
# =========================================
@app.route("/api/users/<int:user_index>", methods=["PUT"])
def api_update_user(user_index):
    """Update user by index. Requires JSON body."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    if user_index < 0:
        return jsonify({"success": False, "message": "User index out of range"}), 404

    data = request.get_json(silent=True) or {}
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT role FROM users WHERE LOWER(email)=LOWER(%s) LIMIT 1",
            ((user_email or "").strip(),),
        )
        current = cur.fetchone()
        if not current or normalize_role(current[0]) not in ["superadmin", "admin"]:
            return jsonify({"success": False, "message": "Only Super Admin/Admin can edit users."}), 403

        cur.execute(
            """
            SELECT id, name, email, phone, role, department, branch
            FROM users
            ORDER BY id DESC
            OFFSET %s LIMIT 1
            """,
            (user_index,),
        )
        base = cur.fetchone()
        if not base:
            return jsonify({"success": False, "message": "User index out of range"}), 404

        user_id = base[0]
        new_name = str(data.get("name", base[1] or "")).strip()
        new_email = str(data.get("email", base[2] or "")).strip()
        new_phone = str(data.get("phone", base[3] or "")).strip()
        new_role = str(data.get("role", base[4] or "User")).strip() or "User"
        new_department = str(data.get("department", base[5] or "")).strip()
        new_branch = str(data.get("branch", base[6] or "")).strip()

        cur.execute(
            """
            UPDATE users
            SET name=%s, email=%s, phone=%s, role=%s, department=%s, branch=%s
            WHERE id=%s
            """,
            (new_name, new_email, new_phone, new_role, new_department, new_branch, user_id),
        )
        conn.commit()

        refreshed = {
            "name": new_name,
            "email": new_email,
            "phone": new_phone,
            "role": new_role,
            "department": new_department,
            "branch": new_branch,
        }
        return jsonify({
            "success": True,
            "message": "User updated",
            "user": user_public_dict(refreshed),
        }), 200
    finally:
        cur.close()
        conn.close()


# =========================================
# ✅ API: DELETE USER (DELETE /api/users/<index>) — JSON for Postman
# =========================================
@app.route("/api/users/<int:user_index>", methods=["DELETE"])
def api_delete_user(user_index):
    """Delete user by index."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    if user_index < 0:
        return jsonify({"success": False, "message": "User index out of range"}), 404

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT role FROM users WHERE LOWER(email)=LOWER(%s) LIMIT 1",
            ((user_email or "").strip(),),
        )
        current = cur.fetchone()
        if not current or normalize_role(current[0]) not in ["superadmin", "admin"]:
            return jsonify({"success": False, "message": "Only Super Admin/Admin can delete users."}), 403

        cur.execute(
            "SELECT id, email FROM users ORDER BY id DESC OFFSET %s LIMIT 1",
            (user_index,),
        )
        target = cur.fetchone()
        if not target:
            return jsonify({"success": False, "message": "User index out of range"}), 404

        user_id, deleted_email = target
        cur.execute("DELETE FROM users WHERE id=%s", (user_id,))
        conn.commit()
        return jsonify({
            "success": True,
            "message": "User deleted successfully",
            "deleted_email": deleted_email or "",
        }), 200
    finally:
        cur.close()
        conn.close()


# 3. MASTERS — Manage Users — Delete API
@app.route("/delete-user/<int:user_id>", methods=["DELETE"])
def delete_user(user_id):
    try:
        user_email = session.get("user")
        if not user_email:
            return jsonify({"success": False, "message": "Not logged in"}), 401

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "SELECT role FROM users WHERE LOWER(email)=LOWER(%s) LIMIT 1",
                ((user_email or "").strip(),),
            )
            current = cur.fetchone()
            if not current:
                return jsonify({"success": False, "message": "Current user not found"}), 403
            if normalize_role(current[0]) != "superadmin":
                return jsonify({
                    "success": False,
                    "message": "Only super admins can delete users."
                }), 403

            if user_id < 0:
                return jsonify({"success": False, "message": "Invalid user ID"}), 404

            cur.execute(
                "SELECT id, email FROM users ORDER BY id DESC OFFSET %s LIMIT 1",
                (user_id,),
            )
            target = cur.fetchone()
            if not target:
                return jsonify({"success": False, "message": "Invalid user ID"}), 404
            db_id, deleted_email = target

            cur.execute("DELETE FROM users WHERE id=%s", (db_id,))
            conn.commit()
            return jsonify({
                "success": True,
                "message": "User deleted successfully",
                "deleted_email": deleted_email or ""
            }), 200
        finally:
            cur.close()
            conn.close()

    except Exception as e:
        print("❌ Delete user error:", e)
        return jsonify({
            "success": False,
            "message": "Server error while deleting user"
        }), 500



# =========================================
# 7. CRM — Enquiry List
# =========================================

def _load_enquiry_file():
    """Load new-enquiry.json as a dict keyed by enquiry_id."""
    if not os.path.exists(ENQUIRY_FILE):
        return {}
    try:
        with open(ENQUIRY_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError, TypeError):
        return {}


def _save_enquiry_file(data):
    """Persist enquiry dict to new-enquiry.json."""
    if not isinstance(data, dict):
        data = {}
    with open(ENQUIRY_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)


def generate_enquiry_id():
    """Next ENQ-#### id from existing keys in ENQUIRY_FILE (dict format)."""
    data = _load_enquiry_file()
    if not data:
        return "ENQ-0001"
    max_n = 0
    for key in data.keys():
        if not isinstance(key, str):
            continue
        k = key.strip().upper()
        if not k.startswith("ENQ"):
            continue
        try:
            tail = k.split("-")[-1] if "-" in k else k[3:]
            tail = tail.lstrip("0") or "0"
            n = int(tail)
            max_n = max(max_n, n)
        except ValueError:
            continue
    return f"ENQ-{max_n + 1:04d}"


def _normalize_incoming_enquiry_details(details: dict) -> dict:
    """Accept phone_number from API clients; store as phone."""
    if not isinstance(details, dict):
        return {}
    out = dict(details)
    if out.get("phone_number") and not out.get("phone"):
        out["phone"] = str(out.pop("phone_number", "")).strip()
    elif "phone_number" in out:
        out.pop("phone_number", None)
    return out


def _enquiry_to_api_dict(enquiry_id: str, enq_data: dict, include_items: bool = True) -> dict:
    """Single enquiry payload for JSON APIs."""
    if not isinstance(enq_data, dict):
        enq_data = {}
    details = _normalize_incoming_enquiry_details(enq_data.get("enquiry_details") or {})
    row = {
        "enquiry_id": enquiry_id,
        "enquiry_details": details,
        "first_name": details.get("first_name", ""),
        "last_name": details.get("last_name", ""),
        "email": details.get("email", ""),
        "phone": details.get("phone", ""),
        "phone_number": details.get("phone", ""),
        "status": details.get("status", "New"),
    }
    if include_items:
        row["items"] = enq_data.get("items") if isinstance(enq_data.get("items"), dict) else {}
    return row




def _get_current_user_role():
    """Get current user's role from session / DB profile (not users.json alone)."""
    user_email = session.get("user")
    if not user_email:
        return None
    sr = session.get("role")
    if sr:
        return (str(sr).strip()).replace(" ", "").replace("_", "").lower()
    prof = get_current_user_profile()
    if prof and prof.get("role"):
        return (prof.get("role") or "User").strip().replace(" ", "").replace("_", "").lower()
    users = load_users()
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            role = (u.get("role") or "User").strip()
            return role.replace(" ", "").replace("_", "").lower()
    return "user"


def _get_logged_in_user_name():
    """Helper to fetch the current logged-in user's display name for profile dropdown."""
    user_email = session.get("user")
    if not user_email:
        return "User"

    prof = get_current_user_profile()
    if prof and prof.get("name"):
        return prof.get("name")
    return "User"


@app.route("/enquiry-list")
def enquiry_list():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"

    # Load enquiries from JSON file (new-enquiry.json)
    enquiries = []
    data = _load_enquiry_file()
    if data:
        # Convert dict to list
        for enq_id, enq_data in data.items():
            details = enq_data.get("enquiry_details", {})
            enquiries.append({
                "enquiry_id": enq_id,
                "first_name": details.get("first_name", ""),
                "last_name": details.get("last_name", ""),
                "email": details.get("email", ""),
                "phone_number": details.get("phone", ""),
                "status": details.get("status", "New"),
                "items": enq_data.get("items", {})
            })

    # JSON API variant for Fetch/XHR (same pattern as manage users)
    if wants_json():
        return jsonify(
            {
                "success": True,
                "data": enquiries,
                "total": len(enquiries),
                "current_user": {
                    "email": user_email,
                    "name": user_name,
                    "role": user_role,
                },
                "permissions": get_effective_permissions_for_session(),
            }
        ), 200

    return render_template(
        "enquiry-list.html",
        title="Enquiry List - Stackly",
        page="enquiry_list",
        section="crm",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
        enquiries=enquiries,
    )





@app.route("/api/enquiry/<enquiry_id>")
def get_enquiry(enquiry_id):
    if not os.path.exists(ENQUIRY_FILE):
        return jsonify(success=False, message="Enquiry file missing")

    with open(ENQUIRY_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    enquiry = data.get(enquiry_id)
    if not enquiry:
        return jsonify(success=False, message="Enquiry not found")

    details = enquiry.get("enquiry_details", {})
    return jsonify(success=True, data={
        "enquiry_id": enquiry_id,
        "first_name": details.get("first_name", ""),
        "last_name": details.get("last_name", ""),
        "phone": details.get("phone", ""),
        "email": details.get("email", "")
    })






@app.route("/update-enquiry/<enquiry_id>", methods=["POST"])
def update_enquiry(enquiry_id):
    role = _get_current_user_role()
    if role not in ("admin", "superadmin"):
        return jsonify(success=False, message="Only Admin or Super Admin can edit enquiries."), 403

    if not os.path.exists(ENQUIRY_FILE):
        return jsonify(success=False, message="Enquiry file missing"), 400

    with open(ENQUIRY_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    if enquiry_id not in data:
        return jsonify(success=False, message="Enquiry not found"), 404

    req_data = request.get_json()
    if not req_data:
        return jsonify(success=False, message="Invalid request"), 400

    # Get existing enquiry details
    details = data[enquiry_id].get("enquiry_details", {})

    # Update only the fields from the request
    details["first_name"] = req_data.get("first_name", details.get("first_name", ""))
    details["last_name"] = req_data.get("last_name", details.get("last_name", ""))
    details["phone"] = req_data.get("phone_number", details.get("phone", ""))
    details["email"] = req_data.get("email", details.get("email", ""))

    # Save updated details back
    data[enquiry_id]["enquiry_details"] = details

    with open(ENQUIRY_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

    return jsonify(success=True, message="Enquiry updated successfully")





@app.route("/delete-enquiry/<enquiry_id>", methods=["DELETE"])
def delete_enquiry(enquiry_id):
    role = _get_current_user_role()
    if role != "superadmin":
        return jsonify(success=False, message="Only Super Admin can delete enquiries."), 403

    if not os.path.exists(ENQUIRY_FILE):
        return jsonify(success=False, message="Enquiry file missing")

    with open(ENQUIRY_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    if enquiry_id not in data:
        return jsonify(success=False, message="Enquiry not found")

    # Remove enquiry
    del data[enquiry_id]

    with open(ENQUIRY_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)

    return jsonify(success=True, message="Enquiry deleted successfully")


# =========================================
# 7b. CRM — Enquiries REST API (Postman / JSON clients)
#   GET    /api/enquiries              — list (optional ?search=&status=)
#   GET    /api/enquiries/<id>         — one enquiry (details + items)
#   POST   /api/enquiries              — create (optional enquiry_id; auto if omitted)
#   PUT    /api/enquiries/<id>         — update details (merge) + optional items (merge)
#   DELETE /api/enquiries/<id>         — delete (Super Admin only)
# =========================================
@app.route("/api/enquiries", methods=["GET"])
def api_enquiries_list():
    """List all enquiries. Requires login. Use Accept: application/json from Postman."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    data = _load_enquiry_file()
    q = (request.args.get("search") or "").strip().lower()
    status_filter = (request.args.get("status") or "").strip()

    users = load_users()
    user_name = "User"
    user_role = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            user_role = (u.get("role") or "User").strip()
            break

    enquiries = []
    for enq_id, enq_data in data.items():
        row = _enquiry_to_api_dict(enq_id, enq_data, include_items=True)
        st = row.get("status") or "New"
        if status_filter and st != status_filter:
            continue
        if q:
            blob = " ".join(
                [
                    str(enq_id),
                    row.get("first_name") or "",
                    row.get("last_name") or "",
                    row.get("email") or "",
                    row.get("phone") or "",
                    st,
                ]
            ).lower()
            if q not in blob:
                continue
        enquiries.append(row)

    return jsonify(
        {
            "success": True,
            "enquiries": enquiries,
            "total": len(enquiries),
            "current_user": {"email": user_email, "name": user_name, "role": user_role},
        }
    ), 200


@app.route("/api/enquiries/<enquiry_id>", methods=["GET"])
def api_enquiries_get_one(enquiry_id):
    """Get one enquiry by id (full details + items)."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    data = _load_enquiry_file()
    enq_data = data.get(enquiry_id)
    if not enq_data:
        return jsonify({"success": False, "message": "Enquiry not found"}), 404

    return jsonify({"success": True, "enquiry": _enquiry_to_api_dict(enquiry_id, enq_data, include_items=True)}), 200


@app.route("/api/enquiries", methods=["POST"])
def api_enquiries_create():
    """Create enquiry. Admin / Super Admin only. Body: { enquiry_details, items?, enquiry_id? }."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    role = _get_current_user_role()
    if role not in ("admin", "superadmin"):
        return jsonify({"success": False, "message": "Only Admin or Super Admin can create enquiries."}), 403

    body = request.get_json(silent=True)
    if not body:
        return jsonify({"success": False, "message": "JSON body required"}), 400

    enquiry_id = (body.get("enquiry_id") or "").strip() or generate_enquiry_id()
    details = _normalize_incoming_enquiry_details(body.get("enquiry_details") or {})
    items = body.get("items") if isinstance(body.get("items"), dict) else {}

    data = _load_enquiry_file()
    if enquiry_id in data:
        return jsonify({"success": False, "message": "Enquiry ID already exists. Omit enquiry_id to auto-generate."}), 409

    if not details.get("status"):
        details["status"] = "New"

    data[enquiry_id] = {"enquiry_details": details, "items": items}
    _save_enquiry_file(data)

    return jsonify(
        {
            "success": True,
            "message": "Enquiry created",
            "enquiry": _enquiry_to_api_dict(enquiry_id, data[enquiry_id], include_items=True),
        }
    ), 201


@app.route("/api/enquiries/<enquiry_id>", methods=["PUT"])
def api_enquiries_update(enquiry_id):
    """Update enquiry: merge enquiry_details; merge items if provided."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    role = _get_current_user_role()
    if role not in ("admin", "superadmin"):
        return jsonify({"success": False, "message": "Only Admin or Super Admin can update enquiries."}), 403

    body = request.get_json(silent=True)
    if not body:
        return jsonify({"success": False, "message": "JSON body required"}), 400

    data = _load_enquiry_file()
    if enquiry_id not in data:
        return jsonify({"success": False, "message": "Enquiry not found"}), 404

    if "enquiry_details" in body and isinstance(body["enquiry_details"], dict):
        existing = data[enquiry_id].setdefault("enquiry_details", {})
        merged = _normalize_incoming_enquiry_details(body["enquiry_details"])
        for k, v in merged.items():
            existing[k] = v
        data[enquiry_id]["enquiry_details"] = existing

    if "items" in body:
        if not isinstance(body["items"], dict):
            return jsonify({"success": False, "message": "items must be an object"}), 400
        old_items = data[enquiry_id].setdefault("items", {})
        old_items.update(body["items"])
        data[enquiry_id]["items"] = old_items

    _save_enquiry_file(data)

    return jsonify(
        {
            "success": True,
            "message": "Enquiry updated",
            "enquiry": _enquiry_to_api_dict(enquiry_id, data[enquiry_id], include_items=True),
        }
    ), 200


@app.route("/api/enquiries/<enquiry_id>", methods=["DELETE"])
def api_enquiries_delete(enquiry_id):
    """Delete enquiry. Super Admin only (same as /delete-enquiry)."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    role = _get_current_user_role()
    if role != "superadmin":
        return jsonify({"success": False, "message": "Only Super Admin can delete enquiries."}), 403

    data = _load_enquiry_file()
    if enquiry_id not in data:
        return jsonify({"success": False, "message": "Enquiry not found"}), 404

    del data[enquiry_id]
    _save_enquiry_file(data)

    return jsonify({"success": True, "message": "Enquiry deleted", "deleted_id": enquiry_id}), 200



@app.route("/api/enquiry-items/<enquiry_id>")
def get_enquiry_items(enquiry_id):
    if not os.path.exists(ENQUIRY_FILE):
        return jsonify(success=False, message="Enquiry file missing")

    with open(ENQUIRY_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    enquiry = data.get(enquiry_id)
    if not enquiry:
        return jsonify(success=False, message="Enquiry not found")

    items = enquiry.get("items", {})  # get items dict

    return jsonify(success=True, data=items)




def read_products():
    """
    Read enquiry-related products from ENQUIRY_PRODUCT_FILE (D:\\POS_Project_Latest\\Pos project\\enquiry_product.json).
    """
    if not os.path.exists(ENQUIRY_PRODUCT_FILE):
        return []
    with open(ENQUIRY_PRODUCT_FILE, "r") as f:
        return json.load(f)


def write_products(data):
    with open(ENQUIRY_PRODUCT_FILE, "w") as f:
        json.dump(data, f, indent=4)                                                                             
@app.route("/update-enquiry-items/<enquiry_id>", methods=["POST"])
def update_enquiry_items(enquiry_id):
    role = _get_current_user_role()
    if role not in ("admin", "superadmin"):
        return jsonify(success=False, message="Only Admin or Super Admin can edit enquiry items."), 403
    if not os.path.exists(ENQUIRY_FILE):
        return jsonify(success=False, message="Enquiry file missing"), 400

    with open(ENQUIRY_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    if enquiry_id not in data:
        return jsonify(success=False, message="Enquiry not found"), 404

    req_data = request.get_json()
    if not req_data or "items" not in req_data:
        return jsonify(success=False, message="Invalid request"), 400

    # Update the items for this enquiry
    for item_code, item_details in req_data["items"].items():
        if "items" not in data[enquiry_id]:
            data[enquiry_id]["items"] = {}
        data[enquiry_id]["items"][item_code] = item_details

    # Save back to JSON file
    with open(ENQUIRY_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

    return jsonify(success=True, message="Enquiry items updated successfully")


    


@app.route("/new-enquiry")
def new_enquiry():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    role = _get_current_user_role()
    if role not in ("admin", "superadmin"):
        return redirect(url_for("enquiry_list") + "?message=create_denied")

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "new-enquiry.html",
        title="New-Enquiry - Stackly",
        page="new_enquiry",
        section="masters",
        user_email=user_email,
        user_name=user_name,
    )



def load_data():
    """Backward-compatible alias for enquiry JSON dict (used by /add-item, /save-enquiry)."""
    return _load_enquiry_file()


def save_data(data):
    """Backward-compatible save for enquiry JSON dict."""
    _save_enquiry_file(data)


@app.route("/generate-enquiry-id")
def generate_id():
    return jsonify(enquiry_id=generate_enquiry_id())


@app.route("/save-enquiry", methods=["POST"])
def save_enquiry():
    role = _get_current_user_role()
    if role not in ("admin", "superadmin"):
        return jsonify(success=False, message="Only Admin or Super Admin can create or update enquiries."), 403

    payload = request.get_json(silent=True)
    if not payload:
        payload = dict(request.form) if request.form else {}
    if not payload:
        return jsonify(success=False, message="Invalid request"), 400

    enquiry_id = payload.get("enquiry_id")
    if not enquiry_id:
        return jsonify(success=False, message="Enquiry ID required"), 400
    enquiry_details = payload.get("enquiry_details") or {}
    new_items = payload.get("items") or {}

    data = _load_enquiry_file()

    if enquiry_id in data:
        # update details
        # data[enquiry_id]["enquiry_details"] = enquiry_details

        # merge items
        old_items = data[enquiry_id].get("items", {})
        old_items.update(new_items)
        data[enquiry_id]["items"] = old_items

    else:
        data[enquiry_id] = {
            "enquiry_details": enquiry_details,
            "items": new_items
        }

    _save_enquiry_file(data)

    return jsonify({"success": True})





@app.route("/get-enquiry-add-items/<enquiry_id>")
def get_enquiry_add_items_(enquiry_id):
    try:
        if not os.path.exists(ENQUIRY_FILE):
            return jsonify({"success": False, "items": {}})

        with open(ENQUIRY_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)

        enquiry = data.get(enquiry_id)
        if not enquiry:
            return jsonify({"success": False, "items": {}})

        # ✅ Return ALL items - NO FILTERING NEEDED since items are physically removed
        all_items = enquiry.get("items", {})
        
        print(f"Enquiry {enquiry_id} has {len(all_items)} items: {list(all_items.keys())}")
        
        return jsonify({
            "success": True,
            "items": all_items  # Return everything - deleted items are GONE from file
        })
        
    except Exception as e:
        print(f"Error in get_enquiry_add_items_: {e}")
        return jsonify({"success": False, "items": {}})

@app.route("/add-item", methods=["POST"])
def add_item():
    payload = request.get_json(force=True)

    enquiry_id = payload["enquiry_id"]
    item = payload["item"]

    data = load_data()

    if enquiry_id not in data:
        return jsonify(error="Invalid enquiry id"), 400

    item_code = item["item_code"]

    data[enquiry_id]["items"][item_code] = item

    save_data(data)

    return jsonify(status="item added")

@app.route("/check-email-enquiry")
def check_email_enquiry():
    try:
        email = request.args.get("email", "").lower()

        if not os.path.exists(ENQUIRY_FILE):
            return jsonify({"exists": False})

        with open(ENQUIRY_FILE, "r") as f:
            enquiries = json.load(f)   # THIS IS A DICT

        # Loop through dict values
        for enquiry_id, enquiry in enquiries.items():
            details = enquiry.get("enquiry_details", {})
            if details.get("email", "").lower() == email:
                return jsonify({
                    "exists": True,
                    "enquiry_id": enquiry_id,
                    "customer": details
                })

        return jsonify({"exists": False})

    except Exception as e:
        print("❌ CHECK EMAIL ERROR:", e)
        return jsonify({"exists": False, "error": str(e)}), 500


def load_products():
    if not os.path.exists(PRODUCT_FILE):
        return []
    with open(PRODUCT_FILE, "r") as f:
        return json.load(f)

@app.route("/get-product/<product_id>")
def get_product(product_id):
    products = load_products()
    for p in products:
        if p["product_id"] == product_id:
            return jsonify({"success": True, "product": p})
    return jsonify({"success": False, "message": "Product not found"}), 404


@app.route("/delete-enquiry-item/<enquiry_id>/<item_code>", methods=["DELETE"])
def delete_enquiry_item(enquiry_id, item_code):
    role = _get_current_user_role()
    if role != "superadmin":
        return jsonify(success=False, message="Only Super Admin can delete enquiry items."), 403
    try:
        print(f"\n=== HARD DELETE REQUEST ===")
        print(f"Enquiry ID: {enquiry_id}")
        print(f"Item Code: {item_code}")
        
        if not os.path.exists(ENQUIRY_FILE):
            return jsonify(success=False, message="Enquiry file missing"), 404

        # Read the file
        with open(ENQUIRY_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Check if enquiry exists
        if enquiry_id not in data:
            return jsonify(success=False, message="Enquiry not found"), 404

        # Check if items exist
        if "items" not in data[enquiry_id]:
            return jsonify(success=False, message="No items found"), 404

        items = data[enquiry_id]["items"]

        # Check if item exists
        if item_code not in items:
            return jsonify(success=False, message="Item not found"), 404

        # 🔴 HARD DELETE - Completely remove the item from dictionary
        del items[item_code]
        print(f"✓ Item {item_code} completely removed from JSON")
        print(f"Remaining items: {list(items.keys())}")

        # Save back to file
        with open(ENQUIRY_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        
        print(f"✓ File saved successfully")
        print(f"=== DELETE COMPLETED ===\n")
        
        return jsonify(success=True, message="Item permanently deleted")

    except Exception as e:
        print(f"Error deleting item: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, message=f"Error: {str(e)}"), 500                                                                                                                                                                                                                    
@app.route("/get-product-config")
def get_product_config():
    if not os.path.exists(PRODUCT_FILE):
        return jsonify(success=False, message="Product file missing")

    with open(PRODUCT_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        return jsonify(success=False, message="Invalid product data")

    # Get all product IDs
    product_ids = [p.get("product_id", "") for p in data if "product_id" in p]

    # Calculate max length
    max_id_length = max(len(pid) for pid in product_ids) if product_ids else 4

    return jsonify(
        success=True,
        max_id_length=max_id_length,
        product_ids=product_ids
    )











# =========================================
# ✅ Helper function for quotation (QUOTATION)
# =========================================
# In-memory cache to avoid repeated file reads (speeds up list + get-quotation + filters)
_quotation_cache = None
_quotation_cache_time = 0.0
QUOTATION_CACHE_TTL = 3  # seconds


def _invalidate_quotation_cache():
    global _quotation_cache, _quotation_cache_time
    _quotation_cache = None
    _quotation_cache_time = 0.0


def load_quotations():
    global _quotation_cache, _quotation_cache_time
    now = time.time()
    if _quotation_cache is not None and (now - _quotation_cache_time) < QUOTATION_CACHE_TTL:
        return _quotation_cache

    base = os.path.dirname(QUOTATION_FILE)
    if base:
        os.makedirs(base, exist_ok=True)
    if not os.path.exists(QUOTATION_FILE):
        with open(QUOTATION_FILE, "w", encoding="utf-8") as f:
            json.dump([], f)
        _quotation_cache = []
        _quotation_cache_time = time.time()
        return []

    with open(QUOTATION_FILE, "r", encoding="utf-8") as f:
        try:
            data = json.load(f)
            out = data if isinstance(data, list) else []
        except Exception:
            out = []
    _quotation_cache = out
    _quotation_cache_time = time.time()
    return out


def save_quotations(items):
    base = os.path.dirname(QUOTATION_FILE)
    if base:
        os.makedirs(base, exist_ok=True)
    with open(QUOTATION_FILE, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2)
    _invalidate_quotation_cache()

def generate_quotation_id(items):
    max_no = 0
    for q in items:
        qid = str(q.get("quotation_id", ""))
        if qid.startswith("Q") and qid[1:].isdigit():
            max_no = max(max_no, int(qid[1:]))
    return f"Q{max_no + 1}"




# ================================
# QUOTATION  PAGE ROUTE
# ================================


@app.route("/quotation")
def quotation():
    user_email = session.get("user")
    if not user_email:
       return redirect(url_for("login", message="session_expired"))

    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"

    return render_template(
        "quotation.html",
        title="Quotation - Stackly",
        page="quotation",
        section="crm",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
    )

# ============ API LIST ============
@app.route("/api/quotations", methods=["GET"])
def api_quotations():
    if "user" not in session:
        return jsonify(success=False, message="Session expired"), 401

    items = load_quotations()

    q = (request.args.get("q") or "").strip().lower()
    status = (request.args.get("status") or "").strip().lower()
    qtype = (request.args.get("type") or "").strip().lower()
    sales_rep = (request.args.get("sales_rep") or "").strip().lower()

    page = int(request.args.get("page") or 1)
    per_page = int(request.args.get("per_page") or 7)

    # ---- Filter ----
    filtered = []
    for it in items:
        if q:
            hay = f"{it.get('quotation_id','')} {it.get('customer_name','')}".lower()
            if q not in hay:
                continue
        if status and (it.get("status","").lower() != status):
            continue
        if qtype and (it.get("quotation_type","").lower() != qtype):
            continue
        if sales_rep and (it.get("sales_rep","").lower() != sales_rep):
            continue
        filtered.append(it)

    # ---- Collect sales reps for dropdown ----
    reps = sorted({ (x.get("sales_rep") or "").strip() for x in items if (x.get("sales_rep") or "").strip() })

    total = len(filtered)
    total_pages = max(1, math.ceil(total / per_page))
    page = max(1, min(page, total_pages))

    start = (page - 1) * per_page
    end = start + per_page
    page_items = filtered[start:end]

    return jsonify(
        success=True,
        total=total,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        items=page_items,
        sales_reps=reps
    )

# ============ API CREATE (optional starter) ============
@app.route("/api/quotations", methods=["POST"])
def api_create_quotation():
    if "user" not in session:
        return jsonify(success=False, message="Session expired"), 401

    data = request.get_json(force=True) or {}
    items = load_quotations()
    new_id = generate_quotation_id(items)

    new_item = {
        "quotation_id": new_id,
        "quotation_type": (data.get("quotation_type") or "service").lower(),
        "customer_name": data.get("customer_name") or "",
        "sales_rep": data.get("sales_rep") or "",
        "quotation_date": data.get("quotation_date") or datetime.now().strftime("%Y-%m-%d"),
        "status": "draft",
        "grand_total": float(data.get("grand_total") or 0),
    }

    items.insert(0, new_item)
    save_quotations(items)
    return jsonify(success=True, item=new_item)



# ================================
# ADD NEW QUOTATION PAGE ROUTE
# ================================

@app.route("/add-new-quotation")
def add_new_quotation():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"

    # RBAC guard: users without quotation create/edit access cannot open this page.
    role_norm = normalize_role(user_role)
    can_by_role = role_norm in ["superadmin", "admin"]
    q_perm = (get_effective_permissions_for_session() or {}).get("quotation", {})
    can_by_matrix = bool(q_perm.get("full_access") or q_perm.get("create") or q_perm.get("edit"))
    if not (can_by_role or can_by_matrix):
        return redirect(url_for("quotation"))

    return render_template(
        "add-new-quotation.html", 
        title="Add-New-Quotation - Stackly",
        page="quotation",
        section="crm",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
    )

# automatically fill dropdown customer type,sales rep,payment term
@app.route("/get-customers-quotation")
def get_customers_quotation():
    try:
        # with open(CUSTOMER_FILE, "r") as file:
        #     customers = json.load(file)
        customers = get_customers_from_db()
        return jsonify(customers)
    except Exception as e:
        return jsonify({"error": str(e)}), 500




# ===================================================
# PRODUCT ENDPOINTS
# ===================================================

@app.route('/get-products')
def get_products():
    try:
        with open(PRODUCT_FILE, 'r') as f:
            products = json.load(f)
        return jsonify(products)
    except FileNotFoundError:
        return jsonify([])



# ===================================================
# GENERATE QUOTATION ID - FIXED VERSION
# ===================================================

def generate_quotation_id():
    """Generate quotation ID in format QA-0001, QA-0002, etc."""
    try:
        # Check if file exists
        if not os.path.exists(QUOTATION_FILE):
            print("📄 Quotation file not found, starting with QA-0001")
            return "QA-0001"
        
        with open(QUOTATION_FILE, "r") as file:
            quotations = json.load(file)

        if not quotations:
            print("📄 No quotations found, starting with QA-0001")
            return "QA-0001"

        # Find the highest QA-XXXX number
        max_number = 0
        
        for q in quotations:
            q_id = q.get("quotation_id", "")
            
            # Only look for IDs in format QA-XXXX
            if q_id and q_id.startswith("QA-"):
                try:
                    # Extract the number part (after "QA-")
                    number_part = q_id.split("-")[1]
                    # Convert to integer
                    num = int(number_part)
                    if num > max_number:
                        max_number = num
                        print(f"  Found QA-{num:04d}")
                except (ValueError, IndexError):
                    # Skip if format is wrong
                    continue
        
        # Increment by 1
        new_number = max_number + 1
        
        # Format with leading zeros (0001, 0002, etc.)
        new_id = f"QA-{new_number:04d}"
        print(f"✅ Generated new quotation ID: {new_id}")
        
        return new_id

    except FileNotFoundError:
        print("📄 Quotation file not found, starting with QA-0001")
        return "QA-0001"
    except Exception as e:
        print(f"❌ Error generating quotation ID: {e}")
        return "QA-0001"

# ===================================================
# API ROUTE FOR QUOTATION ID
# ===================================================

@app.route('/generate-quotation-id')
def generate_quotation_id_route():
    """API endpoint to generate quotation ID"""
    try:
        quotation_id = generate_quotation_id()
        print(f"🚀 Returning ID: {quotation_id}")
        return jsonify({
            'success': True,
            'quotation_id': quotation_id
        })
    except Exception as e:
        print(f"❌ Route error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ===================================================
# DEBUG ROUTE TO CHECK QUOTATION IDs
# ===================================================

@app.route('/debug-quotation-ids')
def debug_quotation_ids():
    """Debug endpoint to see what IDs are in the file"""
    try:
        if not os.path.exists(QUOTATION_FILE):
            return jsonify({
                'file_exists': False,
                'message': 'Quotation file not found'
            })
        
        with open(QUOTATION_FILE, 'r') as f:
            quotations = json.load(f)
        
        # Extract all QA-XXXX IDs
        qa_ids = []
        other_ids = []
        
        for q in quotations:
            q_id = q.get('quotation_id', 'NO ID')
            if q_id and q_id.startswith('QA-'):
                qa_ids.append(q_id)
            else:
                other_ids.append(q_id)
        
        # Find the highest QA number
        max_qa_number = 0
        for q_id in qa_ids:
            try:
                num = int(q_id.split('-')[1])
                if num > max_qa_number:
                    max_qa_number = num
            except:
                pass
        
        return jsonify({
            'total_quotations': len(quotations),
            'qa_format_ids': qa_ids,
            'other_format_ids': other_ids,
            'highest_qa_number': max_qa_number,
            'next_qa_id': f"QA-{max_qa_number + 1:04d}" if max_qa_number > 0 else "QA-0001",
            'file_path': QUOTATION_FILE,
            'file_exists': os.path.exists(QUOTATION_FILE)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
# ===================================================
# SAVE QUOTATION
# ===================================================

@app.route('/save-quotation', methods=['POST'])
def save_quotation():
    try:
        data = request.json
        quotation_id = data.get('quotation_id')
        status = data.get('status', 'draft')
        
        # Load existing quotations
        try:
            with open(QUOTATION_FILE, 'r') as f:
                quotations = json.load(f)
        except FileNotFoundError:
            quotations = []
        
        # Add timestamps
        data['last_updated'] = datetime.now().isoformat()
        data['created_at'] = data.get('created_at', datetime.now().isoformat())
        
        # Initialize status history if not exists
        if 'status_history' not in data:
            data['status_history'] = []
        
        # Add status change to history
        status_entry = {
            'status': status,
            'date': data.get('status_date', datetime.now().isoformat()),
            'user': data.get('submitted_by', 'System'),
            'notes': f'Quotation {status}'
        }
        
        if status == 'rejected' and data.get('rejection_reason'):
            status_entry['notes'] = f'Quotation rejected: {data["rejection_reason"]}'
        
        data['status_history'].append(status_entry)
        
        # Check if quotation already exists
        existing_index = None
        for i, q in enumerate(quotations):
            if q.get('quotation_id') == quotation_id:
                existing_index = i
                break
        
        # Prevent duplicate Customer PO Reference (case-insensitive)
        customer_po = (data.get('customer_po') or '').strip()
        if customer_po:
            customer_po_lower = customer_po.lower()
            for i, q in enumerate(quotations):
                if existing_index is not None and i == existing_index:
                    continue
                existing_po = (q.get('customer_po') or '').strip()
                if existing_po and existing_po.lower() == customer_po_lower:
                    return jsonify({
                        'success': False,
                        'error': 'Customer PO Reference already exists. Please use a unique value.',
                        'duplicate_field': 'customer_po'
                    }), 400
        
        if existing_index is not None:
            quotations[existing_index] = data
            message = f'Quotation {quotation_id} updated with status: {status}'
        else:
            quotations.append(data)
            message = f'New quotation {quotation_id} created with status: {status}'
        
        # Save back to file
        with open(QUOTATION_FILE, 'w', encoding='utf-8') as f:
            json.dump(quotations, f, indent=2)
        _invalidate_quotation_cache()
        return jsonify({
            'success': True,
            'message': message,
            'quotation_id': quotation_id,
            'status': status
        })
        
    except Exception as e:
        print(f"Error saving quotation: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ===================================================
# CHECK CUSTOMER PO REFERENCE (LIVE DUPLICATE CHECK)
# ===================================================

@app.route('/check-customer-po', methods=['GET'])
def check_customer_po():
    """
    GET /check-customer-po?value=CP-0001&exclude_quotation_id=QA-0012
    Returns { "duplicate": true/false } (case-insensitive).
    exclude_quotation_id: when editing, exclude this quotation from the check.
    """
    try:
        value = (request.args.get('value') or '').strip()
        exclude_quotation_id = (request.args.get('exclude_quotation_id') or '').strip()

        if not value:
            return jsonify({'success': True, 'duplicate': False}), 200

        quotations = load_quotations()
        value_lower = value.lower()

        for q in quotations:
            if exclude_quotation_id and q.get('quotation_id') == exclude_quotation_id:
                continue
            existing_po = (q.get('customer_po') or '').strip()
            if existing_po and existing_po.lower() == value_lower:
                return jsonify({
                    'success': True,
                    'duplicate': True,
                    'message': 'Customer PO Reference already exists. Please use a unique value.'
                }), 200

        return jsonify({'success': True, 'duplicate': False}), 200
    except Exception as e:
        print(f"Error in check_customer_po: {e}")
        return jsonify({'success': False, 'duplicate': False, 'error': str(e)}), 500


# ===================================================
# GET SINGLE QUOTATION
# ===================================================

@app.route('/get-quotation/<quotation_id>')
def get_quotation(quotation_id):
    try:
        quotations = load_quotations()
        for quotation in quotations:
            if quotation.get('quotation_id') == quotation_id:
                return jsonify({
                    'success': True,
                    'quotation': quotation
                })
        return jsonify({
            'success': False,
            'error': 'Quotation not found'
        }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ===================================================
# GET QUOTATIONS BY STATUS
# ===================================================

@app.route('/get-quotations/<status>')
def get_quotations_by_status(status):
    try:
        quotations = load_quotations()
        if status and status != 'all':
            filtered = [q for q in quotations if q.get('status') == status]
        else:
            filtered = quotations
        return jsonify({
            'success': True,
            'quotations': filtered,
            'count': len(filtered)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ===================================================
# COMMENTS ENDPOINTS
# ===================================================

@app.route('/add-comment', methods=['POST'])
def add_comment():
    try:
        data = request.json
        quotation_id = data.get('quotation_id')
        comment = data.get('comment')
        user = data.get('user', 'Admin')
        
        # Load existing comments
        try:
            with open(COMMENTS_FILE, 'r') as f:
                all_comments = json.load(f)
        except FileNotFoundError:
            all_comments = {}
        
        # Initialize comments for this quotation if not exists
        if quotation_id not in all_comments:
            all_comments[quotation_id] = []
        
        # Add new comment
        all_comments[quotation_id].append({
            'id': str(uuid.uuid4()),
            'user': user,
            'comment': comment,
            'time': datetime.now().isoformat()
        })
        
        # Save back to file
        with open(COMMENTS_FILE, 'w') as f:
            json.dump(all_comments, f, indent=2)
        
        return jsonify({
            'success': True,
            'message': 'Comment added successfully'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
@app.route('/get-comments/<quotation_id>')
def get_comments(quotation_id):
    try:
        with open(COMMENTS_FILE, 'r') as f:
            all_comments = json.load(f)
        
        comments = all_comments.get(quotation_id, [])
        
        # Sort by time descending (newest first)
        comments.sort(key=lambda x: x['time'], reverse=True)
        
        # Pagination parameters
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 5))
        
        total = len(comments)
        start = (page - 1) * per_page
        end = start + per_page
        paginated = comments[start:end]
        
        # Format for display
        formatted = []
        for comment in paginated:
            formatted.append({
                'user': comment['user'],
                'comment': comment['comment'],
                'time': datetime.fromisoformat(comment['time']).strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return jsonify({
            'comments': formatted,
            'total': total,
            'page': page,
            'per_page': per_page,
            'has_more': end < total
        })
        
    except FileNotFoundError:
        return jsonify({'comments': [], 'total': 0, 'has_more': False})
    except Exception as e:
        return jsonify({'comments': [], 'total': 0, 'has_more': False})





def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
@app.route('/upload-attachment', methods=['POST'])
def upload_attachment():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        quotation_id = request.form.get('quotation_id')
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        # Check file size
        file.seek(0, os.SEEK_END)
        file_length = file.tell()
        file.seek(0)
        if file_length > MAX_FILE_SIZE_BYTES:
            return jsonify({
                'success': False,
                'error': f'File size exceeds {MAX_FILE_SIZE_MB} MB'
            }), 400
        
        # Check file extension
        if not allowed_file(file.filename):
            return jsonify({
                'success': False,
                'error': f'File type not allowed. Allowed: {", ".join(ALLOWED_EXTENSIONS)}'
            }), 400
        
        # Load existing attachments for this quotation
        metadata_file = os.path.join(ATTACHMENTS_FOLDER, 'metadata.json')
        try:
            with open(metadata_file, 'r') as f:
                attachments = json.load(f)
        except FileNotFoundError:
            attachments = []
        
        # Count current attachments for this quotation
        current_count = sum(1 for a in attachments if a['quotation_id'] == quotation_id)
        if current_count >= MAX_ATTACHMENTS_PER_QUOTATION:
            return jsonify({
                'success': False,
                'error': f'Maximum {MAX_ATTACHMENTS_PER_QUOTATION} files allowed per quotation'
            }), 400
        
        # Generate unique filename
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        unique_filename = f"{quotation_id}_{uuid.uuid4().hex}.{file_ext}"
        file_path = os.path.join(ATTACHMENTS_FOLDER, unique_filename)
        
        # Save file
        file.save(file_path)
        
        # Create attachment record
        attachment = {
            'id': str(uuid.uuid4()),
            'quotation_id': quotation_id,
            'original_filename': file.filename,
            'stored_filename': unique_filename,
            'size': file_length,
            'upload_date': datetime.now().isoformat()
        }
        
        attachments.append(attachment)
        
        with open(metadata_file, 'w') as f:
            json.dump(attachments, f, indent=2)
        
        return jsonify({
            'success': True,
            'attachment': attachment
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
@app.route('/get-attachments/<quotation_id>')
def get_attachments(quotation_id):
    try:
        metadata_file = os.path.join(ATTACHMENTS_FOLDER, 'metadata.json')
        
        try:
            with open(metadata_file, 'r') as f:
                all_attachments = json.load(f)
        except FileNotFoundError:
            all_attachments = []
        
        # Filter attachments for this quotation
        attachments = [
            {
                'id': a['id'],
                'original_filename': a['original_filename'],
                'size': a['size'],
                'upload_date': datetime.fromisoformat(a['upload_date']).strftime('%Y-%m-%d %H:%M:%S')
            }
            for a in all_attachments if a['quotation_id'] == quotation_id
        ]
        
        return jsonify({
            'success': True,
            'attachments': attachments
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/view-attachment/<attachment_id>')
def view_attachment(attachment_id):
    try:
        metadata_file = os.path.join(ATTACHMENTS_FOLDER, 'metadata.json')
        
        with open(metadata_file, 'r') as f:
            attachments = json.load(f)
        
        attachment = next((a for a in attachments if a['id'] == attachment_id), None)
        
        if not attachment:
            return jsonify({'success': False, 'error': 'Attachment not found'}), 404
        
        file_path = os.path.join(ATTACHMENTS_FOLDER, attachment['stored_filename'])
        
        return send_file(file_path, download_name=attachment['original_filename'], as_attachment=False)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/download-attachment/<attachment_id>')
def download_attachment(attachment_id):
    try:
        metadata_file = os.path.join(ATTACHMENTS_FOLDER, 'metadata.json')
        
        with open(metadata_file, 'r') as f:
            attachments = json.load(f)
        
        attachment = next((a for a in attachments if a['id'] == attachment_id), None)
        
        if not attachment:
            return jsonify({'success': False, 'error': 'Attachment not found'}), 404
        
        file_path = os.path.join(ATTACHMENTS_FOLDER, attachment['stored_filename'])
        
        return send_file(file_path, download_name=attachment['original_filename'], as_attachment=True)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/delete-attachment/<attachment_id>', methods=['DELETE'])
def delete_attachment(attachment_id):
    try:
        metadata_file = os.path.join(ATTACHMENTS_FOLDER, 'metadata.json')
        
        with open(metadata_file, 'r') as f:
            attachments = json.load(f)
        
        attachment = next((a for a in attachments if a['id'] == attachment_id), None)
        
        if not attachment:
            return jsonify({'success': False, 'error': 'Attachment not found'}), 404
        
        # Delete physical file
        file_path = os.path.join(ATTACHMENTS_FOLDER, attachment['stored_filename'])
        if os.path.exists(file_path):
            os.remove(file_path)
        
        # Remove from metadata
        attachments = [a for a in attachments if a['id'] != attachment_id]
        
        with open(metadata_file, 'w') as f:
            json.dump(attachments, f, indent=2)
        
        return jsonify({'success': True, 'message': 'Attachment deleted'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    #=============================================
    # PDF GENERATION TABLAB
    # =========================================
    

@app.route('/check-quotation/<quotation_id>')
def check_quotation(quotation_id):
    try:
        with open(QUOTATION_FILE, 'r') as f:
            quotations = json.load(f)
        
        exists = any(q.get('quotation_id') == quotation_id for q in quotations)
        
        return jsonify({
            'success': True,
            'exists': exists,
            'quotation_id': quotation_id
        })
    except FileNotFoundError:
        return jsonify({
            'success': True,
            'exists': False,
            'quotation_id': quotation_id
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500       



@app.route('/debug-quotations')
def debug_quotations():
    try:
        with open(QUOTATION_FILE, 'r') as f:
            quotations = json.load(f)
        
        return jsonify({
            'success': True,
            'count': len(quotations),
            'quotation_ids': [q.get('quotation_id') for q in quotations],
            'file_path': QUOTATION_FILE,
            'file_exists': os.path.exists(QUOTATION_FILE)
        })
    except FileNotFoundError:
        return jsonify({
            'success': False,
            'error': 'Quotation file not found',
            'file_path': QUOTATION_FILE
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

#=============================================
# PDF GENERATION TABLAB
# =========================================
@app.route('/generate-pdf/<quotation_id>')
def generate_pdf(quotation_id):
    try:
        # Get quotation data
        with open(QUOTATION_FILE, 'r') as f:
            quotations = json.load(f)
        
        quotation = next((q for q in quotations if q['quotation_id'] == quotation_id), None)
        
        if not quotation:
            return jsonify({'success': False, 'error': 'Quotation not found'}), 404
        
        # Get status for behavior control
        status = quotation.get('status', 'draft').lower()
        
        # Get currency code from JSON
        currency_code = quotation.get('currency', 'USD')
        
        # DYNAMIC CURRENCY MAP - Add all currencies you need
        currency_map = {
            'USD': '$', 'EUR': '€', 'GBP': '£', 'JPY': '¥', 'IND': '₹',
            'SGD': 'S$', 'CAD': 'C$', 'AUD': 'A$', 'CHF': 'Fr', 'CNY': '¥',
            'HKD': 'HK$', 'NZD': 'NZ$', 'KRW': '₩', 'MXN': 'Mex$', 'BRL': 'R$',
            'RUB': '₽', 'ZAR': 'R', 'TRY': '₺', 'PLN': 'zł', 'THB': '฿',
            'IDR': 'Rp', 'MYR': 'RM', 'PHP': '₱', 'CZK': 'Kč', 'HUF': 'Ft',
            'ILS': '₪', 'SAR': '﷼', 'AED': 'د.إ', 'SEK': 'kr', 'NOK': 'kr',
            'DKK': 'kr', 'RON': 'lei', 'BGN': 'лв', 'HRK': 'kn', 'ISK': 'kr',
            'TRY': '₺', 'NGN': '₦', 'EGP': 'E£', 'PKR': '₨', 'LKR': 'Rs',
            'NPR': 'रू', 'BDT': '৳', 'VND': '₫', 'ARS': '$', 'CLP': '$',
            'COP': '$', 'PEN': 'S/', 'UYU': '$U', 'PYG': '₲', 'BOB': 'Bs',
            'GTQ': 'Q', 'HNL': 'L', 'NIO': 'C$', 'CRC': '₡', 'PAB': 'B/.'
        }
        
        # Get the correct symbol based on currency code
        currency_symbol = currency_map.get(currency_code, currency_code)
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=72)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C3E50'),
            alignment=1,
            spaceAfter=20
        )
        
        heading_style = ParagraphStyle(
            'Heading2',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#34495E'),
            spaceAfter=10,
            spaceBefore=20
        )
        
        normal_style = styles['Normal']
        
        # Company Header
        elements.append(Paragraph("STACKLY", title_style))
        elements.append(Paragraph("MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008", normal_style))
        elements.append(Paragraph("Phone: +917010792745 ", normal_style))
        elements.append(Paragraph("Email: info@stackly.com", normal_style))

        elements.append(Spacer(1, 20))
        
        # Quotation Title with Status
        status_display = quotation.get('status', 'draft').upper()
        status_color = {
            'DRAFT': colors.orange,
            'SENT': colors.blue,
            'SEND': colors.blue,
            'SUBMITTED': colors.blue,
            'APPROVED': colors.green,
            'REJECTED': colors.red,
            'EXPIRED': colors.HexColor('#FFA500'),
            'CANCELLED': colors.gray
        }.get(status_display, colors.black)
        
        elements.append(Paragraph(f"QUOTATION - {status_display}", ParagraphStyle(
            'Status',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=status_color,
            alignment=1,
            spaceAfter=30
        )))
        
        # ============================================
        # ADD WATERMARK FOR REJECTED/EXPIRED (VIEW-ONLY)
        # ============================================
        if status in ['rejected', 'expired']:
            watermark_text = "⚠️ REJECTED - FOR REFERENCE ONLY ⚠️" if status == 'rejected' else "⚠️ EXPIRED - FOR REFERENCE ONLY ⚠️"
            watermark_color = colors.red if status == 'rejected' else colors.orange
            
            elements.append(Paragraph(
                watermark_text,
                ParagraphStyle(
                    'Watermark',
                    parent=styles['Normal'],
                    fontSize=16,
                    textColor=watermark_color,
                    alignment=1,
                    spaceAfter=20,
                    spaceBefore=10,
                    backColor=colors.lightgrey
                )
            ))
            elements.append(Spacer(1, 10))
        
        # Quotation Info Table
        info_data = [
            ['Quotation Number:', quotation['quotation_id'], 'Date:', quotation.get('quotation_date', '')],
            ['Customer:', quotation.get('customer_name', ''), 'Expiry Date:', quotation.get('expiry_date', '')],
            ['Sales Rep:', quotation.get('sales_rep', ''), 'Currency:', f"{currency_code} ({currency_symbol})"],
            ['PO Reference:', quotation.get('customer_po', 'N/A'), 'Payment Terms:', quotation.get('payment_term', 'N/A')],
        ]
        
        info_table = Table(info_data, colWidths=[100, 150, 100, 150])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 30))
        
        # Items Table
        if quotation.get('items') and len(quotation['items']) > 0:
            elements.append(Paragraph("QUOTATION ITEMS", heading_style))
            
            # Table headers
            table_data = [['S.No', 'Product Name', 'Qty', 'UOM', 'Unit Price', 'Tax %', 'Disc %', 'Total']]
            
            for item in quotation['items']:
                sl_no = str(item.get('sl_no', ''))
                product_name = item.get('product_name', '')
                
                # Get EXACT values from JSON
                quantity = float(item.get('quantity', 0))
                uom = item.get('uom', '')
                unit_price = float(item.get('unit_price', 0))
                tax_percent = float(item.get('tax', 0))
                discount_percent = float(item.get('discount', 0))
                
                # Use stored line total if available
                if 'total' in item and item['total']:
                    line_total = float(item['total'])
                else:
                    # Calculate if not stored
                    line_subtotal = quantity * unit_price
                    discount_amount = line_subtotal * (discount_percent / 100) if discount_percent > 0 else 0
                    line_after_discount = line_subtotal - discount_amount
                    tax_amount = line_after_discount * (tax_percent / 100) if tax_percent > 0 else 0
                    line_total = line_after_discount + tax_amount
                
                table_data.append([
                    sl_no,
                    product_name,
                    f"{quantity:.2f}",
                    uom,
                    f"{currency_symbol}{unit_price:.2f}",
                    f"{tax_percent:.1f}%" if tax_percent > 0 else "-",
                    f"{discount_percent:.1f}%" if discount_percent > 0 else "-",
                    f"{currency_symbol}{line_total:.2f}"
                ])
            
            # Create items table
            items_table = Table(table_data, colWidths=[40, 150, 50, 45, 80, 55, 55, 80])
            items_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
                ('ALIGN', (7, 1), (7, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(items_table)
            
            elements.append(Spacer(1, 20))
            
            # ============================================
            # TAX AND TOTALS SUMMARY - FIXED FOR IND
            # ============================================
            elements.append(Paragraph("TAX AND TOTALS SUMMARY", heading_style))
            
            # Get values from the nested 'totals' object
            totals = quotation.get('totals', {})
            
            # DEBUG: Print totals to see what's coming from JSON
            print(f"Totals for {quotation_id}: {totals}")
            
            # SIMPLIFIED extract_value function - handles IND specifically
            def extract_value(value_str):
                if not value_str:
                    return 0.0
                
                # Convert to string
                cleaned = str(value_str)
                
                # Remove currency symbols - including IND ₹
                cleaned = cleaned.replace('S$', '').replace('$', '').replace('€', '').replace('£', '')
                cleaned = cleaned.replace('¥', '').replace('₹', '').replace('C$', '').replace('A$', '')
                cleaned = cleaned.replace('HK$', '').replace('NZ$', '').replace('Mex$', '').replace('R$', '')
                cleaned = cleaned.replace('₽', '').replace('R', '').replace('₺', '').replace('zł', '')
                cleaned = cleaned.replace('฿', '').replace('Rp', '').replace('RM', '').replace('₱', '')
                cleaned = cleaned.replace('Kč', '').replace('Ft', '').replace('₪', '').replace('﷼', '')
                cleaned = cleaned.replace('د.إ', '').replace('kr', '').replace('lei', '').replace('лв', '')
                cleaned = cleaned.replace('kn', '').replace('₦', '').replace('E£', '').replace('₨', '')
                cleaned = cleaned.replace('Rs', '').replace('रू', '').replace('৳', '').replace('₫', '')
                cleaned = cleaned.replace('S/', '').replace('$U', '').replace('₲', '').replace('Bs', '')
                cleaned = cleaned.replace('Q', '').replace('L', '').replace('C$', '').replace('₡', '')
                cleaned = cleaned.replace('B/.', '')
                
                # Remove commas and spaces
                cleaned = cleaned.replace(',', '').strip()
                
                try:
                    result = float(cleaned)
                    print(f"Extracted value: '{value_str}' -> {result}")  # Debug
                    return result
                except:
                    print(f"Failed to extract: '{value_str}'")
                    return 0.0
            
            # Extract EXACT values from JSON
            subtotal = extract_value(totals.get('subtotal', 0))
            total_tax = extract_value(totals.get('tax_summary', 0))
            shipping_charge = extract_value(totals.get('shipping_charge', 0))
            grand_total = extract_value(totals.get('grand_total', 0))
            
            # Get global discount percentage
            global_discount_percent = 0.0
            if totals.get('global_discount_percent'):
                try:
                    global_discount_percent = float(totals['global_discount_percent'])
                    print(f"Global discount percent: {global_discount_percent}")
                except:
                    global_discount_percent = 0.0
            
            # Get rounding adjustment
            rounding_adjustment = extract_value(totals.get('rounding_adjustment', 0))
            
            # Calculate global discount amount
            global_discount_amount = subtotal * (global_discount_percent / 100) if global_discount_percent > 0 else 0
            
            # Calculate item level discount from items
            total_discount = 0.0
            for item in quotation.get('items', []):
                quantity = float(item.get('quantity', 0))
                unit_price = float(item.get('unit_price', 0))
                discount_percent = float(item.get('discount', 0))
                
                if discount_percent > 0:
                    line_subtotal = quantity * unit_price
                    discount_amount = line_subtotal * (discount_percent / 100)
                    total_discount += discount_amount
            
            # DEBUG: Print all values before creating table
            print(f"subtotal: {subtotal}, shipping: {shipping_charge}, global%: {global_discount_percent}, grand_total: {grand_total}")
            
            # Create summary data with DYNAMIC currency symbol
            summary_data = [
                ['Subtotal:', f"{currency_symbol}{subtotal:.2f}"],
                ['Total Discount (Item Level):', f"{currency_symbol}{total_discount:.2f}"],
                ['Total Tax:', f"{currency_symbol}{total_tax:.2f}"],
            ]
            
            # ✅ ADD SHIPPING CHARGE - Make sure it's added regardless of currency
            if shipping_charge >= 0:
                summary_data.append(['Shipping Charge:', f"{currency_symbol}{shipping_charge:.2f}"])
                print(f"Added Shipping Charge: {shipping_charge}")
            else:
                print(f"Shipping charge is 0 or not found: {shipping_charge}")
            
            # ✅ ADD GLOBAL DISCOUNT - Make sure it's added regardless of currency
            if global_discount_percent >= 0:
                summary_data.append([f'Global Discount ({global_discount_percent:.1f}%):', f"-{currency_symbol}{global_discount_amount:.2f}"])
                print(f"Added Global Discount: {global_discount_percent}%, amount: {global_discount_amount}")
            else:
                print(f"Global discount percent is 0 or not found: {global_discount_percent}")
            
            # Add Rounding Adjustment if not zero
            if rounding_adjustment != 0:
                sign = "+" if rounding_adjustment > 0 else ""
                summary_data.append(['Rounding Adjustment:', f"{sign}{currency_symbol}{abs(rounding_adjustment):.2f}"])
            
            # Add separator
            summary_data.append(['-' * 30, '-' * 15])
            
            # Grand Total - use EXACT value from JSON
            summary_data.append(['GRAND TOTAL:', f"{currency_symbol}{grand_total:.2f}"])
            
            # Create summary table
            summary_table = Table(summary_data, colWidths=[200, 150])
            
            # Table styling
            table_style = [
                ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
                ('FONTSIZE', (0, 0), (-1, -2), 10),
                ('FONTSIZE', (0, -1), (-1, -1), 12),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('BACKGROUND', (0, -2), (1, -2), colors.lightgrey),
                ('BACKGROUND', (0, -1), (1, -1), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, -1), (1, -1), colors.whitesmoke),
                ('FONTWEIGHT', (0, -1), (1, -1), 'BOLD'),
                ('LINEABOVE', (0, -2), (1, -2), 1, colors.black),
                ('LINEBELOW', (0, -2), (1, -2), 1, colors.black),
                ('LINEABOVE', (0, -1), (1, -1), 2, colors.black),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]
            
            summary_table.setStyle(TableStyle(table_style))
            
            # Color code rounding adjustment
            if rounding_adjustment != 0:
                summary_table.setStyle(TableStyle([
                    ('TEXTCOLOR', (1, -3), (1, -3), colors.green if rounding_adjustment > 0 else colors.red),
                ]))
            
            # Right-align the summary table
            summary_container = Table([[summary_table]], colWidths=[350])
            summary_container.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
            ]))
            
            elements.append(summary_container)
        
        elements.append(Spacer(1, 30))
        
        # Terms and Conditions
        elements.append(Paragraph("Terms and Conditions", heading_style))
        terms_text = """
        1. This quotation is valid until the expiry date mentioned above.<br/>
        2. Prices are subject to change without prior notice.<br/>
        3. Payment terms as agreed upon.<br/>
        4. Delivery charges extra if not specified.<br/>
        5. Goods once sold will not be taken back.<br/>
        6. All taxes and duties as applicable.
        """
        elements.append(Paragraph(terms_text, normal_style))
        
        elements.append(Spacer(1, 30))
        
        # Footer
        footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        elements.append(Paragraph(footer_text, normal_style))
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF from buffer
        pdf = buffer.getvalue()
        buffer.close()
        
        # Create response
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        
        # ============================================
        # PROFESSIONAL ERP: STATUS-BASED CONTENT-DISPOSITION
        # ============================================
        if status == 'draft':
            # Draft should not have PDF at all (handled by JS)
            return jsonify({'success': False, 'error': 'PDF not available for draft quotations'}), 403
            
        elif status in ['rejected', 'expired']:
            # REJECTED and EXPIRED - VIEW ONLY (inline with reference filename)
            response.headers['Content-Disposition'] = f'inline; filename=quotation_{quotation_id}_REFERENCE.pdf'
            print(f"📄 View-only PDF for {status} quotation: {quotation_id}")
            
        else:  # send, submitted, approved
            # SENT, SUBMITTED, APPROVED - Full access (attachment for download)
            response.headers['Content-Disposition'] = f'attachment; filename=quotation_{quotation_id}.pdf'
            print(f"📄 Downloadable PDF for {status} quotation: {quotation_id}")
        
        # Add this right after getting totals
        print(f"Raw totals for {quotation_id}: {totals}")
        print(f"shipping_charge raw: {totals.get('shipping_charge')}")
        print(f"global_discount_percent raw: {totals.get('global_discount_percent')}")
        return response
        
    except Exception as e:
        print(f"Error generating PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    



# ===================================================
# OTP GENERATION
# ===================================================

def generate_otp():
    """Generate 6-digit OTP"""
    return ''.join(random.choices(string.digits, k=6))

# ===================================================
# OTP RATE LIMITING FUNCTIONS
# ===================================================

def check_otp_limits(email, quotation_id):
    """
    Check OTP attempt limits
    Returns: (allowed, reason, attempts_left)
    """
    try:
        now = datetime.now()
        key = f"{email}:{quotation_id}"
        
        # Check if user is blocked
        if key in otp_blocked:
            block_info = otp_blocked[key]
            unblock_time = datetime.fromisoformat(block_info['unblock_time'])
            
            if now < unblock_time:
                wait_minutes = int((unblock_time - now).total_seconds() / 60)
                return False, f"Too many attempts. Try again in {wait_minutes} minutes.", 0
            else:
                del otp_blocked[key]
        
        # Get recent failed attempts
        recent_attempts = [a for a in otp_attempts.get(key, []) 
                          if (now - a['timestamp']) < timedelta(minutes=30)]
        
        failed_attempts = [a for a in recent_attempts if not a['success']]
        
        if len(failed_attempts) >= RATE_LIMIT_CONFIG['max_otp_attempts']:
            unblock_time = now + timedelta(minutes=RATE_LIMIT_CONFIG['otp_cooldown_minutes'])
            otp_blocked[key] = {'unblock_time': unblock_time.isoformat()}
            return False, f"Too many failed attempts. Try again after {RATE_LIMIT_CONFIG['otp_cooldown_minutes']} minutes.", 0
        
        attempts_left = RATE_LIMIT_CONFIG['max_otp_attempts'] - len(failed_attempts)
        return True, "Allowed", attempts_left
        
    except Exception as e:
        print(f"Error in check_otp_limits: {e}")
        return True, "Allowed", 5

def record_otp_attempt(email, quotation_id, success):
    """Record an OTP attempt"""
    key = f"{email}:{quotation_id}"
    otp_attempts[key].append({
        'timestamp': datetime.now(),
        'success': success
    })
    
    # Clean up old records
    cutoff = datetime.now() - timedelta(hours=24)
    otp_attempts[key] = [a for a in otp_attempts[key] if a['timestamp'] > cutoff]

def get_otp_attempts_left(email, quotation_id):
    """Get remaining OTP attempts"""
    key = f"{email}:{quotation_id}"
    now = datetime.now()
    
    if key in otp_blocked:
        return 0
    
    recent_attempts = [a for a in otp_attempts.get(key, []) 
                      if (now - a['timestamp']) < timedelta(minutes=30)]
    failed_attempts = [a for a in recent_attempts if not a['success']]
    return RATE_LIMIT_CONFIG['max_otp_attempts'] - len(failed_attempts)

# ===================================================
# OTP RESEND LIMITING
# ===================================================

def check_resend_limits(email, quotation_id):
    """Check if resend is allowed (max 5 per 24 hours)"""
    now = datetime.now()
    key = f"{email}:{quotation_id}"
    
    resends = [a for a in otp_resend_attempts.get(key, []) 
               if (now - a['timestamp']) < timedelta(hours=24)]
    
    if len(resends) >= 5:
        return False, "Maximum resend attempts reached. Try again after 24 hours.", 0
    
    attempts_left = 5 - len(resends)
    return True, "Allowed", attempts_left

def record_resend_attempt(email, quotation_id):
    """Record a resend attempt"""
    key = f"{email}:{quotation_id}"
    otp_resend_attempts[key].append({'timestamp': datetime.now()})
    
    # Clean up
    cutoff = datetime.now() - timedelta(hours=24)
    otp_resend_attempts[key] = [a for a in otp_resend_attempts[key] if a['timestamp'] > cutoff]
    
    return get_resend_attempts_left(email, quotation_id)

def get_resend_attempts_left(email, quotation_id):
    """Get remaining resend attempts"""
    key = f"{email}:{quotation_id}"
    now = datetime.now()
    resends = [a for a in otp_resend_attempts.get(key, []) 
               if (now - a['timestamp']) < timedelta(hours=24)]
    return 5 - len(resends)

# ===================================================
# EMAIL RATE LIMITING
# ===================================================

def check_email_limits(quotation_id, customer_email, recipient_email):
    """
    Check if email can be sent
    Returns: (allowed, reason, requires_approval)
    """
    now = datetime.now()
    key = f"{quotation_id}:{customer_email}"
    
    attempts = email_attempts.get(key, [])
    
    # Per-quotation limit
    quotation_emails = [a for a in attempts if a['quotation_id'] == quotation_id]
    
    if len(quotation_emails) >= RATE_LIMIT_CONFIG['max_emails_per_quotation']:
        return False, f"Maximum {RATE_LIMIT_CONFIG['max_emails_per_quotation']} emails reached", False
    
    # Per-recipient limit
    recipient_emails = [a for a in quotation_emails if a['recipient'] == recipient_email]
    
    if len(recipient_emails) >= RATE_LIMIT_CONFIG['max_emails_per_recipient']:
        return False, f"Already sent {RATE_LIMIT_CONFIG['max_emails_per_recipient']} emails to this recipient", False
    
    # Daily limit
    daily_emails = [a for a in attempts if a['timestamp'].date() == now.date()]
    
    if len(daily_emails) >= RATE_LIMIT_CONFIG['max_daily_emails_per_customer']:
        return False, f"Daily limit of {RATE_LIMIT_CONFIG['max_daily_emails_per_customer']} emails reached", False
    
    # Throttle
    if attempts:
        last_email = attempts[-1]
        time_diff = now - last_email['timestamp']
        min_wait = timedelta(minutes=RATE_LIMIT_CONFIG['min_time_between_emails_minutes'])
        
        if time_diff < min_wait:
            wait_minutes = RATE_LIMIT_CONFIG['min_time_between_emails_minutes'] - (time_diff.seconds // 60)
            return False, f"Please wait {wait_minutes} minutes between emails", False
    
    # Check if approval required
    requires_approval = len(quotation_emails) >= RATE_LIMIT_CONFIG['requires_approval_after']
    
    return True, "Allowed", requires_approval

def record_email_sent(quotation_id, customer_email, recipient_email, approved=False):
    """Record that an email was sent"""
    key = f"{quotation_id}:{customer_email}"
    
    email_attempts[key].append({
        'quotation_id': quotation_id,
        'recipient': recipient_email,
        'timestamp': datetime.now(),
        'approved': approved
    })
    
    # Clean up old records
    cutoff = datetime.now() - timedelta(days=30)
    email_attempts[key] = [a for a in email_attempts[key] if a['timestamp'] > cutoff]

def get_email_count(quotation_id, customer_email):
    """Get number of emails sent for this quotation"""
    key = f"{quotation_id}:{customer_email}"
    return len(email_attempts.get(key, []))

# ===================================================
# SEND QUOTATION OTP EMAIL (signup uses send_otp_email above — do not shadow it)
# ===================================================

def send_quotation_otp_email(email, otp, quotation_id=None):
    """Send OTP via email for quotation / email flow (not signup)."""
    try:
        print(f"📧 Sending quotation OTP to {email}")
        
        msg = MIMEMultipart()
        msg['Subject'] = f"Your OTP for Quotation {quotation_id}" if quotation_id else "Your OTP for Quotation"
        msg['From'] = SENDER_EMAIL
        msg['To'] = email
        
        body = f"""
        Your OTP for verification is: {otp}
        
        This OTP is valid for {OTP_EXPIRY_MINUTES} minutes.
        
        Please enter this OTP to complete your quotation request.
        
        If you didn't request this, please ignore this email.
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Try port 587 with TLS
        try:
            server = smtplib.SMTP(SMTP_SERVER, 587, timeout=30)
            server.starttls()
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.send_message(msg)
            server.quit()
            print(f"✅ OTP sent successfully")
            return True
        except Exception as e:
            print(f"❌ Port 587 failed: {e}")
            
            # Try port 465 with SSL
            try:
                context = ssl.create_default_context()
                server = smtplib.SMTP_SSL(SMTP_SERVER, 465, context=context, timeout=30)
                server.login(SENDER_EMAIL, SENDER_PASSWORD)
                server.send_message(msg)
                server.quit()
                print(f"✅ OTP sent via SSL")
                return True
            except Exception as e2:
                print(f"❌ Both ports failed: {e2}")
                # For development, just log the OTP
                print(f"📧 [DEV MODE] OTP for {email}: {otp}")
                return True
            
    except Exception as e:
        print(f"❌ Error sending OTP: {e}")
        # For development, just log the OTP
        print(f"📧 [DEV MODE] OTP for {email}: {otp}")
        return True


# ===================================================
# SINGLE SOURCE OF TRUTH - ONE PDF GENERATOR FOR ALL
# ===================================================

def generate_quotation_pdf(quotation, quotation_id):
    """Single PDF generator used by both route and email"""
    try:
        import io
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch, mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from datetime import datetime
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=72)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Get currency code and symbol
        currency_code = quotation.get('currency', 'USD')
        currency_map = {
            'USD': '$', 'EUR': '€', 'GBP': '£', 'JPY': '¥', 'IND': '₹',
            'SGD': 'S$', 'CAD': 'C$', 'AUD': 'A$', 'CHF': 'Fr', 'CNY': '¥',
            'HKD': 'HK$', 'NZD': 'NZ$', 'KRW': '₩', 'MXN': 'Mex$', 'BRL': 'R$',
            'RUB': '₽', 'ZAR': 'R', 'TRY': '₺', 'PLN': 'zł', 'THB': '฿',
            'IDR': 'Rp', 'MYR': 'RM', 'PHP': '₱', 'CZK': 'Kč', 'HUF': 'Ft',
            'ILS': '₪', 'SAR': '﷼', 'AED': 'د.إ', 'SEK': 'kr', 'NOK': 'kr',
            'DKK': 'kr', 'RON': 'lei', 'BGN': 'лв', 'HRK': 'kn', 'ISK': 'kr',
            'TRY': '₺', 'NGN': '₦', 'EGP': 'E£', 'PKR': '₨', 'LKR': 'Rs',
            'NPR': 'रू', 'BDT': '৳', 'VND': '₫', 'ARS': '$', 'CLP': '$',
            'COP': '$', 'PEN': 'S/', 'UYU': '$U', 'PYG': '₲', 'BOB': 'Bs',
            'GTQ': 'Q', 'HNL': 'L', 'NIO': 'C$', 'CRC': '₡', 'PAB': 'B/.'
        }
        currency_symbol = currency_map.get(currency_code, currency_code)
        
        # Helper function to extract numeric values
        def extract_value(value_str):
            if not value_str:
                return 0.0
            cleaned = str(value_str)
            # Remove all currency symbols
            for symbol in currency_map.values():
                cleaned = cleaned.replace(symbol, '')
            cleaned = cleaned.replace(',', '').strip()
            try:
                return float(cleaned)
            except:
                return 0.0
        
        # Get totals
        totals = quotation.get('totals', {})
        
        # Extract all values
        subtotal_value = extract_value(totals.get('subtotal', 0))
        total_tax = extract_value(totals.get('tax_summary', 0))
        shipping_charge = extract_value(totals.get('shipping_charge', 0))
        grand_total_value = extract_value(totals.get('grand_total', 0))
        
        # Get global discount
        global_discount_percent = 0.0
        if totals.get('global_discount_percent'):
            try:
                global_discount_percent = float(totals['global_discount_percent'])
            except:
                global_discount_percent = 0.0
        
        # Get rounding adjustment
        rounding_adjustment = extract_value(totals.get('rounding_adjustment', 0))
        
        # Calculate global discount amount
        global_discount_amount = subtotal_value * (global_discount_percent / 100) if global_discount_percent > 0 else 0
        
        # Calculate item level discount
        total_discount = 0.0
        for item in quotation.get('items', []):
            quantity = float(item.get('quantity', 0))
            unit_price = float(item.get('unit_price', 0))
            discount_percent = float(item.get('discount', 0))
            if discount_percent > 0:
                line_subtotal = quantity * unit_price
                discount_amount = line_subtotal * (discount_percent / 100)
                total_discount += discount_amount
        
        # Styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C3E50'),
            alignment=1,
            spaceAfter=20
        )
        
        heading_style = ParagraphStyle(
            'Heading2',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#34495E'),
            spaceAfter=10,
            spaceBefore=20
        )
        
        normal_style = styles['Normal']
        
        # Company Header
        elements.append(Paragraph("STACKLY", title_style))
        elements.append(Paragraph("Address:MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008", normal_style))
        elements.append(Paragraph("Phone: + 917010792745", normal_style))
        elements.append(Paragraph("Eamil: info@stackly.com", normal_style))

        elements.append(Spacer(1, 20))
        
        # Quotation Title with Status
        status = quotation.get('status', 'draft').upper()
        status_color = {
            'DRAFT': colors.orange,
            'SENT': colors.blue,
            'SEND': colors.blue,
            'SUBMITTED': colors.blue,
            'APPROVED': colors.green,
            'REJECTED': colors.red,
            'EXPIRED': colors.HexColor('#FFA500'),
            'CANCELLED': colors.gray
        }.get(status, colors.black)
        
        elements.append(Paragraph(f"QUOTATION - {status}", ParagraphStyle(
            'Status',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=status_color,
            alignment=1,
            spaceAfter=30
        )))
        
        # Quotation Info Table
        info_data = [
            ['Quotation Number:', quotation['quotation_id'], 'Date:', quotation.get('quotation_date', '')],
            ['Customer:', quotation.get('customer_name', ''), 'Expiry Date:', quotation.get('expiry_date', '')],
            ['Sales Rep:', quotation.get('sales_rep', ''), 'Currency:', f"{currency_code} ({currency_symbol})"],
            ['PO Reference:', quotation.get('customer_po', 'N/A'), 'Payment Terms:', quotation.get('payment_term', 'N/A')],
        ]
        
        info_table = Table(info_data, colWidths=[100, 150, 100, 150])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 30))
        
        # Items Table
        if quotation.get('items') and len(quotation['items']) > 0:
            elements.append(Paragraph("QUOTATION ITEMS", heading_style))
            
            # Table headers
            table_data = [['S.No', 'Product Name', 'Qty', 'UOM', 'Unit Price', 'Tax %', 'Disc %', 'Total']]
            
            for idx, item in enumerate(quotation['items'], 1):
                sl_no = str(idx)
                product_name = item.get('product_name', '')
                
                # Get EXACT values from JSON
                quantity = float(item.get('quantity', 0))
                uom = item.get('uom', '')
                unit_price = float(item.get('unit_price', 0))
                tax_percent = float(item.get('tax', 0))
                discount_percent = float(item.get('discount', 0))
                
                # Use stored line total if available
                if 'total' in item and item['total']:
                    line_total = float(extract_value(item['total']))
                else:
                    # Calculate if not stored
                    line_subtotal = quantity * unit_price
                    discount_amount = line_subtotal * (discount_percent / 100) if discount_percent > 0 else 0
                    line_after_discount = line_subtotal - discount_amount
                    tax_amount = line_after_discount * (tax_percent / 100) if tax_percent > 0 else 0
                    line_total = line_after_discount + tax_amount
                
                table_data.append([
                    sl_no,
                    product_name,
                    f"{quantity:.2f}",
                    uom,
                    f"{currency_symbol}{unit_price:.2f}",
                    f"{tax_percent:.1f}%" if tax_percent > 0 else "-",
                    f"{discount_percent:.1f}%" if discount_percent > 0 else "-",
                    f"{currency_symbol}{line_total:.2f}"
                ])
            
            # Create items table
            items_table = Table(table_data, colWidths=[40, 150, 50, 45, 80, 55, 55, 80])
            items_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
                ('ALIGN', (7, 1), (7, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(items_table)
            
            elements.append(Spacer(1, 20))
            
            # TAX AND TOTALS SUMMARY
            elements.append(Paragraph("TAX AND TOTALS SUMMARY", heading_style))
            
            # Create summary data
            summary_data = [
                ['Subtotal:', f"{currency_symbol}{subtotal_value:.2f}"],
                ['Total Discount (Item Level):', f"{currency_symbol}{total_discount:.2f}"],
                ['Total Tax:', f"{currency_symbol}{total_tax:.2f}"],
            ]
            
            # Add Shipping Charge
            if shipping_charge >= 0:
                summary_data.append(['Shipping Charge:', f"{currency_symbol}{shipping_charge:.2f}"])
            
            # Add Global Discount
            if global_discount_percent >= 0:
                summary_data.append([f'Global Discount ({global_discount_percent:.1f}%):', f"-{currency_symbol}{global_discount_amount:.2f}"])
            
            # Add Rounding Adjustment
            if rounding_adjustment != 0:
                sign = "+" if rounding_adjustment > 0 else ""
                summary_data.append(['Rounding Adjustment:', f"{sign}{currency_symbol}{abs(rounding_adjustment):.2f}"])
            
            # Add separator
            summary_data.append(['-' * 30, '-' * 15])
            
            # Grand Total
            summary_data.append(['GRAND TOTAL:', f"{currency_symbol}{grand_total_value:.2f}"])
            
            # Create summary table
            summary_table = Table(summary_data, colWidths=[200, 150])
            
            # Table styling
            table_style = [
                ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
                ('FONTSIZE', (0, 0), (-1, -2), 10),
                ('FONTSIZE', (0, -1), (-1, -1), 12),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('BACKGROUND', (0, -2), (1, -2), colors.lightgrey),
                ('BACKGROUND', (0, -1), (1, -1), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, -1), (1, -1), colors.whitesmoke),
                ('FONTWEIGHT', (0, -1), (1, -1), 'BOLD'),
                ('LINEABOVE', (0, -2), (1, -2), 1, colors.black),
                ('LINEBELOW', (0, -2), (1, -2), 1, colors.black),
                ('LINEABOVE', (0, -1), (1, -1), 2, colors.black),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]
            
            summary_table.setStyle(TableStyle(table_style))
            
            # Color code rounding adjustment
            if rounding_adjustment != 0:
                summary_table.setStyle(TableStyle([
                    ('TEXTCOLOR', (1, -3), (1, -3), colors.green if rounding_adjustment > 0 else colors.red),
                ]))
            
            # Right-align the summary table
            summary_container = Table([[summary_table]], colWidths=[350])
            summary_container.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
            ]))
            
            elements.append(summary_container)
        
        elements.append(Spacer(1, 30))
        
        # Terms and Conditions
        elements.append(Paragraph("Terms and Conditions", heading_style))
        terms_text = """
        1. This quotation is valid until the expiry date mentioned above.<br/>
        2. Prices are subject to change without prior notice.<br/>
        3. Payment terms as agreed upon.<br/>
        4. Delivery charges extra if not specified.<br/>
        5. Goods once sold will not be taken back.<br/>
        6. All taxes and duties as applicable.
        """
        elements.append(Paragraph(terms_text, normal_style))
        
        elements.append(Spacer(1, 30))
        
        # Footer
        footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        elements.append(Paragraph(footer_text, normal_style))
        
        # Build PDF
        doc.build(elements)
        
        pdf = buffer.getvalue()
        buffer.close()
        
        return pdf
        
    except Exception as e:
        print(f"Error generating PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
from email.message import EmailMessage



import json

def get_quotation_data(quotation_id):
    """
    Fetch a quotation by ID from your JSON file.
    Replace 'quotations.json' with the path to your JSON database.
    """
    try:
        with open("quotations.json", "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get(quotation_id, None)
    except FileNotFoundError:
        return None
    except json.JSONDecodeError:
        return None

# Your existing generate_quotation_pdf function here
# Your get_quotation_data function to fetch from JSON

@app.route("/send-quotation/<quotation_id>", methods=["POST"])
def send_quotation(quotation_id):
    # Fetch quotation data from JSON
    quotation = get_quotation_data(quotation_id)
    if not quotation:
        return jsonify({"success": False, "message": "Quotation not found"}), 404

    # Generate PDF
    pdf_bytes = generate_quotation_pdf(quotation, quotation_id)
    if not pdf_bytes:
        return jsonify({"success": False, "message": "Error generating PDF"}), 500

    # Get customer email from JSON
    customer_email = quotation.get("customer_email")
    if not customer_email:
        return jsonify({"success": False, "message": "Customer email not found"}), 400

    try:
        # Create email
        msg = EmailMessage()
        msg['Subject'] = f"Quotation {quotation_id}"
        msg['From'] = 'yourcompany@example.com'
        msg['To'] = customer_email
        msg.set_content(f"Dear {quotation.get('customer_name', 'Customer')},\n\nPlease find attached your quotation {quotation_id}.\n\nBest regards,\nYour Company")

        # Attach PDF
        msg.add_attachment(pdf_bytes, maintype='application', subtype='pdf', filename=f"Quotation_{quotation_id}.pdf")

        # Send email via SMTP
        with smtplib.SMTP('smtp.example.com', 587) as smtp:
            smtp.starttls()
            smtp.login('your_email@example.com', 'your_password')
            smtp.send_message(msg)

        return jsonify({"success": True, "message": f"Quotation sent to {customer_email}"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


def log_email_sent(quotation_id, recipient, status):
    log_entry = {
        'timestamp': datetime.now().isoformat(),
        'quotation_id': quotation_id,
        'recipient': recipient,
        'status': status
    }
    
    # Append to email log file
    with open('email_log.json', 'a') as f:
        f.write(json.dumps(log_entry) + '\n')        
# ===================================================
# SEND QUOTATION EMAIL WITH PDF
# ===================================================
def send_quotation_email_internal(quotation_id, recipient_email):
    """Send quotation email with PDF attachment"""
    try:
        # Get quotation data
        with open(QUOTATION_FILE, 'r') as f:
            quotations = json.load(f)
        
        quotation = next((q for q in quotations if q['quotation_id'] == quotation_id), None)
        
        if not quotation:
            return {'success': False, 'error': 'Quotation not found'}
        
        # Generate PDF attachment using the SAME common function
        pdf_attachment = generate_quotation_pdf(quotation, quotation_id)
        
        # Generate HTML email
        html_body = render_template(
            'email_quotation.html', 
            quotation=quotation,
            now=datetime.now(),
            recipient_email=recipient_email
        )
        
        # Create email
        msg = MIMEMultipart('mixed')
        msg['Subject'] = f"Quotation {quotation_id} from Your Company"
        msg['From'] = SENDER_EMAIL
        msg['To'] = recipient_email
        
        # Plain text version
        text_body = f"""
        Hi {quotation.get('customer_name', 'Customer')},
        
        Your quotation {quotation_id} has been generated.
        
        Please find the attached PDF.
        """
        
        msg_alternative = MIMEMultipart('alternative')
        msg_alternative.attach(MIMEText(text_body, 'plain'))
        msg_alternative.attach(MIMEText(html_body, 'html'))
        msg.attach(msg_alternative)
        
        # Attach PDF (generated by common function)
        if pdf_attachment:
            attachment = MIMEApplication(pdf_attachment, _subtype="pdf")
            attachment.add_header('Content-Disposition', 'attachment', 
                                filename=f"Quotation_{quotation_id}.pdf")
            msg.attach(attachment)
        
        # Send email
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.send_message(msg)
        
        print(f"✅ Quotation {quotation_id} sent to {recipient_email}")
        return {'success': True, 'message': f'Quotation sent to {recipient_email}'}
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}
# ===================================================
# API ROUTES
# ===================================================

@app.route('/api/check-email-limit', methods=['POST'])
def api_check_email_limit():
    """Check if email can be sent"""
    try:
        data = request.json
        quotation_id = data.get('quotation_id')
        recipient_email = data.get('recipient')
        
        customer_email = session.get('user_email', 'unknown@example.com')
        
        allowed, reason, requires_approval = check_email_limits(
            quotation_id, customer_email, recipient_email
        )
        
        return jsonify({
            'success': True,
            'allowed': allowed,
            'reason': reason if not allowed else None,
            'requires_approval': requires_approval
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/send-quotation-email', methods=['POST'])
def api_send_quotation_email():
    """API endpoint to send quotation email directly (no OTP modal)."""
    try:
        data = request.get_json(silent=True) or {}
        quotation_id = data.get('quotation_id')
        email = (data.get('email') or '').strip()

        if not quotation_id or not email:
            return jsonify({'success': False, 'error': 'Missing quotation ID or email'}), 400

        # Basic email format check
        if '@' not in email or '.' not in email.split('@')[-1]:
            return jsonify({'success': False, 'error': 'Invalid email address'}), 400

        # Use existing logic (limits + sending) via internal helper
        result = send_quotation_email_internal(quotation_id, email)
        if result.get('success'):
            return jsonify({'success': True, 'message': 'Quotation sent successfully'})
        return jsonify({'success': False, 'error': result.get('error')}), 500
    except Exception as e:
        print(f"❌ Error in api_send_quotation_email: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/request-approval', methods=['POST'])
def api_request_approval():
    """Request manager approval"""
    try:
        data = request.json
        quotation_id = data.get('quotation_id')
        recipient_email = data.get('recipient')
        reason = data.get('reason')
        
        print(f"📨 APPROVAL REQUESTED: Quotation {quotation_id}, Recipient {recipient_email}, Reason: {reason}")
        
        return jsonify({'success': True, 'message': 'Approval request sent'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/otp/send', methods=['POST'])
def api_send_otp():
    """Send OTP email"""
    try:
        data = request.json
        email = data.get('email')
        quotation_id = data.get('quotation_id')
        
        if not email or '@' not in email:
            return jsonify({'success': False, 'error': 'Invalid email'}), 400
        
        # Check OTP limits
        allowed, reason, attempts_left = check_otp_limits(email, quotation_id)
        
        if not allowed:
            return jsonify({'success': False, 'error': reason, 'attempts_left': 0}), 429
        
        # Generate and store OTP
        otp = generate_otp()
        
        session[f'otp_{email}'] = {
            'otp': otp,
            'quotation_id': quotation_id,
            'created_at': datetime.now().isoformat()
        }
        
        # Send OTP email
        result = send_quotation_otp_email(email, otp, quotation_id)
        
        if result:
            record_otp_attempt(email, quotation_id, True)
            return jsonify({
                'success': True, 
                'message': 'OTP sent successfully',
                'attempts_left': attempts_left
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to send OTP'}), 500
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/otp/verify', methods=['POST'])
def api_verify_otp():
    """Verify OTP and send quotation"""
    try:
        data = request.json
        email = data.get('email')
        otp = data.get('otp')
        quotation_id = data.get('quotation_id')
        
        # Check OTP limits
        allowed, reason, attempts_left = check_otp_limits(email, quotation_id)
        if not allowed:
            return jsonify({'success': False, 'error': reason}), 429
        
        otp_data = session.get(f'otp_{email}')
        
        if not otp_data:
            record_otp_attempt(email, quotation_id, False)
            return jsonify({'success': False, 'error': 'OTP not found'}), 400
        
        # Check expiry
        created = datetime.fromisoformat(otp_data['created_at'])
        if datetime.now() - created > timedelta(minutes=OTP_EXPIRY_MINUTES):
            session.pop(f'otp_{email}', None)
            record_otp_attempt(email, quotation_id, False)
            return jsonify({'success': False, 'error': 'OTP expired'}), 400
        
        # Verify OTP
        if otp_data['otp'] != otp:
            record_otp_attempt(email, quotation_id, False)
            new_attempts_left = get_otp_attempts_left(email, quotation_id)
            
            if new_attempts_left <= 0:
                return jsonify({'success': False, 'error': 'Maximum attempts exceeded'}), 429
            
            return jsonify({
                'success': False, 
                'error': f'Invalid OTP. {new_attempts_left} attempts left.'
            }), 400
        
        # OTP verified
        record_otp_attempt(email, quotation_id, True)
        session.pop(f'otp_{email}', None)
        
        # Check email limits
        customer_email = session.get('user_email', 'unknown@example.com')
        allowed, reason, requires_approval = check_email_limits(
            quotation_id, customer_email, email
        )
        
        if not allowed:
            return jsonify({'success': False, 'error': reason}), 400
        
        # Send quotation email
        result = send_quotation_email_internal(quotation_id, email)
        
        if result.get('success'):
            record_email_sent(quotation_id, customer_email, email)
            return jsonify({'success': True, 'message': 'Quotation sent successfully'})
        else:
            return jsonify({'success': False, 'error': result.get('error')}), 500
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/otp/resend', methods=['POST'])
def api_resend_otp():
    """Resend OTP"""
    try:
        data = request.json
        email = data.get('email')
        quotation_id = data.get('quotation_id')
        
        # Check resend limits
        allowed, reason, attempts_left = check_resend_limits(email, quotation_id)
        
        if not allowed:
            return jsonify({'success': False, 'error': reason}), 429
        
        # Generate new OTP
        otp = generate_otp()
        
        session[f'otp_{email}'] = {
            'otp': otp,
            'quotation_id': quotation_id,
            'created_at': datetime.now().isoformat()
        }
        
        # Send OTP email
        result = send_quotation_otp_email(email, otp, quotation_id)
        
        if result:
            remaining = record_resend_attempt(email, quotation_id)
            return jsonify({
                'success': True,
                'message': 'OTP resent successfully',
                'resend_attempts_left': remaining
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to resend OTP'}), 500
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get-email-count/<quotation_id>')
def get_email_count_route(quotation_id):
    """Get email count for a quotation"""
    try:
        customer_email = session.get('user_email', 'unknown@example.com')
        count = get_email_count(quotation_id, customer_email)
        
        return jsonify({
            'success': True,
            'count': count,
            'max': RATE_LIMIT_CONFIG['max_emails_per_quotation']
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ===================================================
# DIAGNOSTIC ROUTES
# ===================================================

@app.route('/test-smtp-connection')
def test_smtp_connection():
    """Test SMTP connection"""
    results = []
    
    try:
        import socket
        ip = socket.gethostbyname('smtp.gmail.com')
        results.append(f"✅ DNS resolved: {ip}")
    except Exception as e:
        results.append(f"❌ DNS failed: {e}")
    
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587, timeout=10)
        server.starttls()
        results.append("✅ Port 587 OK")
        server.quit()
    except Exception as e:
        results.append(f"❌ Port 587 failed: {e}")
    
    return "<br>".join(results)









def find_quotation_by_id(quotations, quotation_id):
    """Find quotation by ID"""
    for i, q in enumerate(quotations):
        if q.get('quotation_id') == quotation_id:
            return i, q
    return None, None

# ===================================================
# UPDATE SINGLE QUOTATION STATUS
# ===================================================

@app.route('/update-quotation-status', methods=['POST'])
def update_quotation_status():
    """
    Update the status of a single quotation in your existing JSON file
    Expected JSON payload:
    {
        "quotation_id": "QA-0001",
        "status": "expired",
        "status_date": "2026-02-20T10:30:00.000Z",
        "rejection_reason": "Auto-expired",
        "status_history": {
            "status": "expired",
            "date": "2026-02-20T10:30:00.000Z",
            "user": "System",
            "notes": "Auto-expired"
        }
    }
    """
    try:
        data = request.json
        quotation_id = data.get('quotation_id')
        new_status = data.get('status')
        status_date = data.get('status_date', datetime.now().isoformat())
        rejection_reason = data.get('rejection_reason', '')
        status_history = data.get('status_history', {})
        
        # Validate required fields
        if not quotation_id:
            return jsonify({
                'success': False,
                'error': 'quotation_id is required',
                'code': 400
            }), 400
            
        if not new_status:
            return jsonify({
                'success': False,
                'error': 'status is required',
                'code': 400
            }), 400
        
        # Load quotations from your existing JSON file
        quotations = load_quotations()
        
        if not quotations:
            return jsonify({
                'success': False,
                'error': 'No quotations found in file',
                'code': 404
            }), 404
        
        # Find the quotation
        index, quotation = find_quotation_by_id(quotations, quotation_id)
        
        if index is None:
            return jsonify({
                'success': False,
                'error': f'Quotation {quotation_id} not found',
                'code': 404,
                'quotation_id': quotation_id
            }), 404
        
        # Store old status for logging
        old_status = quotation.get('status', 'unknown')
        
        # Update status
        quotations[index]['status'] = new_status
        quotations[index]['status_date'] = status_date
        quotations[index]['rejection_reason'] = rejection_reason
        quotations[index]['last_updated'] = datetime.now().isoformat()
        
        # Update status history
        if 'status_history' not in quotations[index]:
            quotations[index]['status_history'] = []
        elif quotations[index]['status_history'] is None:
            quotations[index]['status_history'] = []
        
        # Add new history entry
        new_history_entry = {
            'status': new_status,
            'date': status_date,
            'user': status_history.get('user', 'System'),
            'notes': status_history.get('notes', f'Quotation {new_status}')
        }
        quotations[index]['status_history'].append(new_history_entry)
        
        # Save back to JSON file
        if save_quotations(quotations):
            print(f"✅ Quotation {quotation_id} status updated from {old_status} to {new_status}")
            
            return jsonify({
                'success': True,
                'message': f'Quotation {quotation_id} updated to {new_status}',
                'quotation_id': quotation_id,
                'old_status': old_status,
                'new_status': new_status,
                'updated_at': status_date
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to save to JSON file',
                'code': 500
            }), 500
        
    except Exception as e:
        print(f"❌ Error updating quotation status: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'code': 500
        }), 500



# ===================================================
# EXPIRED QUOTATION CHECK - BACKEND
# ===================================================

def check_and_update_expired_quotations():
    """
    Check all quotations and update expired ones in JSON
    This runs when loading the quotations list
    """
    
    # Use the existing QUOTATION_FILE variable
    json_file_path = QUOTATION_FILE  # ← USING YOUR EXISTING VARIABLE
    
    # Check if file exists
    if not os.path.exists(json_file_path):
        print(f"⚠️ Quotation file not found: {json_file_path}")
        return []
    
    try:
        # Load existing quotations
        with open(json_file_path, 'r') as file:
            quotations = json.load(file)
    except json.JSONDecodeError:
        print("❌ Error reading JSON file")
        return []
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        return []
    
    # Get today's date
    today = datetime.now().date()
    updated = False
    
    # Check each quotation
    for quotation in quotations:
        # Get current status
        current_status = quotation.get('status', '').lower()
        
        # Only check Sent quotations (send, sent, submitted)
        if current_status in ['send', 'sent', 'submitted']:
            
            # Get expiry date
            expiry_date_str = quotation.get('expiry_date')
            if not expiry_date_str:
                continue
            
            try:
                # Parse expiry date
                expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
                
                # Check if expired
                if expiry_date < today:
                    print(f"✅ Quotation {quotation.get('quotation_id')} expired - updating")
                    
                    # Update status
                    quotation['status'] = 'expired'
                    
                    # Add to history
                    if 'status_history' not in quotation:
                        quotation['status_history'] = []
                    
                    quotation['status_history'].append({
                        'status': 'expired',
                        'date': today.strftime('%Y-%m-%d'),
                        'time': datetime.now().strftime('%H:%M:%S'),
                        'reason': 'Auto-expired',
                        'notes': f'Expired on {today.strftime("%Y-%m-%d")} (valid until {expiry_date_str})'
                    })
                    
                    # Add expired date
                    quotation['expired_date'] = today.strftime('%Y-%m-%d')
                    
                    updated = True
                    
            except ValueError as e:
                print(f"❌ Date parsing error for {quotation.get('quotation_id')}: {e}")
                continue
            except Exception as e:
                print(f"❌ Error processing {quotation.get('quotation_id')}: {e}")
                continue
    
    # Save changes if any
    if updated:
        try:
            with open(json_file_path, 'w') as file:
                json.dump(quotations, file, indent=2)
            print(f"✅ Expired quotations updated in {json_file_path}")
        except Exception as e:
            print(f"❌ Error saving JSON: {e}")
    
    return quotations


# ===================================================
# API ENDPOINTS
# ===================================================

@app.route('/api/quotations', methods=['GET'])
def get_quotations():
    """
    Get all quotations with automatic expiry check
    This runs every time user opens the quotations list
    """
    try:
        quotations = check_and_update_expired_quotations()
        return jsonify({
            'success': True,
            'items': quotations,
            'count': len(quotations)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/get-quotation/<quotation_id>', methods=['GET'])
def get_single_quotation(quotation_id):
    """
    Get single quotation and check expiry
    """
    try:
        # First update all expired (optional but safe)
        quotations = check_and_update_expired_quotations()
        
        # Find the specific quotation
        quotation = None
        for q in quotations:
            if q.get('quotation_id') == quotation_id:
                quotation = q
                break
        
        if not quotation:
            return jsonify({
                'success': False,
                'error': 'Quotation not found'
            }), 404
        
        return jsonify({
            'success': True,
            'quotation': quotation
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/check-expired-now', methods=['POST'])
def manual_expiry_check():
    """
    Manual endpoint to trigger expiry check
    Can be called from frontend button
    """
    try:
        quotations = check_and_update_expired_quotations()
        
        # Count expired
        expired_count = sum(1 for q in quotations if q.get('status') == 'expired')
        
        return jsonify({
            'success': True,
            'message': f'Expiry check completed',
            'expired_count': expired_count,
            'total': len(quotations)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ===================================================
# OPTIONAL: Daily Scheduler (Run via cron)
# ===================================================

def daily_expiry_check():
    """
    Run this once per day via cron job
    Example cron: 0 0 * * * python3 daily_expiry.py
    """
    print(f"🕒 Daily expiry check started at {datetime.now()}")
    
    quotations = check_and_update_expired_quotations()
    
    expired = [q for q in quotations if q.get('status') == 'expired']
    sent = [q for q in quotations if q.get('status') in ['send', 'sent', 'submitted']]
    
    print(f"📊 Summary:")
    print(f"   - Total quotations: {len(quotations)}")
    print(f"   - Still valid (Sent): {len(sent)}")
    print(f"   - Expired: {len(expired)}")
    print(f"✅ Daily expiry check completed")

# Run daily check if script executed directly
if __name__ == '__main__':
    daily_expiry_check()



# ===================================================
# SOFT-C INTEGRATION
# ===================================================

@app.route('/sync-softc', methods=['POST'])
def sync_softc():
    try:
        data = request.json
        quotation_id = data.get('quotation_id')
        
        # Get quotation data
        with open(QUOTATION_FILE, 'r') as f:
            quotations = json.load(f)
        
        quotation = next((q for q in quotations if q['quotation_id'] == quotation_id), None)
        
        if not quotation:
            return jsonify({'success': False, 'error': 'Quotation not found'}), 404
        
        # 👇 DIRECT ASSIGNMENT - No .env file needed
        softc_api_url = "https://api.soft-c.com/sync"
        softc_api_key = "your-actual-api-key-here"  # Replace with your real key
        
        # Prepare data for Soft-C
        softc_data = {
            "quotation_number": quotation['quotation_id'],
            "customer_name": quotation.get('customer_name'),
            "date": quotation.get('quotation_date'),
            "expiry_date": quotation.get('expiry_date'),
            "currency": quotation.get('currency'),
            "items": quotation.get('items', []),
            "total": quotation.get('grand_total', '0.00'),
            "status": quotation.get('status')
        }
        
        # Send to Soft-C API
        headers = {
            'Authorization': f'Bearer {softc_api_key}',
            'Content-Type': 'application/json'
        }
        
        response = requests.post(softc_api_url, json=softc_data, headers=headers, timeout=10)
        
        if response.status_code == 200:
            return jsonify({'success': True})
        else:
            return jsonify({
                'success': False, 
                'error': f'Soft-C API error: {response.status_code}'
            }), 500
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
# ===================================================
# TEST PDF ROUTE
# ===================================================

@app.route('/test-pdf')
def test_pdf():
    """Simple test to verify PDF generation is working"""
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph("PDF Generation Test", styles['Title']))
        elements.append(Paragraph("If you can see this, ReportLab is working!", styles['Normal']))
        
        doc.build(elements)
        
        pdf = buffer.getvalue()
        buffer.close()
        
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'inline; filename=test.pdf'
        
        return response
    except Exception as e:
        return f"Error: {str(e)}"

@app.route('/delete-quotation/<quotation_id>', methods=['DELETE'])
def delete_quotation(quotation_id):
    """Delete a quotation (only allowed for draft status)"""
    try:
        # Load quotations from JSON file
        with open(QUOTATION_FILE, 'r') as f:
            quotations = json.load(f)
        
        # Find the quotation
        quotation_to_delete = None
        for q in quotations:
            if q.get('quotation_id') == quotation_id:
                quotation_to_delete = q
                break
        
        if not quotation_to_delete:
            return jsonify({
                'success': False,
                'error': 'Quotation not found'
            }), 404
        
        # Check if status is draft (only draft can be deleted)
        if quotation_to_delete.get('status') != 'draft':
            return jsonify({
                'success': False,
                'error': 'Only draft quotations can be deleted'
            }), 403
        
        # Remove the quotation
        quotations = [q for q in quotations if q.get('quotation_id') != quotation_id]
        
        # Save back to JSON file
        with open(QUOTATION_FILE, 'w') as f:
            json.dump(quotations, f, indent=2)
        
        # Also delete associated comments
        try:
            with open(COMMENTS_FILE, 'r') as f:
                comments = json.load(f)
            
            if quotation_id in comments:
                del comments[quotation_id]
                
                with open(COMMENTS_FILE, 'w') as f:
                    json.dump(comments, f, indent=2)
        except FileNotFoundError:
            pass  # Comments file doesn't exist, ignore
        
        # Delete associated attachments
        try:
            metadata_file = os.path.join(ATTACHMENTS_FOLDER, 'metadata.json')
            if os.path.exists(metadata_file):
                with open(metadata_file, 'r') as f:
                    attachments = json.load(f)
                
                # Filter out attachments for this quotation
                attachments_to_keep = [a for a in attachments if a.get('quotation_id') != quotation_id]
                
                # Delete physical files
                for a in attachments:
                    if a.get('quotation_id') == quotation_id:
                        file_path = os.path.join(ATTACHMENTS_FOLDER, a.get('stored_filename', ''))
                        if os.path.exists(file_path):
                            os.remove(file_path)
                
                # Save updated metadata
                with open(metadata_file, 'w') as f:
                    json.dump(attachments_to_keep, f, indent=2)
        except FileNotFoundError:
            pass  # Attachments folder doesn't exist, ignore
        
        return jsonify({
            'success': True,
            'message': f'Quotation {quotation_id} deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting quotation: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500








# EMAIL CHECK (uses Pos project/new-enquiry.json)
@app.route("/check-email", methods=["POST"])
def check_emails():
    data = request.get_json()
    email = data.get("email")

    try:
        with open(ENQUIRY_FILE, "r", encoding="utf-8") as f:
            enquiries = json.load(f)
    except FileNotFoundError:
        enquiries = []

    exists = any(e.get("email") == email for e in enquiries)

    return jsonify({"exists": exists})

@app.route("/quick-billing")
def quick_billing():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "quick-billing.html",
        page="quick_billing",
        title="Quick Billing - Stackly",
        user_email=user_email,
        user_name=user_name,
    )

@app.route("/api/products/qb") 
def api_products_qb():
    products = load_products()
    return jsonify({"success": True, "products": products})

@app.route("/quick-billing/deleted")
def quick_billing_deleted():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    role = session.get("role", "")
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            role = (u.get("role") or role or "").strip()
            break

    return render_template(
        "quickbilling-deleted.html",
        page="quick-billing-deleted",
        role=role,
        title="Removed Items - Stackly",
        user_email=user_email,
        user_name=user_name,
    )


@app.get("/removed-items")
def removed_items_metadata():
    """Small JSON endpoint so /quick-billing/deleted page has a named Fetch/XHR entry."""
    user_email = session.get("user")
    if not user_email:
        return jsonify(
            {"success": False, "message": "Session expired. Please login first."}
        ), 401

    users = load_users()
    user_name = "User"
    role = session.get("role", "")
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            role = (u.get("role") or role or "").strip()
            break

    return jsonify(
        {
            "success": True,
            "page": "removed-items",
            "current_user": {"email": user_email, "name": user_name, "role": role},
        }
    ), 200


# ---------- Quick Billing: load/save helpers (used by API and save-quick-bill) ----------
def load_bills():
    """Read bills from bills.json; return list (empty if missing/invalid)."""
    if not os.path.exists(BILLS_FILE):
        return []
    try:
        with open(BILLS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception as e:
        print(f"❌ Error reading {BILLS_FILE}: {e}")
        return []


def save_bills(bills):
    """Write bills list to bills.json."""
    if not isinstance(bills, list):
        raise ValueError("bills must be a list")
    with open(BILLS_FILE, "w", encoding="utf-8") as f:
        json.dump(bills, f, indent=2, ensure_ascii=False)


def generate_bill_id(bills):
    """Return next numeric bill id (max existing id + 1, or 1 if empty)."""
    if not bills:
        return 1
    ids = []
    for b in bills:
        bid = b.get("id")
        if bid is not None:
            try:
                ids.append(int(bid))
            except (TypeError, ValueError):
                pass
    return max(ids) + 1 if ids else 1


@app.route("/api/hold-bill", methods=["GET", "POST", "DELETE"])
def handle_hold_bill():
    """
    Temporary hold for a single quick-billing bill.
    Data is stored in Hold-Billing.json for now.
    """
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        try:
            with open(HOLD_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            return jsonify({"status": "success"})
        except Exception as e:  # pragma: no cover - defensive
            return jsonify({"status": "error", "message": str(e)}), 500

    if request.method == "GET":
        try:
            if os.path.exists(HOLD_FILE):
                with open(HOLD_FILE, "r", encoding="utf-8") as f:
                    bill = json.load(f)
                return jsonify({"held": True, "bill": bill})
            return jsonify({"held": False})
        except Exception as e:  # pragma: no cover - defensive
            return jsonify({"status": "error", "message": str(e)}), 500

    # DELETE
    try:
        if os.path.exists(HOLD_FILE):
            os.remove(HOLD_FILE)
        return jsonify({"status": "success"})
    except Exception as e:  # pragma: no cover - defensive
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/save-quick-bill", methods=["POST"])
def save_quick_bill():
    """
    Save a quick billing bill to bills.json.
    The frontend sends a payload like:
    {
        "items": [...],
        "totals": { "invoice_total": number },
        "payment": { "mode": "Cash" | "UPI" | "Card" | "Multiple" | "-" }
    }
    """
    try:
        data = request.get_json(silent=True) or {}
        items = data.get("items") or []
        totals = data.get("totals") or {}
        payment = data.get("payment") or {}

        if not items:
            return (
                jsonify({"success": False, "message": "No items to save"}),
                400,
            )

        bills = load_bills()
        bill_id = generate_bill_id(bills)

        bill_entry = {
            "id": bill_id,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "user": session.get("user") or "",
            "items": items,
            "totals": totals,
            "payment": payment,
        }

        bills.append(bill_entry)

        try:
            save_bills(bills)
        except Exception as e:
            print(f"❌ Error writing bills.json: {e}")
            return (
                jsonify(
                    {"success": False, "message": "Could not save bill to file"}
                ),
                500,
            )

        return jsonify({"success": True, "billId": bill_id}), 201

    except Exception as e:  # pragma: no cover - defensive
        print(f"❌ Unexpected error in save_quick_bill: {e}")
        return (
            jsonify(
                {"success": False, "message": "Server error while saving bill"}
            ),
            500,
        )


# =========================================
# Quick Billing REST API (same pattern as /api/products, /api/customer)
# =========================================

@app.route("/api/quick-billing", methods=["GET"])
def api_quick_billing_list():
    """
    GET /api/quick-billing
    List all quick bills with optional filters and pagination.
    Query params: q (search), page, page_size, user, date_from, date_to.
    Requires login.
    """
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    bills = load_bills()

    q = (request.args.get("q") or "").strip().lower()
    user_filter = (request.args.get("user") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    try:
        page = max(1, int(request.args.get("page") or 1))
        page_size = min(1000, max(1, int(request.args.get("page_size") or 10)))
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "Invalid page or page_size"}), 400

    def match(b):
        if q:
            hay = " ".join([
                str(b.get("id", "")),
                str(b.get("user", "")),
                (b.get("created_at") or ""),
            ]).lower()
            if q not in hay:
                return False
        if user_filter and (b.get("user") or "").strip().lower() != user_filter.lower():
            return False
        created = b.get("created_at") or ""
        if date_from and created < date_from:
            return False
        if date_to and created > date_to:
            return False
        return True

    filtered = [b for b in bills if match(b)]
    total_items = len(filtered)
    total_pages = max(1, (total_items + page_size - 1) // page_size)
    page = min(page, total_pages)
    start = (page - 1) * page_size
    items = filtered[start : start + page_size]

    return jsonify({
        "success": True,
        "data": {
            "items": items,
            "page": page,
            "total_pages": total_pages,
            "total_items": total_items,
        }
    }), 200


@app.route("/api/quick-billing/<int:bill_id>", methods=["GET"])
def api_quick_billing_get(bill_id):
    """
    GET /api/quick-billing/<bill_id>
    Return a single bill by id. Requires login.
    """
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    bills = load_bills()
    bill = next((b for b in bills if b.get("id") == bill_id), None)
    if not bill:
        return jsonify({"success": False, "message": "Bill not found"}), 404
    return jsonify({"success": True, "data": bill}), 200


@app.route("/api/quick-billing", methods=["POST"])
def api_quick_billing_create():
    """
    POST /api/quick-billing
    Create a new quick bill. Body: { "items": [...], "totals": {}, "payment": {} }.
    Requires login. Same behavior as /api/save-quick-bill but returns full bill and follows REST naming.
    """
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    if not request.is_json:
        return jsonify({"success": False, "message": "Content-Type must be application/json"}), 400

    data = request.get_json(silent=True) or {}
    items = data.get("items") or []
    totals = data.get("totals") or {}
    payment = data.get("payment") or {}

    if not items:
        return jsonify({"success": False, "message": "At least one item is required"}), 400

    bills = load_bills()
    bill_id = generate_bill_id(bills)
    bill_entry = {
        "id": bill_id,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "user": session.get("user") or "",
        "items": items,
        "totals": totals,
        "payment": payment,
    }
    bills.append(bill_entry)
    try:
        save_bills(bills)
    except Exception as e:
        print(f"❌ Error writing bills: {e}")
        return jsonify({"success": False, "message": "Could not save bill"}), 500

    return jsonify({
        "success": True,
        "message": "Bill created successfully",
        "data": bill_entry,
    }), 201


@app.route("/api/quick-billing/<int:bill_id>", methods=["PUT"])
def api_quick_billing_update(bill_id):
    """
    PUT /api/quick-billing/<bill_id>
    Update an existing bill. Body can include items, totals, payment (partial update supported).
    Requires login.
    """
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    if not request.is_json:
        return jsonify({"success": False, "message": "Content-Type must be application/json"}), 400

    data = request.get_json(silent=True) or {}
    bills = load_bills()
    idx = next((i for i, b in enumerate(bills) if b.get("id") == bill_id), None)
    if idx is None:
        return jsonify({"success": False, "message": "Bill not found"}), 404

    bill = bills[idx]
    if "items" in data and data["items"] is not None:
        bill["items"] = data["items"]
    if "totals" in data and data["totals"] is not None:
        bill["totals"] = data["totals"]
    if "payment" in data and data["payment"] is not None:
        bill["payment"] = data["payment"]
    bill["updated_at"] = datetime.now().isoformat(timespec="seconds")

    try:
        save_bills(bills)
    except Exception as e:
        print(f"❌ Error writing bills: {e}")
        return jsonify({"success": False, "message": "Could not update bill"}), 500

    return jsonify({"success": True, "message": "Bill updated", "data": bill}), 200


@app.route("/api/quick-billing/<int:bill_id>", methods=["DELETE"])
def api_quick_billing_delete(bill_id):
    """
    DELETE /api/quick-billing/<bill_id>
    Remove a bill. Requires login.
    """
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    bills = load_bills()
    new_list = [b for b in bills if b.get("id") != bill_id]
    if len(new_list) == len(bills):
        return jsonify({"success": False, "message": "Bill not found"}), 404

    try:
        save_bills(new_list)
    except Exception as e:
        print(f"❌ Error writing bills: {e}")
        return jsonify({"success": False, "message": "Could not delete bill"}), 500

    return jsonify({"success": True, "message": "Bill deleted successfully"}), 200


@app.route("/api/quick-billing/new-id", methods=["GET"])
def api_quick_billing_new_id():
    """
    GET /api/quick-billing/new-id
    Return the next bill id (for UI use). Requires login.
    """
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status
    bills = load_bills()
    next_id = generate_bill_id(bills)
    return jsonify({"billId": next_id}), 200

#======SALES ORDER====#

# def find_sales_order_by_id(so_id: str):
#     """
#     Find a sales order by SO ID.
#     """
#     orders = load_sales_orders()
#     so_id = (so_id or "").strip()

#     return next(
#         (order for order in orders if str(order.get("so_id", "")).strip() == so_id),
#         None
#     )

# def generate_sales_order_id():
#     """
#     Generate the next Sales Order ID in SO-0001 format.
#     Supports older formats like SO0001 or SO_0001.
#     """
#     orders = load_sales_orders()
#     last_num = 0

#     for order in orders:
#         so = str(
#             order.get("so_id") or
#             order.get("order_id") or
#             order.get("id") or
#             ""
#         ).strip()

#         cleaned = so.replace("SO-", "").replace("SO_", "").replace("SO", "")
#         match = re.search(r"(\d+)$", cleaned)

#         if match:
#             try:
#                 last_num = max(last_num, int(match.group(1)))
#             except Exception:
#                 pass

#     return f"SO-{last_num + 1:04d}"

# def upsert_sales_order(payload: dict, status_value: str):
#     """
#     Insert a new sales order or update an existing one.
#     Supports older key names like order_id for backward compatibility.
#     """
#     orders = load_sales_orders()

#     so_id = (payload.get("so_id") or payload.get("order_id") or "").strip()
#     if not so_id:
#         so_id = generate_sales_order_id()

#     existing = next(
#         (
#             order for order in orders
#             if str(order.get("so_id") or order.get("order_id") or "").strip() == so_id
#         ),
#         None
#     )

#     # Ignore placeholder customer names
#     customer_name = (payload.get("customer_name") or "").strip()
#     if customer_name.lower() in ["select customer", "—", "-", ""]:
#         customer_name = ""

#     # Autofill customer details from master
#     customer = find_customer_by_name(customer_name) if customer_name else None
#     if customer:
#         payload["customer_id"] = customer.get("customer_id", "")
#         payload["email"] = customer.get("email", "")
#         payload["phone"] = customer.get("phone", "")
#         payload["billing_address"] = customer.get("billingAddress", "")
#         payload["shipping_address"] = customer.get("shippingAddress", "")

#     now_iso = datetime.now().isoformat(timespec="seconds")

#     base = {
#         "so_id": so_id,
#         "order_date": "",
#         "sales_rep": "",
#         "order_type": "",
#         "status": status_value,
#         "stock_status": "",

#         "customer_name": "",
#         "customer_id": "",
#         "billing_address": "",
#         "shipping_address": "",
#         "email": "",
#         "phone": "",

#         "payment_method": "",
#         "currency": "",
#         "due_date": "",
#         "terms": "",

#         "shipping_method": "",
#         "delivery_date": "",
#         "tracking_number": "",
#         "internal_notes": "",
#         "customer_notes": "",

#         "items": [],

#         "subtotal": 0,
#         "tax_total": 0,
#         "rounding": 0,
#         "global_discount": 0,
#         "shipping_charges": 0,
#         "grand_total": 0,

#         "comments": [],
#         "status_history": [],

#         "cancel_reason": "",
#         "cancelled_by": "",
#         "cancelled_at": "",

#         "created_at": existing.get("created_at") if existing else now_iso,
#         "updated_at": now_iso,
#     }

#     # Merge in the correct order:
#     # defaults -> existing data -> incoming payload -> forced system fields
#     doc = {**base, **(existing or {}), **payload}
#     doc["so_id"] = so_id
#     doc["status"] = status_value
#     doc["updated_at"] = now_iso

#     if not doc.get("created_at"):
#         doc["created_at"] = now_iso

#     # Remove older key if present
#     doc.pop("order_id", None)

#     # Ensure expected list types
#     if not isinstance(doc.get("items"), list):
#         doc["items"] = []

#     if not isinstance(doc.get("comments"), list):
#         doc["comments"] = []

#     if not isinstance(doc.get("status_history"), list):
#         doc["status_history"] = []

#     # Update existing or insert new
#     if existing:
#         idx = orders.index(existing)
#         orders[idx] = doc
#     else:
#         orders.insert(0, doc)

#     save_sales_orders(orders)
#     return so_id

#======SALES ORDER====#

def generate_sales_order_id():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM sales_orders")
    count = cur.fetchone()[0] + 1

    cur.close()
    conn.close()

    return f"SO-{str(count).zfill(4)}"

@app.post("/api/sales-orders")
def create_sales_order():
    data = request.get_json()

    conn = get_db_connection()
    cur = conn.cursor()

    # INSERT HEADER
    cur.execute("""
        INSERT INTO sales_orders (
            so_id, order_date, sales_rep, order_type, status,
            customer_id, billing_address, shipping_address,
            email, phone,
            payment_method, currency, due_date, terms,
            shipping_method, delivery_date, tracking_number,internal_notes, customer_notes,
            subtotal, tax_total, global_discount, shipping_charges, grand_total
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        data["so_id"],
        data["order_date"],
        data["sales_rep"],
        data["order_type"],
        data["status"],
        data["customer_id"],
        data["billing_address"],
        data["shipping_address"],
        data["email"],
        data["phone"],
        data["payment_method"],
        data["currency"],
        data["due_date"],
        data["terms"],
        data["shipping_method"],
        data["delivery_date"],
        data["tracking_number"],
        data["internal_notes"],   
        data["customer_notes"], 
        data["subtotal"],
        data["tax_total"],
        data["global_discount"],
        data["shipping_charges"],
        data["grand_total"]
    ))

    # INSERT ITEMS
    for item in data["items"]:
        cur.execute("""
            INSERT INTO sales_order_items (
                so_id, product_id, product_name,
                qty, uom, price, tax_pct, disc_pct, line_total
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            data["so_id"],
            item["product_id"],
            item["product_name"],
            item["qty"],
            item["uom"],
            item["price"],
            item["tax_pct"],
            item["disc_pct"],
            item["line_total"]
        ))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})

@app.get("/api/sales-orders/<so_id>")
def get_sales_order(so_id):
    conn = get_db_connection()
    cur = conn.cursor()

    # HEADER
    cur.execute("SELECT * FROM sales_orders WHERE so_id=%s", (so_id,))
    row = cur.fetchone()

    if not row:
        return jsonify({"success": False}), 404

    columns = [desc[0] for desc in cur.description]
    data = dict(zip(columns, row))

    # ✅ DATE FIXES
    data["order_date"] = str(data.get("order_date") or "")
    data["due_date"] = str(data.get("due_date") or "")
    data["delivery_date"] = str(data.get("delivery_date") or "")
    data["internal_notes"] = data.get("internal_notes") or ""
    data["customer_notes"] = data.get("customer_notes") or ""

    # ✅ CUSTOMER NAME
    cur.execute("SELECT name FROM customers WHERE customer_id=%s", (data["customer_id"],))
    cust = cur.fetchone()
    data["customer_name"] = cust[0] if cust else ""
    # 🔥 ADD ITEMS
    cur.execute("SELECT * FROM sales_order_items WHERE so_id=%s", (so_id,))
    items_rows = cur.fetchall()

    items = []
    for i in items_rows:
        items.append({
            "product_id": i[2],
            "product_name": i[3],
            "qty": i[4],
            "uom": i[5],
            "price": i[6],
            "tax_pct": i[7],
            "disc_pct": i[8],
            "line_total": i[9]
        })

    data["items"] = items
    cur.execute("""
        SELECT comment, created_by, created_at
        FROM sales_order_comments
        WHERE so_id=%s
        ORDER BY created_at ASC
    """, (so_id,))

    rows = cur.fetchall()

    comments = []
    for r in rows:
        comments.append({
            "text": r[0],                     # comment
            "user": r[1],                     # created_by
            "created_at": str(r[2])
        })

    data["comments"] = comments

    cur.close()
    conn.close()

    return jsonify({
    "success": True,
    "order": data
})

@app.get("/api/sales-orders")
def list_sales_orders():
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT so_id, order_type, customer_id, sales_rep,
                   order_date, status, stock_status, grand_total
            FROM sales_orders
            ORDER BY created_at DESC
        """)

        rows = cur.fetchall()
        print("ROWS:", rows)   # 🔥 ADD THIS

        orders = []
        for r in rows:
            print("ROW:", r)   # 🔥 ADD THIS

            orders.append({
                "so_id": r[0],
                "order_type": r[1],
                "customer_name": r[2],
                "sales_rep": r[3],
                "order_date": str(r[4]),
                "status": r[5],
                "stock_status": r[6],
                "grand_total": float(r[7] or 0)
            })

        return jsonify({"orders": orders})

    except Exception as e:
        print("🔥 ERROR:", e)   # 🔥 VERY IMPORTANT
        return jsonify({"error": str(e)}), 500

    finally:
        cur.close()
        conn.close()
# =========================================
# SALES ORDER - PAGE ROUTES
# =========================================
@app.get("/sales-order")
def sales_order():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:


        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "sales-order.html",
        page="sales_order",
        title="Sales Order - Stackly",
        user_email=user_email,
        user_name=user_name,
    )


@app.get("/sales_order")
def sales_order_compat():
    return redirect("/sales-order", code=302)


@app.get("/sales-order/new")
def sales_order_new():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    # =========================
    # GENERATE SO ID
    # =========================
    so_id = generate_sales_order_id()

    # =========================
    # GET LOGGED-IN USER NAME FROM DB
    # =========================
    user_name = "User"

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT name 
        FROM users 
        WHERE LOWER(email)=LOWER(%s)
    """, (user_email,))

    row = cur.fetchone()

    if row:
        user_name = (row[0] or "").strip() or "User"

    print("SESSION EMAIL:", user_email)
    print("FINAL USER NAME:", user_name)

    # =========================
    # GET SALES REPS FROM DB
    # =========================
    cur.execute("""
        SELECT name, email, role 
        FROM users 
        WHERE role IN ('Admin','User','Sales')
    """)

    sales_reps_rows = cur.fetchall()

    sales_reps = []
    for r in sales_reps_rows:
        sales_reps.append({
            "name": r[0],
            "email": r[1],
            "role": r[2]
        })

    # =========================
    # GET CUSTOMERS FROM DB
    # =========================
    customers = get_customers_from_db()

    cur.close()
    conn.close()

    # =========================
    # RENDER PAGE
    # =========================
    response = make_response(render_template(
        "sales-new.html",
        mode="new",
        so_id=so_id,
        sales_reps=sales_reps,
        customers=customers,
        page="sales_order",
        user_email=user_email,
        user_name=user_name,
    ))

    # =========================
    # PREVENT CACHE
    # =========================
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"

    return response

    # =========================
    # GENERATE SO ID
    # =========================
    so_id = generate_sales_order_id()

    # =========================
    # GET LOGGED-IN USER NAME FROM DB
    # =========================
    user_name = "User"

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT name 
        FROM users 
        WHERE LOWER(email)=LOWER(%s)
    """, (user_email,))

    row = cur.fetchone()

    if row:
        user_name = (row[0] or "").strip() or "User"

    print("SESSION EMAIL:", user_email)
    print("FINAL USER NAME:", user_name)

    # =========================
    # GET SALES REPS FROM DB
    # =========================
    cur.execute("""
        SELECT name, email, role 
        FROM users 
        WHERE role IN ('Admin','User','Sales')
    """)

    sales_reps_rows = cur.fetchall()

    sales_reps = []
    for r in sales_reps_rows:
        sales_reps.append({
            "name": r[0],
            "email": r[1],
            "role": r[2]
        })

    # =========================
    # GET CUSTOMERS FROM DB
    # =========================
    customers = get_customers_from_db()

    cur.close()
    conn.close()

    # =========================
    # RENDER PAGE
    # =========================
    response = make_response(render_template(
        "sales-new.html",
        mode="new",
        so_id=so_id,
        sales_reps=sales_reps,
        customers=customers,
        page="sales_order",
        user_email=user_email,
        user_name=user_name,
    ))

    # =========================
    # PREVENT CACHE
    # =========================
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"

    return response

@app.get("/sales-order/edit/<so_id>")
def sales_order_edit(so_id):
    users = load_users()
    sales_reps = [u for u in users if u.get("role") in ["Admin", "User", "Sales"]]
#==================================from here
    # customers = load_customer()
    customers = get_customers_from_db()
    #=======================================upto replace

    return render_template(
        "sales-new.html",
        mode="edit",
        so_id=so_id,
        sales_reps=sales_reps,
        customers=customers,
        page="sales_order"
    )


# =========================================
# SALES ORDER - API ROUTES
# =========================================
@app.get("/api/sales-orders/next-id")
def api_sales_orders_next_id():
    return jsonify({
        "success": True,
        "so_id": generate_sales_order_id()
    })


# @app.get("/api/sales-orders")
# def api_sales_orders_list():
#     """
#     Return all sales orders from sales_orders.json.
#     Used by Delivery Note 'Sales Order Reference' dropdown.
#     """
#     orders = load_sales_orders()
#     return jsonify(orders)


# @app.get("/api/sales-orders/all")
# def api_sales_orders_all():
#     orders = load_sales_orders()
#     return jsonify({"orders": orders})


# @app.get("/api/sales-orders/<so_id>")
# def get_one_sales_order(so_id):
#     so = find_sales_order_by_id(so_id)

#     if not so:
#         return jsonify({
#             "success": False,
#             "message": "Not found"
#         }), 404

#     return jsonify({
#         "success": True,
#         "order": so
#     })


@app.post("/api/sales-orders/save-draft")
def api_sales_orders_save_draft():
    data = request.get_json()

    conn = get_db_connection()
    cur = conn.cursor()

    # INSERT HEADER
    cur.execute("""
        INSERT INTO sales_orders (
            so_id, order_date, sales_rep, order_type, status,
            customer_id, billing_address, shipping_address,
            email, phone,
            payment_method, currency, due_date, terms,
            shipping_method, delivery_date, tracking_number,internal_notes, customer_notes,
            subtotal, tax_total, global_discount, shipping_charges, grand_total
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        data["so_id"],
        data["order_date"],
        data["sales_rep"],
        data["order_type"],
        "Draft",   # 🔥 IMPORTANT
        data["customer_id"],
        data["billing_address"],
        data["shipping_address"],
        data["email"],
        data["phone"],
        data["payment_method"],
        data["currency"],
        data["due_date"],
        data["terms"],
        data["shipping_method"],
        data["delivery_date"],
        data["tracking_number"],
        data["internal_notes"],   
        data["customer_notes"],
        data["subtotal"],
        data["tax_total"],
        data["global_discount"],
        data["shipping_charges"],
        data["grand_total"]
    ))

    # INSERT ITEMS
    for item in data["items"]:
        cur.execute("""
            INSERT INTO sales_order_items (
                so_id, product_id, product_name,
                qty, uom, price, tax_pct, disc_pct, line_total
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            data["so_id"],
            item["product_id"],
            item["product_name"],
            item["qty"],
            item["uom"],
            item["price"],
            item["tax_pct"],
            item["disc_pct"],
            item["line_total"]
        ))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})

@app.post("/api/sales-orders/submit")
def api_sales_orders_submit():
    data = request.get_json()

    conn = get_db_connection()
    cur = conn.cursor()

    # INSERT HEADER
    cur.execute("""
        INSERT INTO sales_orders (
            so_id, order_date, sales_rep, order_type, status,
            customer_id, billing_address, shipping_address,
            email, phone,
            payment_method, currency, due_date, terms,
            shipping_method, delivery_date, tracking_number,internal_notes, customer_notes,
            subtotal, tax_total, global_discount, shipping_charges, grand_total
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        data["so_id"],
        data["order_date"],
        data["sales_rep"],
        data["order_type"],
        "Submitted",
        data["customer_id"],
        data["billing_address"],
        data["shipping_address"],
        data["email"],
        data["phone"],
        data["payment_method"],
        data["currency"],
        data["due_date"],
        data["terms"],
        data["shipping_method"],
        data["delivery_date"],
        data["tracking_number"],
        data["internal_notes"],   
        data["customer_notes"],
        data["subtotal"],
        data["tax_total"],
        data["global_discount"],
        data["shipping_charges"],
        data["grand_total"]
    ))

    # INSERT ITEMS
    for item in data["items"]:
        cur.execute("""
            INSERT INTO sales_order_items (
                so_id, product_id, product_name,
                qty, uom, price, tax_pct, disc_pct, line_total
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            data["so_id"],
            item["product_id"],
            item["product_name"],
            item["qty"],
            item["uom"],
            item["price"],
            item["tax_pct"],
            item["disc_pct"],
            item["line_total"]
        ))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})


@app.get("/api/sales-products")
def api_sales_products():
    products = load_products()
    return jsonify({
        "success": True,
        "products": products
    })

@app.post("/api/sales-orders/<so_id>/comments")
def add_sales_order_comment(so_id):
    data = request.get_json()
    comment = (data.get("comment") or "").strip()
    user = (data.get("user") or "User").strip()

    if not comment:
        return jsonify({"success": False, "message": "Empty comment"}), 400

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO sales_order_comments (so_id, comment, created_by)
        VALUES (%s, %s, %s)
    """, (so_id, comment, user))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})

@app.get("/api/sales-orders/<so_id>/comments")
def get_sales_order_comments(so_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT comment, created_by, created_at
        FROM sales_order_comments
        WHERE so_id=%s
        ORDER BY created_at DESC
    """, (so_id,))

    rows = cur.fetchall()

    comments = []
    for r in rows:
        comments.append({
            "comment": r[0],
            "user": r[1],
            "time": r[2].strftime("%d/%m/%Y, %I:%M %p")
        })

    cur.close()
    conn.close()

    return jsonify({"success": True, "comments": comments})

# =========================================
# SALES ORDER - PDF
# =========================================
@app.get("/api/sales-orders/<so_id>/pdf")
def sales_order_pdf(so_id):
    conn = get_db_connection()
    cur = conn.cursor()

    # ========================
    # FETCH HEADER
    # ========================
    cur.execute("SELECT * FROM sales_orders WHERE so_id=%s", (so_id,))
    order = cur.fetchone()

    if not order:
        cur.close()
        conn.close()
        return jsonify({
            "success": False,
            "message": "Sales Order not found"
        }), 404

    # ========================
    # FETCH ITEMS
    # ========================
    cur.execute("SELECT * FROM sales_order_items WHERE so_id=%s", (so_id,))
    items = cur.fetchall()

    cur.close()
    conn.close()

    # ========================
    # CONVERT DB → DICT
    # ========================
    columns = [
        "so_id","order_date","sales_rep","order_type","status",
        "customer_id","billing_address","shipping_address",
        "email","phone",
        "payment_method","currency","due_date","terms",
        "shipping_method","delivery_date","tracking_number",
        "subtotal","tax_total","global_discount","shipping_charges","grand_total",
        "created_at"
    ]

    so = dict(zip(columns, order))

    # ========================
    # ATTACH ITEMS
    # ========================
    so["items"] = []
    for i in items:
        so["items"].append({
            "product_id": i[2],
            "product_name": i[3],
            "qty": float(i[4] or 0),
            "uom": i[5],
            "price": float(i[6] or 0),
            "tax_pct": float(i[7] or 0),
            "disc_pct": float(i[8] or 0),
            "line_total": float(i[9] or 0)
        })

    # ========================
    # GENERATE PDF
    # ========================
    try:
        pdf_bytes = generate_sales_order_pdf_bytes(so)

        response = make_response(pdf_bytes)
        response.headers["Content-Type"] = "application/pdf"
        response.headers["Content-Disposition"] = f'inline; filename="{so_id}.pdf"'

        return response

    except Exception as e:
        print("Sales Order PDF error:", e)
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


def generate_sales_order_pdf_bytes(so):
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=32
    )

    elements = []
    styles = getSampleStyleSheet()

    # =========================================
    # PDF STYLES
    # =========================================
    title_style = ParagraphStyle(
        "SOTitle",
        parent=styles["Heading1"],
        fontSize=24,
        leading=28,
        alignment=1,
        textColor=colors.HexColor("#8c1f1f"),
        spaceAfter=8,
        fontName="DejaVuSans-Bold"
    )

    company_style = ParagraphStyle(
        "SOCompany",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        textColor=colors.black,
        alignment=1,
        spaceAfter=2,
        fontName="DejaVuSans"
    )

    status_style = ParagraphStyle(
        "SOStatus",
        parent=styles["Heading2"],
        fontSize=14,
        leading=18,
        alignment=1,
        textColor=colors.HexColor("#148a08"),
        spaceAfter=14,
        fontName="DejaVuSans-Bold"
    )

    heading_style = ParagraphStyle(
        "SOHeading",
        parent=styles["Heading2"],
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#8c1f1f"),
        spaceBefore=8,
        spaceAfter=8,
        fontName="DejaVuSans-Bold"
    )

    table_cell_style = ParagraphStyle(
        "SOTableCell",
        parent=styles["Normal"],
        fontSize=7.6,
        leading=9,
        fontName="DejaVuSans",
        wordWrap="CJK"
    )

    table_label_style = ParagraphStyle(
        "SOTableLabel",
        parent=styles["Normal"],
        fontSize=7.6,
        leading=9,
        fontName="DejaVuSans-Bold",
        textColor=colors.HexColor("#5f2d2d"),
        wordWrap="CJK"
    )

    terms_heading_style = ParagraphStyle(
        "SOTermsHeading",
        parent=styles["Heading2"],
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#8c1f1f"),
        spaceAfter=6,
        fontName="DejaVuSans-Bold"
    )

    terms_style = ParagraphStyle(
        "SOTerms",
        parent=styles["Normal"],
        fontSize=7.4,
        leading=10,
        fontName="DejaVuSans"
    )

    footer_style = ParagraphStyle(
        "SOFooter",
        parent=styles["Normal"],
        fontSize=7.5,
        textColor=colors.HexColor("#555555"),
        alignment=0
    )

    # =========================================
    # CURRENCY MAPPING
    # =========================================
    currency_code = so.get("currency", "INR")
    currency_map = {
        "USD": "$",
        "EUR": "€",
        "GBP": "£",
        "INR": "₹",
        "IND": "₹",
        "SGD": "S$"
    }
    currency_symbol = currency_map.get(currency_code, currency_code)

    # =========================================
    # PDF HEADER
    # =========================================
    elements.append(Paragraph("STACKLY", title_style))
    elements.append(Paragraph(
        "MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008",
        company_style
    ))
    elements.append(Paragraph("Phone: +91 7010792745", company_style))
    elements.append(Paragraph("Email: info@stackly.com", company_style))
    elements.append(Spacer(1, 10))

    status_text = (so.get("status") or "Submitted").upper()
    elements.append(Paragraph(f"SALES ORDER - {status_text}", status_style))
    elements.append(Spacer(1, 4))

    # =========================================
    # SALES ORDER INFO TABLE
    # =========================================
    info_data = [
        [
            Paragraph("Sales Order Number:", table_label_style),
            Paragraph(str(so.get("so_id", "") or "-"), table_cell_style),
            Paragraph("Date:", table_label_style),
            Paragraph(str(so.get("order_date", "") or "-"), table_cell_style),
        ],
        [
            Paragraph("Customer:", table_label_style),
            Paragraph(str(so.get("customer_name", "") or "-"), table_cell_style),
            Paragraph("Delivery Date:", table_label_style),
            Paragraph(str(so.get("delivery_date", "") or "-"), table_cell_style),
        ],
        [
            Paragraph("Sales Rep:", table_label_style),
            Paragraph(str(so.get("sales_rep", "") or "-"), table_cell_style),
            Paragraph("Currency:", table_label_style),
            Paragraph(str(currency_code or "-"), table_cell_style),
        ],
        [
            Paragraph("Order Type:", table_label_style),
            Paragraph(str(so.get("order_type", "") or "-"), table_cell_style),
            Paragraph("Payment Terms:", table_label_style),
            Paragraph(str(so.get("terms", "N/A") or "N/A"), table_cell_style),
        ],
    ]

    info_table = Table(info_data, colWidths=[110, 145, 95, 130])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#efefef")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#efefef")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 14))

    # =========================================
    # ITEMS TABLE
    # =========================================
    elements.append(Paragraph("SALES ORDER ITEMS", heading_style))

    items = so.get("items", [])
    product_map = {
        str(p.get("product_id", "")).strip(): p
        for p in load_products()
    }

    table_data = [[
        "S.No", "Product Name", "Qty", "UOM", "Unit Price", "Tax %", "Disc %", "Total"
    ]]

    subtotal_calc = 0.0
    total_tax_calc = 0.0
    total_discount_calc = 0.0

    for idx, item in enumerate(items, start=1):
        pid = str(item.get("product_id", "")).strip()
        product = product_map.get(pid, {})

        qty = float(item.get("qty", 0) or 0)
        uom = item.get("uom", "") or "Nos"

        unit_price = float(
            item.get("price")
            or item.get("unit_price")
            or product.get("unit_price")
            or product.get("price")
            or product.get("selling_price")
            or 0
        )

        tax_pct = float(item.get("tax_pct", 0) or 0)
        disc_pct = float(item.get("disc_pct", 0) or 0)

        line_subtotal = qty * unit_price
        discount_amt = line_subtotal * (disc_pct / 100)
        after_discount = line_subtotal - discount_amt
        tax_amt = after_discount * (tax_pct / 100)
        line_total = float(item.get("line_total", 0) or (after_discount + tax_amt))

        subtotal_calc += line_subtotal
        total_tax_calc += tax_amt
        total_discount_calc += discount_amt

        table_data.append([
            str(idx),
            str(item.get("product_name", "") or "-"),
            f"{qty:.2f}",
            str(uom),
            f"{currency_symbol}{unit_price:.2f}",
            f"{tax_pct:.1f}%",
            f"{disc_pct:.1f}%",
            f"{currency_symbol}{line_total:.2f}"
        ])

    items_table = Table(table_data, colWidths=[32, 170, 42, 40, 60, 44, 44, 58])
    items_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#a12828")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 14))

    # =========================================
    # TOTALS SUMMARY
    # =========================================
    elements.append(Paragraph("TAX AND TOTALS SUMMARY", heading_style))

    shipping_charge = float(so.get("shipping_charges", 0) or 0)
    global_discount = float(so.get("global_discount", 0) or 0)
    rounding = float(so.get("rounding", 0) or 0)

    grand_total = float(
        so.get("grand_total", 0)
        or (subtotal_calc - global_discount + total_tax_calc + shipping_charge + rounding)
    )

    summary_data = [
        ["Subtotal:", f"{currency_symbol}{subtotal_calc:.2f}"],
        ["Total Discount (Item Level):", f"{currency_symbol}{total_discount_calc:.2f}"],
        ["Total Tax:", f"{currency_symbol}{total_tax_calc:.2f}"],
        ["Shipping Charge:", f"{currency_symbol}{shipping_charge:.2f}"],
        ["Global Discount:", f"-{currency_symbol}{global_discount:.2f}"],
    ]

    if rounding != 0:
        sign = "+" if rounding > 0 else "-"
        summary_data.append(["Rounding Adjustment:", f"{sign}{currency_symbol}{abs(rounding):.2f}"])

    summary_data.append(["GRAND TOTAL:", f"{currency_symbol}{grand_total:.2f}"])

    summary_table = Table(summary_data, colWidths=[300, 200])
    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -2), "DejaVuSans"),
        ("FONTNAME", (0, -1), (-1, -1), "DejaVuSans-Bold"),
        ("FONTSIZE", (0, 0), (-1, -2), 8),
        ("FONTSIZE", (0, -1), (-1, -1), 9),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 0), (1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#a12828")),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.whitesmoke),
        ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.HexColor("#7f1f1f")),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    # =========================================
    # TERMS AND CONDITIONS
    # =========================================
    elements.append(Paragraph("Terms and Conditions", terms_heading_style))

    terms_lines = [
        "1. This Sales Order is issued based on the confirmed order details.",
        "2. Delivery will be made as per the agreed schedule.",
        "3. Payment should be completed as per agreed terms.",
        "4. Shipping charges extra if applicable.",
        "5. Goods once sold will not be taken back.",
        "6. All taxes and duties as applicable.",
        f"7. Internal Notes: {so.get('internal_notes', '') or 'N/A'}",
        f"8. Customer Notes: {so.get('customer_notes', '') or 'N/A'}",
    ]

    for line in terms_lines:
        elements.append(Paragraph(line, terms_style))

    elements.append(Spacer(1, 18))

    # =========================================
    # FOOTER
    # =========================================
    generated_on = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    elements.append(Paragraph(f"Generated on: {generated_on}", footer_style))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


# =========================================
# SALES ORDER - EMAIL API
# =========================================
@app.post("/api/sales-orders/<so_id>/email")
def sales_order_email(so_id):
    conn = get_db_connection()
    cur = conn.cursor()

    # HEADER
    cur.execute("SELECT * FROM sales_orders WHERE so_id=%s", (so_id,))
    order = cur.fetchone()

    if not order:
        cur.close()
        conn.close()
        return jsonify({
            "success": False,
            "message": "Sales Order not found"
        }), 404

    # ITEMS
    cur.execute("SELECT * FROM sales_order_items WHERE so_id=%s", (so_id,))
    items = cur.fetchall()

    cur.close()
    conn.close()

    # ========================
    # CONVERT DB → DICT
    # ========================
    columns = [
        "so_id","order_date","sales_rep","order_type","status",
        "customer_id","billing_address","shipping_address",
        "email","phone",
        "payment_method","currency","due_date","terms",
        "shipping_method","delivery_date","tracking_number",
        "subtotal","tax_total","global_discount","shipping_charges","grand_total",
        "created_at"
    ]

    so = dict(zip(columns, order))

    so["items"] = []
    for i in items:
        so["items"].append({
            "product_id": i[2],
            "product_name": i[3],
            "qty": float(i[4] or 0),
            "uom": i[5],
            "price": float(i[6] or 0),
            "tax_pct": float(i[7] or 0),
            "disc_pct": float(i[8] or 0),
            "line_total": float(i[9] or 0)
        })

    # ========================
    # EMAIL LOGIC
    # ========================
    customer_email = (so.get("email") or "").strip()
    if not customer_email:
        return jsonify({
            "success": False,
            "message": "Customer email not found"
        }), 400

    try:
        pdf_bytes = generate_sales_order_pdf_bytes(so)

        customer_name = so.get("customer_name", "Customer")
        so_no = so.get("so_id", "")
        order_date = so.get("order_date", "")
        grand_total = so.get("grand_total", 0)
        currency = so.get("currency", "INR")

        subject = f"Sales Order {so_no} from Stackly"

        body = f"""
Dear {customer_name},

Greetings from Stackly.

Please find attached the Sales Order document.

Sales Order No : {so_no}
Order Date     : {order_date}
Grand Total    : {currency} {grand_total}

Thanks & Regards,
Stackly Team
""".strip()

        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = SENDER_EMAIL
        msg["To"] = customer_email
        msg.set_content(body)

        msg.add_attachment(
            pdf_bytes,
            maintype="application",
            subtype="pdf",
            filename=f"{so_no}.pdf"
        )

        context = ssl.create_default_context()
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls(context=context)
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.send_message(msg)

        return jsonify({"success": True})

    except Exception as e:
        print("Email error:", e)
        return jsonify({"success": False, "message": str(e)}), 500

# =========================================
# SALES ORDER - CANCEL API
# =========================================
@app.post("/api/sales-orders/<so_id>/cancel")
def cancel_sales_order(so_id):
    data = request.get_json()
    reason = (data.get("reason") or "").strip()

    if not reason:
        return jsonify({
            "success": False,
            "message": "Cancellation reason is required"
        }), 400

    conn = get_db_connection()
    cur = conn.cursor()

    # check exists
    cur.execute("SELECT status FROM sales_orders WHERE so_id=%s", (so_id,))
    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return jsonify({
            "success": False,
            "message": "Sales Order not found"
        }), 404

    # update status
    cur.execute("""
        UPDATE sales_orders
        SET status='Cancelled'
        WHERE so_id=%s
    """, (so_id,))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        "success": True,
        "message": "Order cancelled"
    })

# =========================================
# DELIVERY NOTE - UTILITIES / HELPERS
# =========================================

# -----------------------------------------
# Find customer by name (customer.json)
# -----------------------------------------
def find_customer_by_name(name: str):
    """
    Find a customer by name from customer.json.
    Returns the full customer object if found, else None.
    """
    if not name:
        return None

    target_name = str(name).strip().lower()
    #==============================from here
    # customers = load_customer()
   
    customers = get_customers_from_db()
     #===============================upto replace
    for customer in customers:
        customer_name = str(customer.get("name", "")).strip().lower()
        if customer_name == target_name:
            return customer
    return None


# -----------------------------------------
# Delivery Notes JSON Storage
# -----------------------------------------
def load_delivery_notes():
    if not os.path.exists(DELIVERY_NOTE_FILE):
        with open(DELIVERY_NOTE_FILE, "w", encoding="utf-8") as f:
            json.dump([], f)
        return []

    with open(DELIVERY_NOTE_FILE, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return []


def save_delivery_notes(data):
    with open(DELIVERY_NOTE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


# -----------------------------------------
# Next DN ID generator (DN-0001 format)
# -----------------------------------------
def next_dn_id(notes):
    max_num = 0
    for n in notes:
        dn = str(n.get("dn_id", ""))
        if dn.startswith("DN-"):
            try:
                num = int(dn.split("-")[1])
                max_num = max(max_num, num)
            except:
                pass
    return f"DN-{max_num+1:04d}"


# -----------------------------------------
# Get DN by ID
# -----------------------------------------
def get_dn_by_id(dn_id: str):
    dns = load_delivery_notes()
    return next((x for x in dns if x.get("dn_id") == dn_id), None)


# -----------------------------------------
# EMAIL ATTACHMENT HELPER
# -----------------------------------------
def send_email_with_attachments(to_email, subject, body, from_email, password, attachments=None):
    smtp_server = "smtp.gmail.com"
    port = 587
    attachments = attachments or []

    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = to_email

    msg.attach(MIMEText(body, "plain", "utf-8"))

    for a in attachments:
        filename = a.get("filename", "attachment")
        content = a.get("content_bytes", b"")
        part = MIMEApplication(content, _subtype="pdf")
        part.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(part)

    context = ssl.create_default_context()

    try:
        with smtplib.SMTP(smtp_server, port) as server:
            server.starttls(context=context)
            server.login(from_email, password)
            server.sendmail(from_email, [to_email], msg.as_string())
        return True
    except Exception as e:
        print("❌ Email send error:", e)
        return False


# =========================================
# DELIVERY NOTE - FIRST PAGE (List Page)
# delivery-note.html + delivery-note.js
# =========================================

@app.route("/delivery_note")
def delivery_note():
    user_email = session.get("user")
    user_name = "User"

    if user_email:
        users = load_users()
        for u in users:
            if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
                user_name = u.get("name") or "User"
                break

    return render_template(
        "delivery-note.html",
        page="delivery_note",
        user_email=user_email,
        user_name=user_name,
    )


# =========================================
# DELIVERY NOTE - SECOND PAGE (New/Edit/View)
# deliverynote-new.html + deliverynote-new.js
# =========================================

@app.route("/delivery_note/new")
def delivery_note_new():
    # Preload data needed for the New Delivery Note page
    sales_orders = load_sales_orders()
    notes = load_delivery_notes()
    next_id = next_dn_id(notes)

    return render_template(
        "deliverynote-new.html",
        page="delivery_note",
        sales_orders=sales_orders,
        next_dn_id=next_id,
        so_id=request.args.get("so_id", "").strip(),
        user_email=session.get("user"),
        user_name=_get_logged_in_user_name(),
    )


@app.route("/delivery_note/form")
def delivery_note_form():
    dn_id = request.args.get("id", "")
    mode  = request.args.get("mode", "edit")
    sales_orders = load_sales_orders()
    return render_template(
        "deliverynote-new.html",
        page="delivery_note",
        dn_id=dn_id,
        mode=mode,
        sales_orders=sales_orders,
        user_email=session.get("user"),
        user_name=_get_logged_in_user_name(),
    )


# =========================================
# DELIVERY NOTE - API (List + Create)
# =========================================

@app.route("/api/delivery-notes", methods=["GET", "POST"])
def api_delivery_notes():

    # GET: list all delivery notes (for first page table)
    if request.method == "GET":
        notes = load_delivery_notes()
        return jsonify({"success": True, "data": notes})

    # POST: create new delivery note (from second page submit/save draft)
    data = request.get_json(force=True) or {}
    notes = load_delivery_notes()

    new_id = (data.get("dn_id") or "").strip()
    if not new_id:
        new_id = next_dn_id(notes)

    record = {
        "dn_id": new_id,
        "delivery_date": data.get("delivery_date", ""),
        "so_ref": data.get("so_ref", ""),
        "customer_name": data.get("customer_name", ""),
        "delivery_type": data.get("delivery_type", ""),
        "destination_address": data.get("destination_address", ""),
        "delivery_by": data.get("delivery_by", ""),
        "delivery_status": data.get("delivery_status", "draft"),
        "vehicle_no": data.get("vehicle_no", ""),
        "tracking_id": data.get("tracking_id", ""),
        "delivery_notes": data.get("delivery_notes", ""),
        "status": data.get("status", "Draft"),
        "items": data.get("items", []),
    }

    # Auto-fetch customer fields for email/phone/address (based on customer_name)
    customer_name = (record.get("customer_name") or "").strip()
    customer = find_customer_by_name(customer_name)

    if customer:
        record["customer_id"] = customer.get("customer_id", "")
        record["email"] = customer.get("email", "")
        record["phone"] = customer.get("phone", "")
        record["billing_address"] = customer.get("billingAddress", "")
        record["shipping_address"] = customer.get("shippingAddress", "")
    else:
        record["customer_id"] = ""
        record["email"] = ""
        record["phone"] = ""
        record["billing_address"] = ""
        record["shipping_address"] = ""

    notes.append(record)
    save_delivery_notes(notes)

    return jsonify({"success": True, "message": "Delivery Note Saved", "dn_id": new_id})


# =========================================
# DELIVERY NOTE - API (Get One + Update One)
# =========================================

@app.route("/api/delivery-notes/<dn_id>", methods=["GET", "PUT"])
def api_delivery_note_one(dn_id):
    notes = load_delivery_notes()
    dn = next((x for x in notes if x.get("dn_id") == dn_id), None)

    # GET: fetch one DN (for edit/view)
    if request.method == "GET":
        if not dn:
            return jsonify({"success": False, "message": "Delivery Note not found"}), 404

        # Ensure missing keys exist (old records safety)
        dn.setdefault("delivery_type", "")
        dn.setdefault("destination_address", "")
        dn.setdefault("vehicle_no", "")
        dn.setdefault("tracking_id", "")
        dn.setdefault("delivery_by", "")
        dn.setdefault("delivery_notes", "")
        dn.setdefault("delivery_status", "draft")

        return jsonify({"success": True, "data": dn})

    # PUT: update DN
    if not dn:
        return jsonify({"success": False, "message": "Delivery Note not found"}), 404

    payload = request.get_json(force=True) or {}

    dn["delivery_date"] = payload.get("delivery_date", "")
    dn["so_ref"] = payload.get("so_ref", "")
    dn["customer_name"] = payload.get("customer_name", "")
    dn["delivery_type"] = payload.get("delivery_type", "")
    dn["destination_address"] = payload.get("destination_address", "")
    dn["delivery_by"] = payload.get("delivery_by", "")
    dn["delivery_status"] = payload.get("delivery_status", "")
    dn["vehicle_no"] = payload.get("vehicle_no", "")
    dn["tracking_id"] = payload.get("tracking_id", "")
    dn["delivery_notes"] = payload.get("delivery_notes", "")
    dn["status"] = payload.get("status", dn.get("status", ""))
    dn["items"] = payload.get("items", [])

    save_delivery_notes(notes)
    return jsonify({"success": True, "message": "Delivery Note Updated"})


# =========================================
# DELIVERY NOTE - API (Cancel DN)
# =========================================

@app.put("/api/delivery-notes/<dn_id>/cancel")
def cancel_delivery_note(dn_id):
    notes = load_delivery_notes()

    idx = next((i for i, x in enumerate(notes) if x.get("dn_id") == dn_id), None)
    if idx is None:
        return jsonify(success=False, message="Delivery Note not found"), 404

    dn = notes[idx]

    current_status = (dn.get("delivery_status") or dn.get("status") or "").strip().lower().replace(" ", "_")
    if current_status == "cancelled":
        return jsonify(success=True, message="Already cancelled")

    dn["delivery_status"] = "cancelled"
    dn["status"] = "Cancelled"
    dn["cancel_reason"] = (request.json or {}).get("reason", "").strip()
    dn["cancelled_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    notes[idx] = dn
    save_delivery_notes(notes)

    return jsonify(success=True, message="Delivery Note cancelled successfully")


# =========================================
# DELIVERY NOTE - API (PDF) - REPORTLAB
# =========================================
@app.get("/api/delivery-notes/<dn_id>/pdf")
def delivery_note_pdf(dn_id):
    dn = get_dn_by_id(dn_id)

    if not dn:
        return jsonify({"success": False, "message": "Delivery Note not found"}), 404

    pdf_bytes = generate_delivery_note_pdf_bytes(dn)

    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f'inline; filename="{dn_id}.pdf"'
    return response


# =========================================
# DELIVERY NOTE - API (Email with PDF) - REPORTLAB
# =========================================
@app.post("/api/delivery-notes/<dn_id>/email")
def email_delivery_note(dn_id):
    dns = load_delivery_notes()
    dn = next((x for x in dns if x.get("dn_id") == dn_id), None)

    if not dn:
        return jsonify({"success": False, "message": "DN not found"}), 404

    # 1) email from DN
    customer_email = (dn.get("email") or "").strip()

    # 2) fallback customer master
    if not customer_email:
        customer = find_customer_by_name(dn.get("customer_name", ""))
        if customer:
            customer_email = (customer.get("email") or "").strip()

    # 3) fallback sales order
    if not customer_email:
        so_ref = dn.get("so_ref")
        if so_ref:
            orders = load_sales_orders()
            so = next((o for o in orders if o.get("so_id") == so_ref), None)
            if so:
                customer_email = (so.get("email") or "").strip()

    if not customer_email:
        return jsonify({"success": False, "message": "Customer email not available"}), 400

    pdf_bytes = generate_delivery_note_pdf_bytes(dn)

    ok = send_email_with_attachments(
        to_email=customer_email,
        subject=f"Delivery Note {dn_id}",
        body=f"Dear {dn.get('customer_name','Customer')},\n\nPlease find attached Delivery Note {dn_id}.\n\nRegards,\nStackly POS",
        from_email=EMAIL_ADDRESS,
        password=EMAIL_PASSWORD,
        attachments=[
            {"filename": f"{dn_id}.pdf", "content_bytes": pdf_bytes}
        ],
    )

    if not ok:
        return jsonify({"success": False, "message": "Email failed. Check SMTP/App password/Spam."}), 500

    return jsonify({"success": True, "message": "Email sent"})


# =========================================
# DELIVERY NOTE - PRINT PAGE (Optional route)
# Uses same PDF generator and streams inline
# =========================================
@app.get("/delivery-note/<dn_id>/print")
def delivery_note_print(dn_id):
    dn = get_dn_by_id(dn_id)
    if not dn:
        return "DN not found", 404

    pdf_bytes = generate_delivery_note_pdf_bytes(dn)
    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f'inline; filename="{dn_id}.pdf"'
    return response


# =========================================
# DELIVERY NOTE - PDF GENERATOR (REPORTLAB)
# =========================================
def generate_delivery_note_pdf_bytes(dn):
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18,
        leftMargin=18,
        topMargin=16,
        bottomMargin=18
    )

    styles = getSampleStyleSheet()

    # ---------------------------------------------------
    # STYLES
    # ---------------------------------------------------
    company_style = ParagraphStyle(
        name="CompanyName",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#8c1f1f"),
        alignment=TA_CENTER,
        spaceAfter=4,
    )

    company_info_style = ParagraphStyle(
        name="CompanyInfo",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=9,
        leading=12,
        textColor=colors.black,
        alignment=TA_CENTER,
        spaceAfter=1,
    )

    page_title_style = ParagraphStyle(
        name="PageTitle",
        parent=styles["Heading1"],
        fontName="DejaVuSans-Bold",
        fontSize=16,
        leading=20,
        textColor=colors.green,
        alignment=TA_CENTER,
        spaceBefore=12,
        spaceAfter=12,
    )

    section_style = ParagraphStyle(
        name="DNSection",
        parent=styles["Heading3"],
        fontName="DejaVuSans-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#8c1f1f"),
        spaceAfter=6,
        spaceBefore=10,
    )

    label_style = ParagraphStyle(
        name="DNLabel",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#6b1a1a"),
    )

    value_style = ParagraphStyle(
        name="DNValue",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=8.5,
        leading=11,
        textColor=colors.black,
    )

    summary_white_style = ParagraphStyle(
        name="DNSummaryWhite",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=8.5,
        leading=11,
        textColor=colors.white,
    )

    small_style = ParagraphStyle(
        name="DNSmall",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#444444"),
    )

    header_small_style = ParagraphStyle(
        name="DNHeaderSmall",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=TA_CENTER,
    )

    terms_style = ParagraphStyle(
        name="TermsStyle",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=8,
        leading=11,
        textColor=colors.black,
        leftIndent=8,
    )

    elements = []

    # ---------------------------------------------------
    # SAFE HELPERS
    # ---------------------------------------------------
    def safe_str(val, default="-"):
        if val is None:
            return default
        s = str(val).strip()
        return s if s else default

    def safe_float(val, default=0.0):
        try:
            if val in (None, ""):
                return default
            return float(val)
        except Exception:
            return default

    # ---------------------------------------------------
    # COMPANY HEADER
    # ---------------------------------------------------
    elements.append(Paragraph("STACKLY", company_style))
    elements.append(Paragraph(
        "MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008",
        company_info_style
    ))
    elements.append(Paragraph("Phone: +91 7010792745", company_info_style))
    elements.append(Paragraph("Email: info@stackly.com", company_info_style))
    elements.append(Spacer(1, 10))

    # ---------------------------------------------------
    # PAGE TITLE
    # ---------------------------------------------------
    status_text = safe_str(dn.get("delivery_status") or dn.get("status") or "SUBMITTED").upper()
    elements.append(Paragraph(f"DELIVERY NOTE - {status_text}", page_title_style))
    elements.append(Spacer(1, 2))

    # ---------------------------------------------------
    # TOP DETAILS TABLE
    # ---------------------------------------------------
    dn_details_data = [
        [
            Paragraph("<b>Delivery Note Number:</b>", label_style),
            Paragraph(safe_str(dn.get("dn_id")), value_style),
            Paragraph("<b>Date:</b>", label_style),
            Paragraph(safe_str(dn.get("delivery_date")), value_style),
        ],
        [
            Paragraph("<b>Customer:</b>", label_style),
            Paragraph(safe_str(dn.get("customer_name")), value_style),
            Paragraph("<b>Sales Order Ref:</b>", label_style),
            Paragraph(safe_str(dn.get("so_ref")), value_style),
        ],
        [
            Paragraph("<b>Delivery Type:</b>", label_style),
            Paragraph(safe_str(dn.get("delivery_type")), value_style),
            Paragraph("<b>Delivery By:</b>", label_style),
            Paragraph(safe_str(dn.get("delivery_by")), value_style),
        ],
        [
            Paragraph("<b>Vehicle Number:</b>", label_style),
            Paragraph(safe_str(dn.get("vehicle_no")), value_style),
            Paragraph("<b>Tracking ID:</b>", label_style),
            Paragraph(safe_str(dn.get("tracking_id")), value_style),
        ],
        [
            Paragraph("<b>Destination Address:</b>", label_style),
            Paragraph(safe_str(dn.get("destination_address")), value_style),
            Paragraph("<b>Status:</b>", label_style),
            Paragraph(status_text, value_style),
        ],
    ]

    details_table = Table(dn_details_data, colWidths=[110, 170, 95, 145])
    details_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f3f3")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#8a8a8a")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#a5a5a5")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(details_table)
    elements.append(Spacer(1, 16))

    # ---------------------------------------------------
    # DELIVERY NOTE ITEMS
    # ---------------------------------------------------
    elements.append(Paragraph("DELIVERY NOTE ITEMS", section_style))
    elements.append(Spacer(1, 2))

    items = dn.get("items", []) or []

    item_data = [[
        Paragraph("S.No", header_small_style),
        Paragraph("Product Name", header_small_style),
        Paragraph("Product ID", header_small_style),
        Paragraph("Qty", header_small_style),
        Paragraph("UOM", header_small_style),
        Paragraph("Rate", header_small_style),
        Paragraph("Tax %", header_small_style),
        Paragraph("Disc %", header_small_style),
        Paragraph("Total", header_small_style),
    ]]

    grand_total = 0.0
    subtotal_sum = 0.0
    total_tax_sum = 0.0
    total_discount_sum = 0.0

    for idx, item in enumerate(items, start=1):
        product_name = safe_str(item.get("product_name"))
        product_id = safe_str(item.get("product_id"))
        qty = safe_float(item.get("qty"), 0.0)
        uom = safe_str(item.get("uom"))
        rate = safe_float(item.get("rate"), 0.0)
        tax = safe_float(item.get("tax"), 0.0)
        discount = safe_float(item.get("discount"), 0.0)

        line_subtotal = rate * qty
        line_tax = (line_subtotal * tax) / 100
        line_discount = (line_subtotal * discount) / 100

        total = item.get("total", None)
        if total in (None, ""):
            total = line_subtotal + line_tax - line_discount
        total = safe_float(total, 0.0)

        subtotal_sum += line_subtotal
        total_tax_sum += line_tax
        total_discount_sum += line_discount
        grand_total += total

        item_data.append([
            Paragraph(str(idx), value_style),
            Paragraph(product_name, value_style),
            Paragraph(product_id, value_style),
            Paragraph(f"{qty:.2f}".rstrip("0").rstrip("."), value_style),
            Paragraph(uom, value_style),
            Paragraph(f"₹{rate:.2f}", value_style),
            Paragraph(f"{tax:.1f}%", value_style),
            Paragraph(f"{discount:.1f}%", value_style),
            Paragraph(f"₹{total:.2f}", value_style),
        ])

    if len(item_data) == 1:
        item_data.append(["-", "No line items available", "-", "-", "-", "-", "-", "-", "-"])

    items_table = Table(
        item_data,
        colWidths=[35, 135, 72, 42, 45, 58, 45, 45, 60],
        repeatRows=1
    )
    items_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#a12828")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "DejaVuSans"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (2, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#999999")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f7f7")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 16))

    # ---------------------------------------------------
    # TAX AND TOTAL SUMMARY
    # ---------------------------------------------------
    elements.append(Paragraph("TAX AND TOTALS SUMMARY", section_style))
    elements.append(Spacer(1, 3))

    shipping_charge = safe_float(dn.get("shipping_charge"), 0.0)
    global_discount = safe_float(dn.get("global_discount"), 0.0)
    final_total = grand_total + shipping_charge - global_discount

    summary_data = [
        [Paragraph("Subtotal:", value_style), Paragraph(f"₹{subtotal_sum:.2f}", value_style)],
        [Paragraph("Total Discount (Item Level):", value_style), Paragraph(f"₹{total_discount_sum:.2f}", value_style)],
        [Paragraph("Total Tax:", value_style), Paragraph(f"₹{total_tax_sum:.2f}", value_style)],
        [Paragraph("Shipping Charge:", value_style), Paragraph(f"₹{shipping_charge:.2f}", value_style)],
        [Paragraph("Global Discount:", value_style), Paragraph(f"₹{global_discount:.2f}", value_style)],
        [Paragraph("GRAND TOTAL:", summary_white_style), Paragraph(f"₹{final_total:.2f}", summary_white_style)],
    ]

    summary_table = Table(summary_data, colWidths=[380, 120])
    summary_table.setStyle(TableStyle([
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (-1, -2), "DejaVuSans"),
        ("FONTNAME", (0, -1), (-1, -1), "DejaVuSans-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#a12828")),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    # ---------------------------------------------------
    # DELIVERY NOTES
    # ---------------------------------------------------
    delivery_notes = safe_str(dn.get("delivery_notes"), "").strip()
    if delivery_notes:
        elements.append(Paragraph("Delivery Notes", section_style))
        notes_table = Table([[Paragraph(delivery_notes, value_style)]], colWidths=[500])
        notes_table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#999999")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fafafa")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(notes_table)
        elements.append(Spacer(1, 12))

    # ---------------------------------------------------
    # TERMS AND CONDITIONS
    # ---------------------------------------------------
    elements.append(Paragraph("Terms and Conditions", section_style))

    terms_list = [
        "1. This Delivery Note is issued based on the confirmed order details.",
        "2. Delivery will be made as per the agreed schedule.",
        "3. Kindly verify the delivered items at the time of receipt.",
        "4. Any shortage or damage should be reported immediately.",
        "5. Goods once delivered will be considered accepted unless otherwise notified.",
    ]

    for term in terms_list:
        elements.append(Paragraph(term, terms_style))

    # ---------------------------------------------------
    # BUILD PDF
    # ---------------------------------------------------
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

# =======================================
# Stock Reciept
# =========================================
 
@app.route('/stock-receipt')
def stock_receipt():
    data = [
        {
            "grn": "GRN-0001",
            "po": "PO-0001",
            "supplier": "Vasu",
            "date": "10-01-2026",
            "total": 20,
            "status": "Draft",
            "received_by": "Mandy",
            "qc_by": "Sans"
        },
        {
            "grn": "GRN-0002",
            "po": "PO-0002",
            "supplier": "Srinu",
            "date": "10-01-2026",
            "total": 20,
            "status": "Submitted",
            "received_by": "Mandy",
            "qc_by": "Sans"
        }
    ]
 
    return render_template(
        'stock-reciept.html',   # ✅ EXACT match
        data=data,
        page='stock_receipt'
    )

# =========================================
# ✅ RUN APP
# =========================================
if __name__ == "__main__":
    print("Application is running successfully")
    app.run(debug=True)